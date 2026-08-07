#!/usr/bin/env python3
"""
生成符合客户格式的完整ESG排名表格
严格按照客户2024年报告的表格格式
"""

import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime

print("=" * 80)
print("生成客户格式ESG排名表格（全部605家企业）")
print("=" * 80)

# 读取最新评分
companies = []
with open('output/audit/esg_scores_v1_2025.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        companies.append({
            'rank': int(row['rank']),
            'code': row['company_code'],
            'name': row['company_name'],
            'esg_score': float(row['esg_score']),
        })

# 读取指标数据
indicator_data = {}
with open('output/audit/ci_merged_all_sources_v1_2025.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        code = row['company_code']
        if code not in indicator_data:
            indicator_data[code] = {}
        indicator_data[code][row['indicator_code']] = float(row['value'])

print(f"\n已加载 {len(companies)} 家企业数据")

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# 设置列宽
ws.column_dimensions['A'].width = 6   # 序号
ws.column_dimensions['B'].width = 2   # 空列
ws.column_dimensions['C'].width = 2   # 空列
ws.column_dimensions['D'].width = 10  # 联系人
ws.column_dimensions['E'].width = 12  # 公司地址
ws.column_dimensions['F'].width = 12  # 证券代码
ws.column_dimensions['G'].width = 15  # 公司简称
ws.column_dimensions['H'].width = 25  # 公司名称
ws.column_dimensions['I'].width = 10  # 指标类型
for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
    ws.column_dimensions[col].width = 12

# 第一行：标题
ws.merge_cells('A1:T1')
title_cell = ws['A1']
title_cell.value = '2026中国能源上市公司可持续发展（ESG）评价'
title_cell.font = Font(name='Microsoft YaHei', size=14, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')
title_cell.fill = PatternFill(start_color='E6F4FF', end_color='E6F4FF', fill_type='solid')

# 第二行：列标题
headers = [
    '序号', None, None, '证券代码', '公司地址', '证券代码', '公司简称', '公司名称', None,
    '温室气体排放强度（千克/万元）', '综合能源消耗强度（千克/万元）',
    '氮氧化物（NOX）排放强度（克/万元营业收入）', '二氧化硫（SO2）排放强度（克/万元营业收入）',
    '新鲜水资源消耗强度（千克/万元营业收入）', '一般固体废弃物排放强度（千克/万元营业收入）',
    '环保/安全生产投入占比（％）', '研发（R&D）费用占比（%）', '现金分红（元/股）',
    '资产负债率（%）', 'ESG分数'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.fill = PatternFill(start_color='D9E9FF', end_color='D9E9FF', fill_type='solid')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

# 指标代码映射
indicator_codes = {
    'Q_E_GHG_INTENSITY': 'J',
    'Q_E_ENERGY_INTENSITY': 'K',
    'Q_E_NOX_INTENSITY': 'L',
    'Q_E_SO2_INTENSITY': 'M',
    'Q_E_WATER_INTENSITY': 'N',
    'Q_E_SOLID_WASTE_INTENSITY': 'O',
    'Q_S_SAFETY_INVEST_RATE': 'P',
    'Q_S_RD_RATE': 'Q',
    'Q_G_DIVIDEND_PER_SHARE': 'R',
    'Q_G_DEBT_RATIO': 'S',
}

# 填充数据
current_row = 3
for company in companies:
    code = company['code']

    # 指标数值行
    ws.cell(row=current_row, column=1).value = company['rank']  # 序号
    ws.cell(row=current_row, column=4).value = '李天平'  # 默认联系人
    ws.cell(row=current_row, column=5).value = ''  # 公司地址（暂无）
    ws.cell(row=current_row, column=6).value = code  # 证券代码
    ws.cell(row=current_row, column=7).value = company['name']  # 公司简称
    ws.cell(row=current_row, column=8).value = company['name']  # 公司名称（同简称）
    ws.cell(row=current_row, column=9).value = '指标数值'

    # 填充指标数值
    company_indicators = indicator_data.get(code, {})
    for ind_code, col_letter in indicator_codes.items():
        value = company_indicators.get(ind_code)
        cell = ws[f'{col_letter}{current_row}']
        if value is not None:
            # 格式化数值
            if 'RATE' in ind_code or 'RATIO' in ind_code:
                cell.value = round(value, 2)
            elif 'DIVIDEND' in ind_code:
                cell.value = round(value, 2)
            else:
                cell.value = round(value, 2)
            cell.number_format = '0.00'
        else:
            cell.value = 0
            cell.number_format = '0.00'
        cell.alignment = Alignment(horizontal='right')

    # ESG分数
    ws.cell(row=current_row, column=20).value = round(company['esg_score'], 2)
    ws.cell(row=current_row, column=20).number_format = '0.00'
    ws.cell(row=current_row, column=20).alignment = Alignment(horizontal='right')

    # 设置行样式
    for col in range(1, 21):
        cell = ws.cell(row=current_row, column=col)
        cell.font = Font(name='Microsoft YaHei', size=9)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        if col == 1:  # 序号列
            cell.alignment = Alignment(horizontal='center')
        elif col == 20:  # ESG分数列
            cell.fill = PatternFill(start_color='FFF4E6', end_color='FFF4E6', fill_type='solid')
            cell.font = Font(name='Microsoft YaHei', size=9, bold=True, color='D2691E')

    current_row += 1

    # 每100行打印进度
    if company['rank'] % 100 == 0:
        print(f"  已处理 {company['rank']} 家企业...")

# 保存文件
output_file = 'output/audit/2026年能源企业ESG评价完整排名_605家.xlsx'
wb.save(output_file)

print(f"\n已保存: {output_file}")
print(f"  企业总数: {len(companies)}")
print(f"  总行数: {current_row - 1} (含2行标题)")
print(f"  文件大小: {len(companies) * 0.015:.1f} KB (预估)")

print("\n" + "=" * 80)
print("✅ 客户格式ESG排名表格生成完成")
print("=" * 80)
