#!/usr/bin/env python3
"""
严格按照客户2024年格式生成ESG排名表格
每个企业2行：指标数值 + 指标分值
"""

import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

print("=" * 80)
print("生成客户格式ESG排名表格（严格按照2024年格式）")
print("=" * 80)

# 读取最新评分和行业分类
companies = []
with open('output/audit/esg_scores_with_industry_v1_2025.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        companies.append({
            'rank': int(row['overall_rank']),
            'code': row['company_code'],
            'name': row['company_name'],
            'industry': row['industry'],
            'esg_score': float(row['esg_score']),
        })

# 读取指标数据和评分
indicator_data = {}
with open('output/audit/ci_merged_all_sources_v1_2025.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        code = row['company_code']
        if code not in indicator_data:
            indicator_data[code] = {}
        indicator_data[code][row['indicator_code']] = {
            'value': float(row['value']),
            'source': row['source']
        }

# 读取方法论用于计算单项分值
import json
with open('data/methodologies/energy_esg_2025_research_sasac.json', 'r', encoding='utf-8') as f:
    methodology = json.load(f)

# 建立指标映射
indicator_map = {ind['code']: ind for ind in methodology['indicators']}

# 读取基准统计数据用于计算分值
indicator_stats = {}
with open('output/audit/ci_incremental_review_summary_v1_2025.csv', 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    for row in reader:
        indicator_stats[row['indicator_code']] = {
            'mean': float(row['mean']),
            'std': float(row['std']),
            'min': float(row['min']),
            'max': float(row['max'])
        }

def calculate_indicator_score(value, indicator_code):
    """计算单个指标的得分（0-100标准化后乘以权重）"""
    if indicator_code not in indicator_map:
        return 0.0

    ind_meta = indicator_map[indicator_code]
    stats = indicator_stats.get(indicator_code)

    if not stats or value is None:
        return 0.0

    # 标准化到0-100
    prefer_lower = ind_meta.get('prefer_lower', False)

    if prefer_lower:
        # 越低越好（如排放强度）
        if stats['max'] == stats['min']:
            normalized = 50.0
        else:
            normalized = 100.0 * (1 - (value - stats['min']) / (stats['max'] - stats['min']))
    else:
        # 越高越好（如研发投入）
        if stats['max'] == stats['min']:
            normalized = 50.0
        else:
            normalized = 100.0 * (value - stats['min']) / (stats['max'] - stats['min'])

    normalized = max(0, min(100, normalized))

    # 乘以权重得到实际得分
    weight = ind_meta['weight']
    return normalized * weight / 100.0

print(f"\n已加载 {len(companies)} 家企业数据")

# 创建Excel工作簿
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sheet1"

# 设置列宽
ws.column_dimensions['A'].width = 6
ws.column_dimensions['D'].width = 10
ws.column_dimensions['E'].width = 12
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 15
ws.column_dimensions['H'].width = 25
ws.column_dimensions['I'].width = 10
for col in ['J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
    ws.column_dimensions[col].width = 12

# 第一行：标题
ws.merge_cells('A1:T1')
title_cell = ws['A1']
title_cell.value = '2026中国能源上市公司可持续发展（ESG）'
title_cell.font = Font(name='Microsoft YaHei', size=14, bold=True)
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# 第二行：列标题
headers = [
    '序号', None, None, '证券代码', '公司地址', '证券代码', '公司简称', '公司名称', None,
    '温室气体排放强度（千克/万元）', '综合能源消耗强度（千克/万元）',
    '氮氧化物（NOX）排放强度（克/万元营业收入）', '二氧化硫（SO2）排放强度（克/万元营业收入）',
    '新鲜水资源消耗强度（千克/万元营业收入）', '一般固体废弃物排放强度（千克/万元营业收入）',
    '环保/安全生产投入占比（％）', '研发（R&D）费用占比（%）', '现金分红（元/股）',
    '资产负债率（%）', 'ESG分数'
]

for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col_idx)
    cell.value = header
    cell.font = Font(name='Microsoft YaHei', size=9, bold=True)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

# 10项关键指标代码
indicator_codes = [
    'Q_E_GHG_INTENSITY',
    'Q_E_ENERGY_INTENSITY',
    'Q_E_NOX_INTENSITY',
    'Q_E_SO2_INTENSITY',
    'Q_E_WATER_INTENSITY',
    'Q_E_SOLID_WASTE_INTENSITY',
    'Q_S_SAFETY_INVEST_RATE',
    'Q_S_RD_RATE',
    'Q_G_DIVIDEND_PER_SHARE',
    'Q_G_DEBT_RATIO',
]

# 填充数据
current_row = 3
for company in companies:
    code = company['code']
    company_indicators = indicator_data.get(code, {})

    # ===== 第1行：指标数值 =====
    ws.cell(row=current_row, column=1).value = company['rank']  # 序号
    ws.cell(row=current_row, column=4).value = '李天平'  # 联系人
    ws.cell(row=current_row, column=5).value = company['industry']  # 公司地址用行业代替
    ws.cell(row=current_row, column=6).value = code
    ws.cell(row=current_row, column=7).value = company['name']
    ws.cell(row=current_row, column=8).value = company['name']
    ws.cell(row=current_row, column=9).value = '指标数值'

    # 填充10项指标数值
    for col_offset, ind_code in enumerate(indicator_codes, 0):
        col = 10 + col_offset
        ind_info = company_indicators.get(ind_code)
        if ind_info:
            ws.cell(row=current_row, column=col).value = round(ind_info['value'], 2)
        else:
            ws.cell(row=current_row, column=col).value = 0

    # ESG总分
    ws.cell(row=current_row, column=20).value = round(company['esg_score'], 2)

    # ===== 第2行：指标分值 =====
    current_row += 1
    ws.cell(row=current_row, column=9).value = '指标分值'

    # 填充10项指标分值
    for col_offset, ind_code in enumerate(indicator_codes, 0):
        col = 10 + col_offset
        ind_info = company_indicators.get(ind_code)
        if ind_info:
            score = calculate_indicator_score(ind_info['value'], ind_code)
            ws.cell(row=current_row, column=col).value = round(score, 2)
        else:
            ws.cell(row=current_row, column=col).value = 0

    # 分值行ESG总分留空
    ws.cell(row=current_row, column=20).value = None

    current_row += 1

    # 每100家打印进度
    if company['rank'] % 100 == 0:
        print(f"  已处理 {company['rank']} 家企业...")

# 保存文件
output_file = 'output/audit/2026年能源企业ESG评价完整排名_客户格式_605家.xlsx'
wb.save(output_file)

print(f"\n已保存: {output_file}")
print(f"  企业总数: {len(companies)}")
print(f"  总行数: {current_row - 1} (含2行标题 + {len(companies) * 2}行数据)")
print(f"  每个企业2行: 指标数值 + 指标分值")

print("\n" + "=" * 80)
print("✅ 客户格式ESG排名表格生成完成")
print("=" * 80)
