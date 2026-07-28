from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Iterable

from .universe import UniverseCompany


INDICATORS = (
    ("Q_E_GHG_INTENSITY", "温室气体排放强度"),
    ("Q_E_ENERGY_INTENSITY", "综合能源消耗强度"),
    ("Q_E_NOX_INTENSITY", "NOx"),
    ("Q_E_SO2_INTENSITY", "SO2"),
    ("Q_E_WATER_INTENSITY", "新鲜水资源消耗强度"),
    ("Q_E_GENERAL_WASTE_INTENSITY", "一般固废排放强度"),
    # 历史表将环保和安全投入合并，不能无依据映射为现行单项指标。
    ("H_ENV_SAFETY_INVEST_RATE", "环保/安全生产投入占比"),
    ("Q_S_RD_RATE", "研发费用占比"),
    ("Q_S_DIVIDEND_PER_SHARE", "现金分红"),
    ("Q_G_DEBT_ASSET_RATE", "资产负债率"),
)


@dataclass(frozen=True)
class HistoricalCompany:
    evaluation_year: int
    report_year: int
    historical_rank: str
    stock_code: str
    company_abbr: str
    company_name: str
    contact_name: str
    company_city: str
    exchange: str
    st_flag: bool
    current_snapshot_status: str
    historical_esg_score: str
    source_file: str
    source_value_row: int
    source_score_row: int


@dataclass(frozen=True)
class HistoricalObservation:
    evaluation_year: int
    report_year: int
    stock_code: str
    company_name: str
    indicator_code: str
    source_indicator_name: str
    raw_value: str
    historical_score: str
    source_file: str
    source_value_row: int
    source_score_row: int


def import_historical_workbook(
    path: str | Path,
    snapshots: Iterable[UniverseCompany],
    evaluation_year: int,
    report_year: int,
) -> tuple[list[HistoricalCompany], list[HistoricalObservation], dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - depends on optional installation
        raise RuntimeError("导入XLSX需要安装 openpyxl（pip install -e '.[universe]'）") from exc

    source = Path(path)
    workbook = load_workbook(source, read_only=True, data_only=True)
    if len(workbook.sheetnames) != 1:
        raise ValueError(f"历史评价工作簿必须只有一个工作表，实际为{len(workbook.sheetnames)}个")
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    header_index = _find_header(rows)
    snapshot_codes = {item.stock_code.upper() for item in snapshots}
    companies: list[HistoricalCompany] = []
    observations: list[HistoricalObservation] = []
    seen: set[str] = set()

    index = header_index + 1
    while index < len(rows):
        value_row = rows[index]
        if _text(_cell(value_row, 8)) != "指标数值":
            if any(_text(value) for value in value_row):
                raise ValueError(f"第{index + 1}行应为指标数值行")
            index += 1
            continue
        if index + 1 >= len(rows) or _text(_cell(rows[index + 1], 8)) != "指标分值":
            raise ValueError(f"第{index + 1}行缺少紧随其后的指标分值行")
        score_row = rows[index + 1]
        code = _normalize_code(_cell(value_row, 5))
        name = _text(_cell(value_row, 7))
        identity = code if code and code != "#N/A" else f"NAME:{name}"
        if not name:
            raise ValueError(f"第{index + 1}行公司名称为空")
        if identity in seen:
            raise ValueError(f"第{index + 1}行公司记录重复: {identity}")
        seen.add(identity)
        exchange = _exchange(code)
        snapshot_status = "not_applicable" if exchange not in ("SSE", "SZSE") else (
            "matched" if code in snapshot_codes else "unmatched"
        )
        companies.append(HistoricalCompany(
            evaluation_year, report_year, _text(_cell(value_row, 0)), code,
            _text(_cell(value_row, 6)), name, _text(_cell(value_row, 3)),
            _text(_cell(value_row, 4)), exchange,
            bool(re.match(r"^\*?ST", _text(_cell(value_row, 6)), re.IGNORECASE)),
            snapshot_status, _text(_cell(value_row, 19)), source.name,
            index + 1, index + 2,
        ))
        for offset, (indicator_code, label) in enumerate(INDICATORS, 9):
            observations.append(HistoricalObservation(
                evaluation_year, report_year, code, name, indicator_code, label,
                _text(_cell(value_row, offset)), _text(_cell(score_row, offset)),
                source.name, index + 1, index + 2,
            ))
        index += 2
    workbook.close()

    exchanges = Counter(item.exchange for item in companies)
    nonmissing = Counter(
        item.indicator_code for item in observations if item.raw_value not in ("", "#N/A")
    )
    audit = {
        "source_file": source.name,
        "source_sha256": hashlib.sha256(source.read_bytes()).hexdigest(),
        "sheet_name": sheet.title,
        "evaluation_year": evaluation_year,
        "report_year": report_year,
        "company_count": len(companies),
        "observation_count": len(observations),
        "exchange_counts": dict(sorted(exchanges.items())),
        "st_company_count": sum(item.st_flag for item in companies),
        "malformed_code_count": sum(item.exchange == "UNKNOWN" for item in companies),
        "mainland_snapshot_matched": sum(item.current_snapshot_status == "matched" for item in companies),
        "mainland_snapshot_unmatched": sum(item.current_snapshot_status == "unmatched" for item in companies),
        "nonmissing_observations": dict(sorted(nonmissing.items())),
    }
    return companies, observations, audit


def write_historical_import(
    companies_path: str | Path, observations_path: str | Path, audit_path: str | Path,
    companies: Iterable[HistoricalCompany], observations: Iterable[HistoricalObservation], audit: dict,
) -> None:
    _write_dataclasses(companies_path, companies, tuple(HistoricalCompany.__annotations__))
    _write_dataclasses(observations_path, observations, tuple(HistoricalObservation.__annotations__))
    output = Path(audit_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(audit, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def _write_dataclasses(path, rows, fields):
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(asdict(item) for item in rows)


def _find_header(rows) -> int:
    for index, row in enumerate(rows):
        texts = {_text(value) for value in row}
        if "公司简称" in texts and "公司名称" in texts and "ESG分数" in texts:
            return index
    raise ValueError("未找到包含公司简称、公司名称和ESG分数的表头")


def _cell(row, index):
    return row[index] if index < len(row) else None


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _normalize_code(value) -> str:
    text = _text(value).upper().replace(" ", "")
    if text in ("", "#N/A"):
        return text or "#N/A"
    if re.fullmatch(r"\d{6}\.(SH|SZ|BJ)", text) or re.fullmatch(r"\d{5}\.HK", text):
        return text
    return text


def _exchange(code: str) -> str:
    suffixes = {".SH": "SSE", ".SZ": "SZSE", ".BJ": "BSE", ".HK": "HKEX"}
    return next((exchange for suffix, exchange in suffixes.items() if code.endswith(suffix)), "UNKNOWN")
