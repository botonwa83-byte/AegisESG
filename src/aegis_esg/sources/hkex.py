from __future__ import annotations

import re
from datetime import date
from pathlib import Path

from ..universe_builder import ExchangeSecurity, normalize_stock_code


EQUITY_SUBCATEGORIES = {
    "Equity Securities (Main Board)",
    "Equity Securities (GEM)",
}


def import_hkex_securities(
    input_path: str | Path, source_url: str, expected_as_of_date: str = "",
) -> tuple[list[ExchangeSecurity], str]:
    """Import ordinary listed equities from HKEX's Full List of Securities.

    The official workbook contains two metadata rows before its header and also
    contains funds, debt and derivatives.  Only Main Board and GEM equity
    securities represent listed-company share lines for the universe snapshot.
    """
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError("读取港交所XLSX需要安装openpyxl") from error

    source = Path(input_path)
    sheet = load_workbook(source, read_only=True, data_only=True).active
    # HKEX workbooks currently carry a stale A1-style worksheet dimension.
    sheet.reset_dimensions()
    values = sheet.iter_rows(values_only=True)
    try:
        title = _text(next(values)[0])
        updated = _text(next(values)[0])
        headers = [_text(value) for value in next(values)]
    except StopIteration as error:
        raise ValueError("港交所证券名单为空或缺少元数据行") from error
    if title != "List of Securities":
        raise ValueError(f"无法识别港交所证券名单标题: {title or '空'}")
    match = re.fullmatch(r"Updated as at (\d{2})/(\d{2})/(\d{4})", updated)
    if not match:
        raise ValueError(f"无法识别港交所名单更新日期: {updated or '空'}")
    day, month, year = map(int, match.groups())
    as_of_date = date(year, month, day).isoformat()
    if expected_as_of_date and expected_as_of_date != as_of_date:
        raise ValueError(f"港交所名单日期不一致: 期望{expected_as_of_date}，文件为{as_of_date}")

    required = {"Stock Code", "Name of Securities", "Category", "Sub-Category"}
    missing = required.difference(headers)
    if missing:
        raise ValueError(f"港交所证券名单缺少字段: {','.join(sorted(missing))}")
    positions = {name: headers.index(name) for name in required}
    rows: list[ExchangeSecurity] = []
    for row_number, row in enumerate(values, 4):
        category = _cell(row, positions["Category"])
        subcategory = _cell(row, positions["Sub-Category"])
        if category != "Equity" or subcategory not in EQUITY_SUBCATEGORIES:
            continue
        raw_code = _cell(row, positions["Stock Code"])
        name = _cell(row, positions["Name of Securities"])
        if not raw_code or not name:
            raise ValueError(f"港交所证券名单第{row_number}行缺少代码或名称")
        code = normalize_stock_code(raw_code, "HKEX")
        rows.append(ExchangeSecurity(
            stock_code=code, company_name=name, exchange="HKEX",
            industry="待分类", entity_id=code, listing_status="上市",
            source_url=source_url, as_of_date=as_of_date,
        ))
    if not rows:
        raise ValueError("港交所证券名单没有主板或GEM普通股记录")
    if len({item.stock_code for item in rows}) != len(rows):
        raise ValueError("港交所普通股名单包含重复证券代码")
    return rows, as_of_date


def _cell(row: tuple, index: int) -> str:
    return _text(row[index] if index < len(row) else None)


def _text(value) -> str:
    return str(value or "").strip()
