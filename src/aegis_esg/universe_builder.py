from __future__ import annotations

import csv
import json
import re
from datetime import date
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable


ENERGY_INDUSTRIES = {
    "煤炭", "石油石化", "石油天然气", "电力", "燃气", "新能源",
    "能源设备", "电气设备", "油气开采", "油气服务",
    "火力发电", "水力发电", "核力发电", "风力发电", "光伏发电",
}
ACTIVE_STATUSES = {"", "上市", "正常", "active", "listed"}
EXCHANGES = {"SSE", "SZSE", "BSE", "HKEX"}
FIELD_ALIASES = {
    "stock_code": ("stock_code", "证券代码", "股票代码", "公司代码", "A股代码", "股份代号", "stock code", "code"),
    "company_name": ("company_name", "公司名称", "公司简称", "证券简称", "股票简称", "A股简称", "股份简称", "name"),
    "industry": ("industry", "行业", "行业名称", "所属行业", "industry name"),
    "entity_id": ("entity_id", "主体标识", "统一社会信用代码", "issuer id"),
    "st_status": ("st_status", "特别处理状态", "ST状态"),
    "listing_status": ("listing_status", "上市状态", "状态", "status"),
    "energy_eligible": ("energy_eligible", "能源行业纳入", "是否能源行业"),
}


@dataclass(frozen=True)
class ExchangeSecurity:
    stock_code: str
    company_name: str
    exchange: str
    industry: str
    entity_id: str
    st_status: str = ""
    listing_status: str = "上市"
    energy_eligible: str = ""
    source_url: str = ""
    as_of_date: str = ""


@dataclass(frozen=True)
class UniverseDecision:
    stock_code: str
    company_name: str
    exchange: str
    sub_industry: str
    included: bool
    exclusion_reason: str
    entity_id: str
    source_url: str
    as_of_date: str


@dataclass(frozen=True)
class SnapshotQuality:
    row_count: int
    exchanges: dict[str, int]
    missing_source_count: int
    missing_date_count: int
    missing_entity_count: int
    unknown_exchange_count: int
    duplicate_code_count: int
    invalid_source_count: int
    invalid_date_count: int
    valid: bool

    def as_dict(self) -> dict:
        return vars(self)


def normalize_exchange_export(
    input_path: str | Path, exchange: str, source_url: str, as_of_date: str,
) -> list[ExchangeSecurity]:
    exchange = exchange.strip().upper()
    if exchange not in EXCHANGES:
        raise ValueError(f"不支持的交易所: {exchange}")
    rows = _read_tabular_rows(input_path)
    if not rows:
        return []
    aliases = {_normalize_header(key): key for key in rows[0]}
    mapping = {}
    for target, candidates in FIELD_ALIASES.items():
        match = next((aliases[_normalize_header(name)] for name in candidates if _normalize_header(name) in aliases), None)
        if match:
            mapping[target] = match
    missing = {"stock_code", "company_name"}.difference(mapping)
    if missing:
        raise ValueError(f"交易所原始文件缺少可识别字段: {','.join(sorted(missing))}")
    result = []
    for line, row in enumerate(rows, 2):
        raw_code = (row.get(mapping["stock_code"]) or "").strip()
        name = (row.get(mapping["company_name"]) or "").strip()
        if not raw_code or not name:
            raise ValueError(f"交易所原始文件第{line}行缺少证券代码或公司名称")
        values = {key: (row.get(column) or "").strip() for key, column in mapping.items()}
        code = normalize_stock_code(raw_code, exchange)
        result.append(ExchangeSecurity(
            stock_code=code, company_name=name, exchange=exchange,
            industry=values.get("industry", "待分类"),
            entity_id=values.get("entity_id") or code,
            st_status=values.get("st_status", ""), listing_status=values.get("listing_status") or "上市",
            energy_eligible=values.get("energy_eligible", "").lower(),
            source_url=source_url.strip(), as_of_date=as_of_date.strip(),
        ))
    return result


def normalize_stock_code(raw_code: str, exchange: str) -> str:
    code = raw_code.strip().upper()
    code = re.sub(r"\.(SH|SS|SZ|BJ|HK)$", "", code)
    if not code.isdigit():
        raise ValueError(f"证券代码格式错误: {raw_code}")
    exchange = exchange.upper()
    width = 5 if exchange == "HKEX" else 6
    suffix = {"SSE": "SH", "SZSE": "SZ", "BSE": "BJ", "HKEX": "HK"}[exchange]
    return f"{code.zfill(width)}.{suffix}"


def audit_snapshot(securities: Iterable[ExchangeSecurity]) -> SnapshotQuality:
    rows = list(securities)
    counts: dict[str, int] = {}
    for item in rows:
        counts[item.exchange] = counts.get(item.exchange, 0) + 1
    missing_source = sum(not item.source_url for item in rows)
    missing_date = sum(not item.as_of_date for item in rows)
    missing_entity = sum(not item.entity_id for item in rows)
    unknown_exchange = sum(item.exchange not in EXCHANGES for item in rows)
    codes = [item.stock_code for item in rows]
    duplicate_codes = len(codes) - len(set(codes))
    invalid_source = sum(bool(item.source_url) and not re.match(r"^https?://", item.source_url, re.I) for item in rows)
    invalid_date = sum(bool(item.as_of_date) and not _is_iso_date(item.as_of_date) for item in rows)
    return SnapshotQuality(
        len(rows), dict(sorted(counts.items())), missing_source, missing_date,
        missing_entity, unknown_exchange, duplicate_codes, invalid_source, invalid_date,
        bool(rows) and not any((missing_source, missing_date, missing_entity, unknown_exchange,
                                duplicate_codes, invalid_source, invalid_date)),
    )


def write_exchange_snapshot(path: str | Path, securities: Iterable[ExchangeSecurity]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = tuple(ExchangeSecurity.__annotations__)
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        writer.writerows(vars(item) for item in securities)


def write_snapshot_quality(path: str | Path, quality: SnapshotQuality) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(quality.as_dict(), ensure_ascii=False, indent=2), encoding="utf-8")


def read_exchange_snapshot(path: str | Path) -> list[ExchangeSecurity]:
    with Path(path).open(encoding="utf-8-sig", newline="") as stream:
        rows = list(csv.DictReader(stream))
    required = {"stock_code", "company_name", "exchange", "industry"}
    missing = required.difference(rows[0] if rows else ())
    if missing:
        raise ValueError(f"交易所快照缺少字段: {','.join(sorted(missing))}")
    result = []
    for line, row in enumerate(rows, 2):
        code = row["stock_code"].strip().upper()
        name = row["company_name"].strip()
        if not code or not name:
            raise ValueError(f"交易所快照第{line}行缺少证券代码或公司名称")
        result.append(ExchangeSecurity(
            stock_code=code, company_name=name, exchange=row["exchange"].strip().upper(),
            industry=row["industry"].strip(), entity_id=(row.get("entity_id") or code).strip().upper(),
            st_status=(row.get("st_status") or "").strip(),
            listing_status=(row.get("listing_status") or "上市").strip(),
            energy_eligible=(row.get("energy_eligible") or "").strip().lower(),
            source_url=(row.get("source_url") or "").strip(),
            as_of_date=(row.get("as_of_date") or "").strip(),
        ))
    return result


def build_energy_universe(securities: Iterable[ExchangeSecurity]) -> list[UniverseDecision]:
    rows = list(securities)
    seen_codes: set[str] = set()
    eligible = [item for item in rows if not _exclusion_reason(item)]
    primary_by_entity: dict[str, str] = {}
    for item in eligible:
        current_code = primary_by_entity.get(item.entity_id)
        if current_code is None or _security_priority(item) < _code_priority(current_code):
            primary_by_entity[item.entity_id] = item.stock_code
    decisions = []
    for item in rows:
        if item.stock_code in seen_codes:
            raise ValueError(f"交易所快照证券代码重复: {item.stock_code}")
        seen_codes.add(item.stock_code)
        reason = _exclusion_reason(item)
        included = not reason
        if included and primary_by_entity[item.entity_id] != item.stock_code:
            included = False
            reason = f"同一主体重复上市，保留{primary_by_entity[item.entity_id]}"
        decisions.append(UniverseDecision(
            item.stock_code, item.company_name, item.exchange, item.industry,
            included, reason, item.entity_id, item.source_url, item.as_of_date,
        ))
    return decisions


def write_universe(path: str | Path, decisions: Iterable[UniverseDecision]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = (
        "stock_code", "company_name", "exchange", "sub_industry", "included",
        "exclusion_reason", "entity_id", "source_url", "as_of_date",
    )
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        for item in decisions:
            row = vars(item).copy()
            row["included"] = "true" if item.included else "false"
            writer.writerow(row)


def write_decision_audit(path: str | Path, decisions: Iterable[UniverseDecision]) -> None:
    output = Path(path)
    output.parent.mkdir(parents=True, exist_ok=True)
    fields = ("stock_code", "company_name", "exchange", "industry", "decision", "reason", "entity_id", "source_url", "as_of_date")
    with output.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields)
        writer.writeheader()
        for item in decisions:
            writer.writerow({
                "stock_code": item.stock_code, "company_name": item.company_name,
                "exchange": item.exchange, "industry": item.sub_industry,
                "decision": "include" if item.included else "exclude",
                "reason": item.exclusion_reason or "能源行业映射命中",
                "entity_id": item.entity_id, "source_url": item.source_url,
                "as_of_date": item.as_of_date,
            })


def _exclusion_reason(item: ExchangeSecurity) -> str:
    status = item.listing_status.strip().lower()
    if status not in ACTIVE_STATUSES:
        return f"非正常上市状态:{item.listing_status}"
    st = item.st_status.upper().replace(" ", "")
    name = item.company_name.upper().replace(" ", "")
    if st in {"ST", "*ST"} or name.startswith("ST") or name.startswith("*ST"):
        return "ST/*ST排除"
    if item.energy_eligible in {"false", "0", "no", "否"}:
        return "行业映射明确排除"
    if item.energy_eligible in {"true", "1", "yes", "是"}:
        return ""
    if item.industry not in ENERGY_INDUSTRIES:
        return f"行业待复核:{item.industry or '未分类'}"
    return ""


def _security_priority(item: ExchangeSecurity) -> tuple[int, str]:
    exchange_order = {"SSE": 0, "SZSE": 1, "BSE": 2, "HKEX": 3}
    return exchange_order.get(item.exchange, 9), item.stock_code


def _code_priority(code: str) -> tuple[int, str]:
    if code.endswith(".SH"):
        exchange = "SSE"
    elif code.endswith(".SZ"):
        exchange = "SZSE"
    elif code.endswith(".BJ"):
        exchange = "BSE"
    elif code.endswith(".HK"):
        exchange = "HKEX"
    else:
        exchange = "UNKNOWN"
    return _security_priority(ExchangeSecurity(code, "", exchange, "", ""))


def _normalize_header(value: str) -> str:
    return re.sub(r"[\s_\-（）()]+", "", value).lower()


def _read_tabular_rows(path: str | Path) -> list[dict[str, str]]:
    source = Path(path)
    if source.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError as error:
            raise RuntimeError("读取XLSX需要安装可选依赖: pip install 'aegis-esg[universe]'") from error
        workbook = load_workbook(source, read_only=True, data_only=True)
        sheet = workbook.active
        # Some exchange-generated workbooks incorrectly declare dimension A1
        # while containing thousands of rows. Ignore that stale dimension.
        sheet.reset_dimensions()
        values = sheet.iter_rows(values_only=True)
        try:
            headers = [str(value or "").strip() for value in next(values)]
        except StopIteration:
            return []
        return [dict(zip(headers, (str(value or "").strip() for value in row))) for row in values]
    with source.open(encoding="utf-8-sig", newline="") as stream:
        sample = stream.read(4096)
        stream.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        return list(csv.DictReader(stream, dialect=dialect))


def _is_iso_date(value: str) -> bool:
    try:
        date.fromisoformat(value)
        return bool(re.fullmatch(r"\d{4}-\d{2}-\d{2}", value))
    except ValueError:
        return False
