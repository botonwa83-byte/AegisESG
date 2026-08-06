import csv
import hashlib
import json
import tempfile
import unittest
import gzip
import urllib.parse
import urllib.request
from dataclasses import replace
from unittest.mock import patch
from decimal import Decimal
from pathlib import Path

from aegis_esg.io import read_observations, write_observations, write_ranking_csv
from aegis_esg.methodology import load_methodology
from aegis_esg.methodology_review import (
    evaluate_thin_population_methodology_review,
    prepare_thin_population_methodology_review,
    write_methodology_review_packet,
)
from aegis_esg.models import Direction, Indicator, IndicatorKind, Observation, ValueStatus
from aegis_esg.scoring import MissingStrategy, PopulationStats, ScoringEngine
from aegis_esg.ranking_analysis import analyze_missing_sensitivity, validate_ranking_mode
from aegis_esg.release_guard import FORMAL_ALGORITHM_VERSION, prepare_release_authorization, validate_release_authorization
from aegis_esg.repository import SQLiteRepository
from aegis_esg.extraction import (
    PageText,
    _extract_chinese_env_table_rows,
    extract_batch_text_exports,
    extract_indicator_candidates,
    read_page_text_export,
    resolve_text_export_path,
    summarize_review_candidates,
)
from aegis_esg.env_intensity import (
    CompanyDocument, derive_env_intensity_candidates, derive_ghg_reduction_candidates,
)
from aegis_esg.financial import FinancialFact, derive_financial_observations
from aegis_esg.gap_priority import prioritize_qualitative_gaps
from aegis_esg.gap_diagnostics import _classify as classify_quantitative_gap
from aegis_esg.embedded_coverage import recognize_embedded_esg_coverage
from aegis_esg.quantitative_gap_priority import build_thin_population_gap_batch, prioritize_quantitative_gaps
from aegis_esg.esg_disclosure import (
    QualitativeEvidenceCandidate, collect_annual_qualitative_evidence,
    collect_esg_qualitative_evidence, scan_annual_esg_disclosure,
)
from aegis_esg.evidence_graph import build_evidence_constraint_graph
from aegis_esg.evidence_experiment import (
    evaluate_e1_validation, prepare_e1_validation_sample, write_e1_validation_sample,
)
from aegis_esg.quantitative_validation import (
    apply_quantitative_validation, evaluate_quantitative_validation,
    prepare_quantitative_validation_sample, write_quantitative_validation_sample,
)
from aegis_esg.qualitative_review import (
    QualitativeReviewAudit, QualitativeReviewDecision, apply_qualitative_review_decisions,
    merge_qualitative_candidate_files, plan_qualitative_review,
    read_qualitative_review_decisions, read_qualitative_review_packets,
    reprioritize_evidence_gaps, write_qualitative_review_plan, write_qualitative_review_template,
)
from aegis_esg.dual_review import (
    ArbitrationDecision, DualReviewDecision, apply_arbitration_decisions,
    apply_dual_review_decisions, read_arbitration_decisions, read_dual_review_decisions,
    requires_dual_review, select_dual_review_cases, write_dual_review_template,
)
from aegis_esg.io import merge_confirmed_observations
from aegis_esg.review_batch import (
    apply_review_batch, create_review_batch, read_batch_ledger, read_batch_rows,
    read_review_progress,
)
from aegis_esg.quality import evaluate_quality
from aegis_esg.resolution import ResolutionDecision, audit_resolution_preview, plan_review_tiers, resolve_pending_candidates, select_manual_review_candidates
from aegis_esg.resolution import ReviewTier
from aegis_esg.review_priority import prioritize_review_by_impact
from aegis_esg.review import ReviewInstruction, apply_conflict_review_instructions, apply_review_instructions, read_review_instructions
from aegis_esg.sources.sse import classify_title, discover_reports, parse_response
from aegis_esg.sources.szse import SZSEDisclosure, classify_title as classify_szse_title, discover_batch as discover_szse_batch, discover_reports as discover_szse_reports, parse_response as parse_szse_response
from aegis_esg.sources.listings import collect_listing_pages, parse_listing_page
from aegis_esg.sources.hkex import import_hkex_securities
from aegis_esg.sources.hkex_profile import collect_hkex_issuer_profiles, parse_hkex_access_token, parse_hkex_quote_payload, prepare_hkex_evidence_drafts
from aegis_esg.sources.hkex_disclosure import HKEXDisclosure, _fetch, classify_continuity_document, discover_hkex_continuity_batch, discover_hkex_continuity_documents, parse_stock_lookup, parse_title_search, select_continuity_downloads
from aegis_esg.sources.bse import collect_bse_listings, parse_bse_code_mapping, parse_bse_page, parse_bse_disclosures, classify_disclosure_title, discover_bse_reports, discover_bse_annual_report
from aegis_esg.collector import DocumentRecord, _decode_document, _download_candidates, _read_document_index, collect_batch, dedupe_document_records, supersede_documents, write_document_index
from aegis_esg.completion import audit_project_completion
from aegis_esg.continuity_evidence import extract_continuity_evidence_candidates, finalize_continuity_reviews, prepare_continuity_review_packets, render_continuity_review_guide, select_continuity_review_batch, write_continuity_evidence_candidates
from aegis_esg.universe import UniverseCompany, audit_universe
from aegis_esg.universe_builder import ExchangeSecurity, audit_snapshot, build_energy_universe, normalize_exchange_export, normalize_stock_code, read_exchange_snapshot, write_universe
from aegis_esg.reference import extract_reference_securities
from aegis_esg.registry import normalize_company_name, reconcile_registry
from aegis_esg.planning import audit_document_coverage, collection_summary, merge_document_indexes, plan_collection
from aegis_esg.historical import import_historical_workbook
from aegis_esg.migration import augment_candidate_universe, bind_snapshot_provenance, plan_historical_migration, write_candidate_universe
from aegis_esg.indicator_plan import plan_candidate_coverage, plan_indicator_tasks
from aegis_esg.incremental import benchmark_incremental_scoring, plan_incremental_recompute
from aegis_esg.dashboard import load_progress_dashboard, render_conflict_review_template, render_progress_dashboard
from aegis_esg.issuer_continuity import apply_issuer_continuity_decisions, audit_hkex_issuer_continuity, plan_continuity_evidence_tasks
from aegis_esg.universe_review import apply_universe_evidence, merge_universe_evidence_batches, plan_universe_evidence


ROOT = Path(__file__).resolve().parents[1]


class MethodologyTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.methodology = load_methodology(ROOT / "data/methodologies/energy_esg_2025.json")

    def test_methodology_matches_report(self):
        self.assertEqual(37, len(self.methodology.quantitative))
        self.assertEqual(43, len(self.methodology.qualitative))
        self.assertAlmostEqual(100, sum(i.weight for i in self.methodology.quantitative))
        self.assertAlmostEqual(100, sum(i.weight for i in self.methodology.qualitative))
        self.assertEqual(10, sum(i.key_indicator for i in self.methodology.quantitative))

    def test_thin_methodology_review_freezes_inputs_and_requires_real_signature(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            summary_path = root / "summary.json"
            diagnostics_path = root / "diagnostics.csv"
            packet_path = root / "packet.csv"
            manifest_path = root / "manifest.json"
            summary_path.write_text(json.dumps({
                "minimum_population_threshold": 20,
                "below_minimum_population_indicator_codes": ["Q_E_SO2_INTENSITY"],
                "indicator_population": {"Q_E_SO2_INTENSITY": 13},
            }), encoding="utf-8")
            with diagnostics_path.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=["indicator_code", "diagnostic_category"])
                writer.writeheader()
                writer.writerow({"indicator_code": "Q_E_SO2_INTENSITY", "diagnostic_category": "no_matching_disclosure_in_text"})
            rows, manifest = prepare_thin_population_methodology_review(summary_path, diagnostics_path)
            self.assertEqual(7, rows[0]["population_deficit"])
            self.assertFalse(manifest["applicable"])
            write_methodology_review_packet(packet_path, manifest_path, rows, manifest)
            with self.assertRaisesRegex(ValueError, "决定无效"):
                evaluate_thin_population_methodology_review(packet_path, manifest_path)
            rows[0].update({
                "decision": "retain_threshold_with_thin_sample_warning",
                "reviewer": "方法论负责人甲",
                "reviewed_at": "2026-08-03T20:00:00+08:00",
                "rationale": "公开披露不足，保留阈值并明确薄样本风险。",
            })
            write_methodology_review_packet(packet_path, manifest_path, rows, manifest)
            report = evaluate_thin_population_methodology_review(packet_path, manifest_path)
            self.assertTrue(report["all_decisions_signed"])
            self.assertFalse(report["methodology_change_authorized"])
            self.assertFalse(report["scoring_authorized"])

    def test_thin_methodology_review_rejects_placeholder_and_manifest_mutation(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            summary_path = root / "summary.json"
            diagnostics_path = root / "diagnostics.csv"
            packet_path = root / "packet.csv"
            manifest_path = root / "manifest.json"
            summary_path.write_text(json.dumps({
                "minimum_population_threshold": 20,
                "below_minimum_population_indicator_codes": ["Q_E_ALTERNATIVE_WATER_RATE"],
                "indicator_population": {"Q_E_ALTERNATIVE_WATER_RATE": 7},
            }), encoding="utf-8")
            diagnostics_path.write_text(
                "indicator_code,diagnostic_category\nQ_E_ALTERNATIVE_WATER_RATE,related_fields_incomplete\n",
                encoding="utf-8-sig",
            )
            rows, manifest = prepare_thin_population_methodology_review(summary_path, diagnostics_path)
            rows[0].update({
                "decision": "commission_additional_data", "reviewer": "system",
                "reviewed_at": "2026-08-03T20:00:00+08:00", "rationale": "需要补充外部数据后再决定。",
            })
            write_methodology_review_packet(packet_path, manifest_path, rows, manifest)
            with self.assertRaisesRegex(ValueError, "真实审核人"):
                evaluate_thin_population_methodology_review(packet_path, manifest_path)
            rows[0]["review_id"] = "tampered"
            rows[0]["reviewer"] = "方法论负责人乙"
            write_methodology_review_packet(packet_path, manifest_path, rows, manifest)
            with self.assertRaisesRegex(ValueError, "冻结清单"):
                evaluate_thin_population_methodology_review(packet_path, manifest_path)

    def test_specialized_gap_diagnostic_requires_numeric_physical_total(self):
        revenue = "营业收入（元）725,540,857.14。"
        self.assertEqual(
            "related_fields_incomplete",
            classify_quantitative_gap(
                "Q_E_CLEAN_ENERGY_INTENSITY",
                "陕西美能清洁能源集团股份有限公司积极探索绿电业务。" + revenue,
            )[0],
        )
        self.assertEqual(
            "possible_clean_energy_revenue_closure",
            classify_quantitative_gap(
                "Q_E_CLEAN_ENERGY_INTENSITY",
                "清洁能源使用总量36,635.24兆瓦时。" + revenue,
            )[0],
        )
        self.assertEqual(
            "related_fields_incomplete",
            classify_quantitative_gap(
                "Q_E_HAZ_WASTE_INTENSITY", "公司依法管理危险废物。" + revenue,
            )[0],
        )
        self.assertEqual(
            "possible_hazardous_waste_revenue_closure",
            classify_quantitative_gap(
                "Q_E_HAZ_WASTE_INTENSITY", "危险废物产生量12.5吨。" + revenue,
            )[0],
        )

    def test_gap_diagnostic_keeps_sox_separate_and_classifies_alternative_water(self):
        self.assertEqual(
            "no_matching_disclosure_in_text",
            classify_quantitative_gap(
                "Q_E_SO2_INTENSITY", "Sulfur oxides (SOx) Tonnes 0.78 1.28",
            )[0],
        )
        self.assertEqual(
            "possible_direct_alternative_water_rate",
            classify_quantitative_gap(
                "Q_E_ALTERNATIVE_WATER_RATE", "再生水使用量占总用水量比例为 18.6%。",
            )[0],
        )
        self.assertEqual(
            "possible_alternative_water_formula_closure",
            classify_quantitative_gap(
                "Q_E_ALTERNATIVE_WATER_RATE",
                "再生水使用量 12,000 立方米。总用水量 80,000 立方米。",
            )[0],
        )

    def test_alternative_water_rate_reads_group_narrative_and_explicit_year_tables(self):
        pages = [
            PageText(10, "2025 年，公司循环水用量 13,785.74 吨，循环水用量占比 6.81%。"),
            PageText(11, "指标 单位 2023年数据 2024年数据 2025年数据\n替代水源占比4 % - 27 38"),
            PageText(12, "水资源利用指标 2024年\n单位\n2025年\n替代水源用水量占比1\n%\n2.63\n5.15"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"]
        self.assertEqual([6.81, 38, 5.15], values)
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertFalse([
            item for item in confirmed if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"
        ])
        self.assertEqual(3, len([
            item for item in unresolved if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"
        ]))

    def test_alternative_water_rate_rejects_site_kpi_and_hydropower_name_collision(self):
        pages = [
            PageText(20, "义乌基地 ESG 指标绩效\n替代水源占比 % 94.42"),
            PageText(21, "中水电装机 113.28 万千瓦，占比 16.4%"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        self.assertFalse([
            item for item in items if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"
        ])

    def test_alternative_water_rate_reads_reclaimed_water_reuse_rate(self):
        pages = [
            PageText(8, "报告期内，公司中水回用率为 61.87%。"),
            PageText(9, "中水使用占比达 40%。"),
            PageText(10, "中水使用占比超 95%。"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"]
        self.assertEqual([61.87, 40.0], values)

    def test_split_chinese_so2_intensity_table_converts_tonnes_per_rmb_million(self):
        pages = [PageText(129, """废气污染物减排情况
指标 单位 2025年
二氧化硫
排放总量
吨
611.85
排放强度
吨/百万元营收
0.0072""")]
        items = extract_indicator_candidates(
            pages, "600438.SH", "通威股份", 2025, "url", "esg_report.pdf",
        )
        so2 = [item for item in items if item.indicator_code == "Q_E_SO2_INTENSITY"]
        self.assertEqual([72], [item.value for item in so2])
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertEqual([72], [
            item.value for item in confirmed if item.indicator_code == "Q_E_SO2_INTENSITY"
        ])
        self.assertFalse([
            item for item in unresolved if item.indicator_code == "Q_E_SO2_INTENSITY"
        ])
        rejected = extract_indicator_candidates(
            [PageText(130, "废气污染物减排情况\n指标 单位 2025年\n硫氧化物\n排放强度\n吨/百万元营收\n0.0072")],
            "A", "甲", 2025, "url", "esg_report.pdf",
        )
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_E_SO2_INTENSITY"])

    def test_environmental_gap_diagnostic_requires_actual_physical_total(self):
        revenue = "营业收入（元）725,540,857.14。"
        self.assertEqual(
            "related_disclosure_without_compatible_intensity",
            classify_quantitative_gap(
                "Q_E_GHG_INTENSITY", "公司建设光伏项目，累计减少二氧化碳排放103.22万吨。" + revenue,
            )[0],
        )
        self.assertEqual(
            "possible_total_plus_rmb_revenue_derivation",
            classify_quantitative_gap(
                "Q_E_GHG_INTENSITY", "温室气体排放总量为12,345.6吨二氧化碳当量。" + revenue,
            )[0],
        )
        self.assertEqual(
            "related_disclosure_without_compatible_intensity",
            classify_quantitative_gap(
                "Q_E_NOX_INTENSITY", "NOx Emission Reduction Tonnes 2,117. Revenue RMB million 725.5.",
            )[0],
        )
        self.assertEqual(
            "disclosed_scope_mismatch_requires_review",
            classify_quantitative_gap(
                "Q_E_WATER_INTENSITY",
                "总耗水量769,323.80吨。披露范围包括三家生产基地。营业收入（元）725,540,857.14。",
            )[0],
        )

    def test_specialized_formula_diagnostic_requires_local_numeric_amounts(self):
        self.assertEqual(
            "related_fields_incomplete",
            classify_quantitative_gap(
                "Q_S_RD_RATE", "公司重视研发创新。营业收入（元）725,540,857.14。",
            )[0],
        )
        self.assertEqual(
            "possible_rd_revenue_formula_closure",
            classify_quantitative_gap(
                "Q_S_RD_RATE", "研发投入金额12,500万元。营业收入725,540万元。",
            )[0],
        )
        self.assertEqual(
            "related_fields_incomplete",
            classify_quantitative_gap(
                "Q_S_SAFETY_INVEST_RATE", "公司完善安全生产投入制度。营业收入725,540万元。",
            )[0],
        )
        self.assertNotEqual(
            "likely_methodology_compatible_rule_gap",
            classify_quantitative_gap(
                "Q_E_SO2_INTENSITY",
                "公司环保投入5,543.47万元，持续推动二氧化硫减排，实现污染物排放强度下降。",
            )[0],
        )
        self.assertEqual(
            "likely_methodology_compatible_rule_gap",
            classify_quantitative_gap(
                "Q_E_SO2_INTENSITY", "二氧化硫排放强度 克/万元营业收入 2.5。",
            )[0],
        )
        self.assertEqual(
            "related_fields_incomplete",
            classify_quantitative_gap(
                "Q_E_GHG_REDUCTION_RATE", "公司在2025年持续推进温室气体管理，较2024年加强治理。",
            )[0],
        )
        self.assertEqual(
            "possible_two_year_ghg_formula_closure",
            classify_quantitative_gap(
                "Q_E_GHG_REDUCTION_RATE",
                "2025 2024\n温室气体排放总量 吨 900 1,000。",
            )[0],
        )

    def test_completion_audit_exposes_release_blockers(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            values = {
                "documents.json": {"annual_coverage_count": 612},
                "quantitative.json": {"company_count": 614, "quantitative_indicator_count": 37,
                                      "candidate_task_count": 7597},
                "qualitative.json": {"qualitative_indicator_count": 43,
                                     "review_packet_count": 10538, "auto_confirmed_count": 0},
                "resolution.json": {"manual_required_group_count": 145,
                                    "freeze_ready": False, "applicable": False},
            }
            for name, value in values.items():
                (root / name).write_text(json.dumps(value), encoding="utf-8")
            report = audit_project_completion(
                root / "documents.json", root / "quantitative.json",
                root / "qualitative.json", root / "resolution.json",
            )
            self.assertFalse(report["publishable"])
            self.assertEqual("universe", report["next_gate"])
            self.assertEqual(18, report["gates"]["universe"]["target"] - report["gates"]["universe"]["current"])
            self.assertEqual(145, report["gates"]["review"]["quantitative_manual_groups"])

    def test_completion_audit_requires_all_six_gates(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            values = (
                {"annual_coverage_count": 2},
                {"company_count": 2, "quantitative_indicator_count": 37, "candidate_task_count": 74},
                {"qualitative_indicator_count": 43, "review_packet_count": 86, "auto_confirmed_count": 86},
                {"manual_required_group_count": 0, "freeze_ready": True, "applicable": True},
            )
            paths = []
            for index, value in enumerate(values):
                path = root / f"{index}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            report = audit_project_completion(*paths, expected_companies=2)
            self.assertTrue(report["publishable"])
            self.assertEqual(6, report["completed_gate_count"])

    def test_completion_audit_v2_accepts_validated_risk_gates_without_dense_grid(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            values = (
                {"annual_coverage_count": 2},
                {"company_count": 2, "quantitative_indicator_count": 37, "candidate_task_count": 20,
                 "zero_coverage_indicator_count": 0, "minimum_population_gate_passed": True,
                 "sampling_accuracy_passed": True},
                {"qualitative_indicator_count": 43, "review_packet_count": 30, "auto_confirmed_count": 10,
                 "classification_thresholds_validated": True, "sampling_accuracy_passed": True,
                 "high_risk_open_count": 0, "open_arbitration_count": 0},
                {"manual_required_group_count": 0, "freeze_ready": True, "applicable": True},
            )
            paths = []
            for index, value in enumerate(values):
                path = root / f"risk-{index}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            report = audit_project_completion(*paths, expected_companies=2)
            self.assertTrue(report["publishable"])
            self.assertEqual("project-completion-risk-v2", report["policy_version"])
            self.assertEqual("risk_gate", report["gates"]["quantitative"]["completion_basis"])
            self.assertEqual("risk_gate", report["gates"]["qualitative"]["completion_basis"])

    def test_completion_audit_v2_does_not_infer_missing_validation_evidence(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            values = (
                {"annual_coverage_count": 2},
                {"company_count": 2, "quantitative_indicator_count": 37, "candidate_task_count": 20,
                 "zero_coverage_indicator_count": 0},
                {"qualitative_indicator_count": 43, "review_packet_count": 30, "auto_confirmed_count": 10,
                 "high_risk_open_count": 0, "open_arbitration_count": 0},
                {"manual_required_group_count": 0, "freeze_ready": True, "applicable": True},
            )
            paths = []
            for index, value in enumerate(values):
                path = root / f"missing-{index}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            report = audit_project_completion(*paths, expected_companies=2)
        self.assertFalse(report["gates"]["quantitative"]["complete"])
        self.assertFalse(report["gates"]["qualitative"]["complete"])

    def test_stage_orchestrator_reports_first_external_gate_without_mutation(self):
        from aegis_esg.stage_orchestrator import assess_next_stage
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "completion.json"
            path.write_text(json.dumps({
                "publishable": False,
                "gates": {
                    "universe": {"complete": False, "blocker": "主体待审核"},
                    "documents": {"complete": False}, "quantitative": {"complete": False},
                    "qualitative": {"complete": False}, "review": {"complete": False},
                    "release": {"complete": False},
                },
            }), encoding="utf-8")
            result = assess_next_stage(path)
            self.assertEqual("blocked_external", result["status"])
            self.assertEqual("M3", result["next_stage"])
            self.assertEqual("universe", result["gate"])
            self.assertFalse(result["publishable"])

    def test_external_readiness_reports_unsigned_inputs(self):
        from aegis_esg.external_readiness import audit_external_readiness
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            completion = root / "completion.json"
            quantitative = root / "quantitative.json"
            thin = root / "thin.json"
            release = root / "release.json"
            patent = root / "patent.md"
            completion.write_text(json.dumps({"publishable": False, "completed_gate_count": 0, "gate_count": 6}), encoding="utf-8")
            quantitative.write_text(json.dumps({"applicable": False, "signed_count": 0, "sample_count": 176}), encoding="utf-8")
            thin.write_text(json.dumps({"applicable": False, "signed_count": 0, "review_count": 3}), encoding="utf-8")
            release.write_text(json.dumps({"authorized": False}), encoding="utf-8")
            patent.write_text("template", encoding="utf-8")
            report = audit_external_readiness(completion, quantitative, thin, release, patent)
            self.assertEqual("blocked_external", report["status"])
            self.assertFalse(report["ready"])
            self.assertEqual("0/176 signed", report["checks"]["quantitative_sampling"]["evidence"])

    def test_external_readiness_can_include_unfinished_patent_experiments(self):
        from aegis_esg.external_readiness import audit_external_readiness
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            paths = []
            for name, value in (
                ("completion", {"publishable": False, "completed_gate_count": 0, "gate_count": 6}),
                ("quantitative", {"applicable": False, "signed_count": 0, "sample_count": 1}),
                ("thin", {"applicable": False, "signed_count": 0, "review_count": 1}),
                ("release", {"authorized": False}),
                ("e1", {"applicable": False, "signed_count": 0, "sample_count": 1}),
                ("e2", {"applicable": False, "task_count": 1}),
            ):
                path = root / f"{name}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            patent = root / "patent.md"
            patent.write_text("template", encoding="utf-8")
            report = audit_external_readiness(paths[0], paths[1], paths[2], paths[3], patent, paths[4], paths[5])
            self.assertIn("e1_constraint_experiment", report["checks"])
            self.assertIn("e2_review_scheduling_experiment", report["checks"])
            self.assertFalse(report["ready"])

    def test_auto_stage_combines_stage_and_external_state(self):
        from aegis_esg.auto_stage import run_auto_stage
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            values = [
                {"publishable": False, "completed_gate_count": 0, "gate_count": 6,
                 "gates": {name: {"complete": False} for name in ("universe", "documents", "quantitative", "qualitative", "review", "release")}},
                {"applicable": False, "signed_count": 0, "sample_count": 1},
                {"applicable": False, "signed_count": 0, "review_count": 1},
                {"authorized": False},
            ]
            paths = []
            for index, value in enumerate(values):
                path = root / f"{index}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            patent = root / "patent.md"
            patent.write_text("template", encoding="utf-8")
            report = run_auto_stage(paths[0], paths[1], paths[2], paths[3], patent)
            self.assertEqual("M3", report["next_stage"])
            self.assertFalse(report["continue_automatically"])

    def test_auto_stage_does_not_continue_when_external_experiments_remain(self):
        from aegis_esg.auto_stage import run_auto_stage
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            gates = {name: {"complete": True} for name in (
                "universe", "documents", "quantitative", "qualitative", "review", "release",
            )}
            values = [
                {"publishable": True, "completed_gate_count": 6, "gate_count": 6, "gates": gates},
                {"applicable": True, "signed_count": 1, "sample_count": 1},
                {"applicable": True, "signed_count": 1, "review_count": 1},
                {"authorized": True},
            ]
            paths = []
            for index, value in enumerate(values):
                path = root / f"{index}.json"
                path.write_text(json.dumps(value), encoding="utf-8")
                paths.append(path)
            patent = root / "patent.md"
            patent.write_text("template", encoding="utf-8")
            e1 = root / "e1.json"
            e1.write_text(json.dumps({"applicable": False, "signed_count": 0, "sample_count": 1}), encoding="utf-8")
            report = run_auto_stage(paths[0], paths[1], paths[2], paths[3], patent, e1)
            self.assertEqual("complete", report["next_stage"])
            self.assertFalse(report["continue_automatically"])

    def test_e2_validation_template_stays_unsigned_and_rejects_incomplete_rows(self):
        from aegis_esg.e2_experiment import prepare_e2_validation_sample, write_e2_validation_sample, evaluate_e2_validation
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            impact = root / "impact.csv"
            impact.write_text("priority,company_code,report_year,indicator_code,impact_score,crosses_top_200,baseline_confidence_rank,baseline_weight_rank\n1,A,2025,Q,80,True,4,2\n", encoding="utf-8")
            rows, summary = prepare_e2_validation_sample(impact)
            self.assertEqual(1, summary["task_count"])
            self.assertFalse(summary["applicable"])
            sample = root / "e2.csv"
            summary_path = root / "summary.json"
            write_e2_validation_sample(sample, summary_path, rows, summary)
            with self.assertRaisesRegex(ValueError, "字段不完整"):
                evaluate_e2_validation(sample)

    def test_normal_direction(self):
        engine = ScoringEngine(self.methodology)
        stat = PopulationStats(3, 10, 2)
        positive = Indicator("P", "E", "x", "x", IndicatorKind.QUANTITATIVE, 100, Direction.POSITIVE)
        negative = Indicator("N", "E", "x", "x", IndicatorKind.QUANTITATIVE, 100, Direction.NEGATIVE)
        self.assertGreater(engine._score_value(positive, 12, stat), engine._score_value(positive, 8, stat))
        self.assertGreater(engine._score_value(negative, 8, stat), engine._score_value(negative, 12, stat))

    def test_dlt2971_grade_bands_and_na_rules(self):
        from aegis_esg.grade import GradeFlags, map_esg_grade
        cases = [
            (95, 80, "AAA"), (90, 80, "AAA"), (75, 80, "AA"), (60, 80, "A"),
            (50, 80, "BBB"), (40, 80, "BB"), (30, 80, "B"), (20, 80, "C"),
            (19.99, 80, "NA"), (0, 80, "NA"),
        ]
        for score, disclosure, expected in cases:
            result = map_esg_grade(score, disclosure)
            self.assertEqual(expected, result.grade, msg=(score, disclosure))
            self.assertEqual("score_band", result.reason)
        low_disclosure = map_esg_grade(95, 49.99)
        self.assertEqual("NA", low_disclosure.grade)
        self.assertEqual("disclosure_below_half", low_disclosure.reason)
        half_disclosure = map_esg_grade(95, 50.0)
        self.assertEqual("AAA", half_disclosure.grade)
        incident = map_esg_grade(95, 90, GradeFlags(major_safety_incident=True))
        self.assertEqual("NA", incident.grade)
        self.assertEqual("major_safety_incident", incident.reason)

    def test_scoring_engine_attaches_dlt2971_grade(self):
        from aegis_esg.grade import GradeFlags
        observations = []
        for index, indicator in enumerate(self.methodology.indicators):
            if indicator.kind == IndicatorKind.QUALITATIVE:
                value = 100
            elif indicator.direction == Direction.NEGATIVE:
                value = 1
            else:
                value = 100 + index
            observations.append(Observation("A", "甲", 2025, indicator.code, value))
        # Sparse company: only first indicator confirmed → disclosure far below 50%.
        sparse = [Observation("B", "乙", 2025, self.methodology.indicators[0].code, 10)]
        engine = ScoringEngine(self.methodology)
        dense = engine.evaluate(observations)[0]
        self.assertIn(dense.grade, {"AAA", "AA", "A", "BBB", "BB", "B", "C", "NA"})
        self.assertTrue(dense.grade)
        self.assertIn("grade", dense.to_dict(include_details=False))
        thin = engine.evaluate(sparse)[0]
        self.assertLess(thin.disclosure_rate, 50)
        self.assertEqual("NA", thin.grade)
        self.assertEqual("disclosure_below_half", thin.grade_reason)
        flagged = engine.evaluate(
            observations, company_flags={"A": GradeFlags(accident_misreport=True)},
        )[0]
        self.assertEqual("NA", flagged.grade)
        self.assertEqual("accident_misreport", flagged.grade_reason)

    def test_evaluation_dense_rank_and_export_shape(self):
        observations = []
        for code, name, offset in (("A", "甲公司", 0), ("B", "乙公司", 1), ("C", "丙公司", 1)):
            for index, indicator in enumerate(self.methodology.indicators):
                if indicator.kind == IndicatorKind.QUALITATIVE:
                    value = 80 if offset == 0 else 50
                elif indicator.direction == Direction.NEGATIVE:
                    value = 10 + index + offset
                elif indicator.direction == Direction.BIDIRECTIONAL:
                    value = 50 + offset
                else:
                    value = 50 + index - offset
                observations.append(Observation(code, name, 2024, indicator.code, value))
        results = ScoringEngine(self.methodology).evaluate(observations)
        self.assertEqual([1, 2, 2], [r.rank for r in results])
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "ranking.csv"
            write_ranking_csv(output, results, self.methodology)
            with output.open(encoding="utf-8-sig") as stream:
                rows = list(csv.reader(stream))
            self.assertEqual(7, len(rows))
            self.assertEqual(18, len(rows[0]))
            self.assertEqual("数值类别", rows[0][3])
            self.assertEqual("披露率%", rows[0][4])
            self.assertEqual("指标数值", rows[1][3])
            self.assertEqual("指标分值", rows[2][3])

    def test_universe_disclosed_population_baseline_excludes_outsiders_and_missing(self):
        from aegis_esg.scoring import build_population_baseline_report

        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 10, ValueStatus.CONFIRMED),
            Observation("B", "乙", 2025, indicator.code, 30, ValueStatus.CONFIRMED),
            Observation("C", "丙", 2025, indicator.code, 1000, ValueStatus.CONFIRMED),
            Observation("D", "丁", 2025, indicator.code, None, ValueStatus.MISSING),
        ]
        engine = ScoringEngine(self.methodology, minimum_population=2)
        results = {
            item.company_code: item
            for item in engine.evaluate(
                observations,
                MissingStrategy.LEGACY_ZERO_V1,
                universe_codes={"A", "B", "D"},
                minimum_population=2,
            )
        }
        self.assertEqual({"A", "B", "D"}, set(results))
        self.assertNotIn("C", results)
        detail_a = next(item for item in results["A"].details if item.indicator_code == indicator.code)
        self.assertEqual(2, detail_a.population_count)
        self.assertEqual(20.0, detail_a.mean)
        self.assertFalse(detail_a.thin_population)
        self.assertEqual(0, results["D"].total_score)
        baseline = build_population_baseline_report(
            observations,
            self.methodology,
            universe_codes={"A", "B", "D"},
            expected_companies=3,
            minimum_population=2,
        )
        row = next(item for item in baseline["indicators"] if item["indicator_code"] == indicator.code)
        self.assertEqual(2, row["disclosed_company_count"])
        self.assertEqual(3, row["universe_company_count"])
        self.assertEqual(3, row["observed_company_count"])
        self.assertFalse(row["thin_population"])
        # Other methodology quantitative indicators remain undisclosed → thin gate fails.
        self.assertFalse(baseline["minimum_population_gate_passed"])
        self.assertFalse(baseline["formal_baseline_ready"])
        incomplete = build_population_baseline_report(
            observations,
            self.methodology,
            universe_codes={"A", "B"},
            expected_companies=3,
            minimum_population=2,
        )
        self.assertEqual(2, incomplete["universe_company_count"])
        self.assertLess(incomplete["universe_company_count"], incomplete["expected_companies"])

    def test_versioned_missing_strategies_do_not_silently_share_zero_behavior(self):
        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 10, ValueStatus.CONFIRMED),
            Observation("B", "乙", 2025, indicator.code, None, ValueStatus.PENDING),
        ]
        engine = ScoringEngine(self.methodology)
        zero = {item.company_code: item for item in engine.evaluate(
            observations, MissingStrategy.LEGACY_ZERO_V1,
        )}
        neutral = {item.company_code: item for item in engine.evaluate(
            observations, MissingStrategy.INDICATOR_NEUTRAL_V1,
        )}
        disclosed = {item.company_code: item for item in engine.evaluate(
            observations, MissingStrategy.DISCLOSED_WEIGHT_V1,
        )}
        self.assertEqual(0, zero["B"].total_score)
        self.assertGreater(neutral["B"].total_score, zero["B"].total_score)
        self.assertGreater(disclosed["A"].total_score, zero["A"].total_score)
        self.assertEqual(0, disclosed["B"].total_score)

    def test_missing_strategy_sensitivity_reports_rank_span(self):
        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 10, ValueStatus.CONFIRMED),
            Observation("B", "乙", 2025, indicator.code, None, ValueStatus.MISSING),
        ]
        report = analyze_missing_sensitivity(observations, self.methodology)
        self.assertEqual(2, report["company_count"])
        self.assertEqual(3, len(report["strategy_versions"]))
        self.assertIn("rank_span", report["companies"][0])
        self.assertEqual(3, len(report["strategy_comparisons"]))
        self.assertEqual(2, sum(report["credibility_grade_counts"].values()))
        self.assertTrue(all("top_200_overlap_rate" in item for item in report["strategy_comparisons"]))

    def test_release_mode_requires_explicit_strategy_and_rejects_pending(self):
        indicator = self.methodology.quantitative[0]
        pending = [Observation("A", "甲", 2025, indicator.code, None, ValueStatus.PENDING)]
        with self.assertRaisesRegex(ValueError, "显式指定"):
            validate_ranking_mode("release", [], None)
        with self.assertRaisesRegex(ValueError, "pending"):
            validate_ranking_mode("release", pending, MissingStrategy.LEGACY_ZERO_V1.value)
        selected = validate_ranking_mode(
            "release",
            [Observation("A", "甲", 2025, indicator.code, 1, ValueStatus.CONFIRMED)],
            MissingStrategy.INDICATOR_NEUTRAL_V1.value,
        )
        self.assertEqual(MissingStrategy.INDICATOR_NEUTRAL_V1, selected)
        research = Observation(
            "A", "甲", 2025, indicator.code, 50, ValueStatus.CONFIRMED,
            evidence_text="heuristic [research-only:auto-qualitative-v1;not-formal]",
        )
        with self.assertRaisesRegex(ValueError, "研究域机器观测"):
            validate_ranking_mode(
                "release", [research], MissingStrategy.INDICATOR_NEUTRAL_V1.value,
            )

    def test_release_authorization_rejects_tamper_machine_and_stale_algorithm(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            observations = root / "observations.csv"
            methodology = root / "methodology.json"
            manifest = root / "release.json"
            observations.write_text("frozen observations\n", encoding="utf-8")
            methodology.write_text("{}\n", encoding="utf-8")
            unsigned = prepare_release_authorization(
                observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
            )
            manifest.write_text(json.dumps(unsigned), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "审批人"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                )
            payload = {
                "manifest_version": "release-authorization-v1",
                "algorithm_version": FORMAL_ALGORITHM_VERSION,
                "scope": "official_release",
                "input_sha256": hashlib.sha256(observations.read_bytes()).hexdigest(),
                "methodology_sha256": hashlib.sha256(methodology.read_bytes()).hexdigest(),
                "missing_strategy_version": MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                "approvals": [
                    {"reviewer": "reviewer-a", "role": "methodology_owner", "reviewed_at": "2026-08-03T10:00:00+08:00", "note": "method frozen"},
                    {"reviewer": "reviewer-b", "role": "data_reviewer", "reviewed_at": "2026-08-03T10:05:00+08:00", "note": "data frozen"},
                ],
            }
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            report = validate_release_authorization(
                manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
            )
            self.assertTrue(report["authorized"])

            observations.write_text("tampered\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "input_sha256"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                )
            observations.write_text("frozen observations\n", encoding="utf-8")
            payload["algorithm_version"] = "auto_prerank_v1"
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "算法版本"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                )
            payload["algorithm_version"] = FORMAL_ALGORITHM_VERSION
            payload["approvals"][1]["reviewer"] = "system"
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "机器身份"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                )

    def test_release_authorization_binds_all_six_completion_gates(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            observations = root / "observations.csv"
            methodology = root / "methodology.json"
            completion = root / "completion.json"
            manifest = root / "release.json"
            observations.write_text("frozen\n", encoding="utf-8")
            methodology.write_text("{}\n", encoding="utf-8")
            completion.write_text(json.dumps({
                "publishable": False,
                "gates": {name: {"complete": False} for name in (
                    "universe", "documents", "quantitative", "qualitative", "review", "release",
                )},
            }), encoding="utf-8")
            payload = prepare_release_authorization(
                observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value, completion,
            )
            payload["approvals"] = [
                {"reviewer": "a", "role": "methodology_owner", "reviewed_at": "2026-08-04T10:00:00+08:00", "note": "x"},
                {"reviewer": "b", "role": "data_reviewer", "reviewed_at": "2026-08-04T10:01:00+08:00", "note": "y"},
            ]
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "六道完成门禁"):
                validate_release_authorization(
                    manifest, observations, methodology,
                    MissingStrategy.INDICATOR_NEUTRAL_V1.value, completion,
                )

    def test_dlt_release_validity_and_committee_gate(self):
        from datetime import datetime, timedelta, timezone
        from aegis_esg.release_guard import check_release_effective, seal_release_validity
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            observations = root / "observations.csv"
            methodology = root / "methodology.json"
            manifest = root / "release.json"
            observations.write_text("frozen observations\n", encoding="utf-8")
            methodology.write_text("{}\n", encoding="utf-8")
            start = datetime(2026, 8, 4, 10, 0, tzinfo=timezone(timedelta(hours=8)))
            payload = prepare_release_authorization(
                observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                seal_validity=True, valid_from=start,
            )
            self.assertEqual("DL/T 2971—2025", payload["standard_ref"])
            self.assertEqual(365, payload["result_validity_days"])
            self.assertTrue(payload["valid_from"])
            self.assertTrue(payload["valid_until"])
            payload["approvals"] = [
                {"reviewer": "a", "role": "methodology_owner", "reviewed_at": "2026-08-04T10:00:00+08:00", "note": "method"},
                {"reviewer": "b", "role": "data_reviewer", "reviewed_at": "2026-08-04T10:01:00+08:00", "note": "data"},
            ]
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            # Without require_dlt_process, sealed validity is checked but expiry does not block.
            report = validate_release_authorization(
                manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                as_of=start + timedelta(days=400),
            )
            self.assertTrue(report["authorized"])
            self.assertTrue(report["validity"]["expired"])
            with self.assertRaisesRegex(ValueError, "一年有效期"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                    require_dlt_process=True, as_of=start + timedelta(days=400),
                )
            with self.assertRaisesRegex(ValueError, "evaluation_lead"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                    require_dlt_process=True, as_of=start + timedelta(days=10),
                )
            payload["committee_approvals"] = [
                {"reviewer": "a", "role": "evaluation_lead", "reviewed_at": "2026-08-04T10:02:00+08:00", "note": "committee"},
            ]
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "同一人"):
                validate_release_authorization(
                    manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                    require_dlt_process=True, as_of=start + timedelta(days=10),
                )
            payload["committee_approvals"] = [
                {"reviewer": "c", "role": "evaluation_lead", "reviewed_at": "2026-08-04T10:02:00+08:00", "note": "committee"},
            ]
            manifest.write_text(json.dumps(payload), encoding="utf-8")
            ok = validate_release_authorization(
                manifest, observations, methodology, MissingStrategy.INDICATOR_NEUTRAL_V1.value,
                require_dlt_process=True, as_of=start + timedelta(days=10),
            )
            self.assertTrue(ok["authorized"])
            self.assertEqual(1, ok["committee_approval_count"])
            self.assertTrue(ok["validity"]["effective"])
            sealed = seal_release_validity(start)
            effective = check_release_effective({**payload, **sealed}, as_of=start)
            self.assertTrue(effective["effective"])

    def test_validate_graded_ranking_requires_table1_grades(self):
        from aegis_esg.release_guard import validate_graded_ranking
        with self.assertRaisesRegex(ValueError, "缺少DL/T 2971级别"):
            validate_graded_ranking([{"company_code": "A", "grade": ""}])
        report = validate_graded_ranking([
            {"company_code": "A", "grade": "AAA"},
            {"company_code": "B", "grade": "NA"},
        ])
        self.assertEqual(2, report["graded_company_count"])
        self.assertEqual(1, report["na_count"])
        self.assertTrue(report["complete"])

    def test_official_domain_review_packet_and_safe_apply(self):
        from aegis_esg.domain_verification import (
            apply_official_domain_review,
            evaluate_official_domain_review,
            prepare_official_domain_review_packet,
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            candidates = root / "candidates.csv"
            index = root / "index.csv"
            queue = root / "queue.csv"
            candidates.write_text(
                "company_code,company_name,official_domain,candidate_url,evidence_file,evidence_count,verification_status,next_action\n"
                "000001.SZ,测试A,example-a.com,https://www.example-a.com/,a.txt,3,candidate_declared_in_issuer_report,核验\n"
                "000001.SZ,测试A,mirror.example.net,http://mirror.example.net/,a2.txt,1,candidate_declared_in_issuer_report,核验\n"
                "000002.SZ,测试B,example-b.com,https://www.example-b.com/esg,b.txt,5,candidate_declared_in_issuer_report,核验\n",
                encoding="utf-8",
            )
            index.write_text(
                "company_code,company_name,report_year,document_type,source_url,local_path,sha256,size\n"
                "000001.SZ,测试A,2025,annual_report,https://x/a.pdf,a.pdf,aa,1\n"
                "000002.SZ,测试B,2025,annual_report,https://x/b.pdf,b.pdf,bb,1\n"
                "000002.SZ,测试B,2025,esg_report,https://x/be.pdf,be.pdf,cc,1\n",
                encoding="utf-8",
            )
            queue.write_text(
                "company_code,company_name,report_year,document_type,source_channel,official_domain,candidate_url,"
                "domain_verification,download_status,next_action,scoring_authorized\n"
                "000001.SZ,测试A,2025,annual_report,issuer_official_website,,,not_submitted,pending_official_url,登记,False\n"
                "000001.SZ,测试A,2025,esg_report,issuer_official_website,,,not_submitted,pending_official_url,登记,False\n"
                "000002.SZ,测试B,2025,annual_report,issuer_official_website,,,not_submitted,pending_official_url,登记,False\n",
                encoding="utf-8",
            )
            summary = prepare_official_domain_review_packet(
                candidates, index,
                csv_path=root / "review.csv",
                html_path=root / "review.html",
                summary_path=root / "summary.json",
                limit=10,
            )
            self.assertEqual(2, summary["row_count"])
            self.assertEqual(1, summary["missing_esg_priority_count"])
            self.assertFalse(summary["download_authorized"])
            with (root / "review.csv").open(encoding="utf-8-sig", newline="") as stream:
                review_rows = list(csv.DictReader(stream))
            self.assertEqual("000001.SZ", review_rows[0]["company_code"])
            blank = evaluate_official_domain_review(review_rows)
            self.assertEqual("blocked_external_review", blank["status"])
            self.assertFalse(blank["download_authorized"])
            for row in review_rows:
                row["verification_decision"] = "verify"
                row["reviewer"] = "alice"
                row["reviewed_at"] = "2026-08-04T17:00:00+08:00"
                row["review_note"] = "与年报封面官网一致"
            with (root / "review.csv").open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=review_rows[0].keys(), lineterminator="\n")
                writer.writeheader()
                writer.writerows(review_rows)
            output_queue = root / "queue_out.csv"
            applied = apply_official_domain_review(
                root / "review.csv", queue,
                output_queue_path=output_queue,
                application_path=root / "application.json",
            )
            self.assertEqual("ready_to_register_verified_domains", applied["status"])
            self.assertTrue(applied["queue_updated"])
            self.assertFalse(applied["download_authorized"])
            self.assertFalse(applied["scoring_authorized"])
            with output_queue.open(encoding="utf-8-sig", newline="") as stream:
                updated = list(csv.DictReader(stream))
            self.assertEqual("example-a.com", updated[0]["official_domain"])
            self.assertEqual("verified", updated[0]["domain_verification"])
            self.assertEqual("pending_report_discovery", updated[0]["download_status"])
            self.assertEqual("", updated[0]["candidate_url"])
            # Illegal decision is rejected and does not touch queue.
            review_rows[0]["verification_decision"] = "auto-approve"
            bad = evaluate_official_domain_review(review_rows)
            self.assertEqual("reject_template", bad["status"])

    def test_official_domain_review_allow_partial_apply(self):
        from aegis_esg.domain_verification import (
            apply_official_domain_review,
            evaluate_official_domain_review,
        )
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            review = root / "review.csv"
            queue = root / "queue.csv"
            review.write_text(
                "priority,company_code,company_name,official_domain,candidate_url,https_ready,"
                "evidence_count,evidence_file,missing_independent_esg,alt_domain_count,"
                "verification_decision,reviewer,reviewed_at,review_note\n"
                "1,000001.SZ,测试A,example-a.com,https://www.example-a.com/,true,2,a.txt,true,0,"
                "verify,alice,2026-08-06T09:00:00+08:00,与年报封面官网一致\n"
                "2,000002.SZ,测试B,example-b.com,https://www.example-b.com/,true,2,b.txt,true,0,"
                ",,,\n",
                encoding="utf-8",
            )
            queue.write_text(
                "company_code,company_name,report_year,document_type,source_channel,official_domain,"
                "candidate_url,domain_verification,download_status,next_action,scoring_authorized\n"
                "000001.SZ,测试A,2025,annual_report,issuer_official_website,,,not_submitted,"
                "pending_official_url,登记,False\n"
                "000002.SZ,测试B,2025,annual_report,issuer_official_website,,,not_submitted,"
                "pending_official_url,登记,False\n",
                encoding="utf-8",
            )
            blocked = evaluate_official_domain_review(list(csv.DictReader(review.open(encoding="utf-8-sig"))))
            self.assertEqual("blocked_external_review", blocked["status"])
            partial = evaluate_official_domain_review(
                list(csv.DictReader(review.open(encoding="utf-8-sig"))),
                allow_partial=True,
            )
            self.assertEqual("ready_to_register_verified_domains", partial["status"])
            self.assertEqual(1, partial["verified_rows"])
            self.assertEqual(1, partial["unsigned_rows"])
            applied = apply_official_domain_review(
                review, queue, output_queue_path=root / "queue_out.csv",
                application_path=root / "app.json", allow_partial=True,
            )
            self.assertTrue(applied["queue_updated"])
            with (root / "queue_out.csv").open(encoding="utf-8-sig", newline="") as stream:
                updated = {row["company_code"]: row for row in csv.DictReader(stream)}
            self.assertEqual("verified", updated["000001.SZ"]["domain_verification"])
            self.assertEqual("not_submitted", updated["000002.SZ"]["domain_verification"])

    def test_document_declared_domain_research_discovery_is_unverified(self):
        from aegis_esg.official_report_discovery import discover_document_declared_domain_reports
        html = '<a href="https://issuer.example.com/esg/2025-sustainability-report.pdf">2025可持续发展报告</a>'

        def fake_fetch(url: str) -> str:
            self.assertTrue(url.startswith("https://issuer.example.com"))
            return html

        rows = discover_document_declared_domain_reports(
            [{
                "company_code": "A",
                "company_name": "甲",
                "official_domain": "issuer.example.com",
                "evidence_count": "3",
            }],
            fetcher=fake_fetch,
            report_year=2025,
            seed_paths=("/esg/",),
        )
        hits = [row for row in rows if row.get("source_url")]
        self.assertEqual(1, len(hits))
        self.assertEqual("research_candidate_unverified", hits[0]["discovery_status"])
        self.assertEqual(
            "https://issuer.example.com/esg/2025-sustainability-report.pdf",
            hits[0]["source_url"],
        )

    def test_same_domain_report_discovery_requires_verified_https(self):
        from aegis_esg.official_report_discovery import (
            extract_same_domain_pdf_candidates,
            prepare_official_report_discovery_packet,
        )
        html = '''
        <a href="/files/2025-annual-report.pdf">2025年年度报告</a>
        <a href="https://other.com/x.pdf">ESG报告</a>
        <a href="https://issuer.example.com/esg/2025-sustainability-report.pdf">2025可持续发展报告</a>
        <a href="http://issuer.example.com/old.pdf">2025年报</a>
        '''
        hits = extract_same_domain_pdf_candidates(
            "https://issuer.example.com/investor/",
            html,
            official_domain="issuer.example.com",
            report_year=2025,
        )
        urls = {item["source_url"] for item in hits}
        self.assertIn("https://issuer.example.com/files/2025-annual-report.pdf", urls)
        self.assertIn("https://issuer.example.com/esg/2025-sustainability-report.pdf", urls)
        self.assertNotIn("https://other.com/x.pdf", urls)
        self.assertTrue(all(item["source_url"].startswith("https://") for item in hits))
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            queue = root / "queue.csv"
            queue.write_text(
                "company_code,company_name,report_year,document_type,source_channel,official_domain,candidate_url,"
                "domain_verification,download_status,next_action,scoring_authorized\n"
                "A,甲,2025,esg_report,issuer_official_website,issuer.example.com,,not_submitted,pending,登记,False\n",
                encoding="utf-8",
            )
            empty = prepare_official_report_discovery_packet(
                queue,
                csv_path=root / "disc.csv",
                html_path=root / "disc.html",
                summary_path=root / "disc.json",
                fetcher=None,
            )
            self.assertEqual("await_verified_domains", empty["status"])
            self.assertFalse(empty["download_authorized"])

            queue.write_text(
                "company_code,company_name,report_year,document_type,source_channel,official_domain,candidate_url,"
                "domain_verification,download_status,next_action,scoring_authorized\n"
                "A,甲,2025,esg_report,issuer_official_website,issuer.example.com,,verified,pending_report_discovery,发现,False\n",
                encoding="utf-8",
            )

            def fake_fetch(url: str) -> str:
                self.assertTrue(url.startswith("https://issuer.example.com"))
                return html

            scanned = prepare_official_report_discovery_packet(
                queue,
                csv_path=root / "disc2.csv",
                html_path=root / "disc2.html",
                summary_path=root / "disc2.json",
                fetcher=fake_fetch,
            )
            self.assertEqual("candidates_pending_review", scanned["status"])
            self.assertGreaterEqual(scanned["candidate_rows"], 2)
            self.assertFalse(scanned["scoring_authorized"])

            from aegis_esg.official_report_discovery import apply_official_report_discovery
            with (root / "disc2.csv").open(encoding="utf-8-sig", newline="") as stream:
                discovery_rows = list(csv.DictReader(stream))
            for row in discovery_rows:
                if not row.get("source_url"):
                    continue
                row["review_decision"] = "accept"
                row["reviewer"] = "bob"
                row["reviewed_at"] = "2026-08-04T17:30:00+08:00"
                row["review_note"] = "同域HTTPS年报链接已人工确认"
            with (root / "disc2.csv").open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=discovery_rows[0].keys(), lineterminator="\n")
                writer.writeheader()
                writer.writerows(discovery_rows)
            applied = apply_official_report_discovery(
                root / "disc2.csv", queue,
                output_queue_path=root / "queue_out.csv",
                application_path=root / "app.json",
            )
            self.assertEqual("ready_to_register_report_urls", applied["status"])
            self.assertTrue(applied["queue_updated"])
            self.assertFalse(applied["download_authorized"])
            with (root / "queue_out.csv").open(encoding="utf-8-sig", newline="") as stream:
                out_rows = list(csv.DictReader(stream))
            self.assertTrue(any(row["candidate_url"].startswith("https://issuer.example.com/") for row in out_rows))
            self.assertEqual("pending_official_download", out_rows[0]["download_status"])

    def test_official_website_queue_preserves_verified_domains(self):
        import subprocess
        import sys
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            # Simulate previous verified queue and a tiny document index.
            queue = root / "output/audit/official_website_source_queue_v1_2025.csv"
            queue.parent.mkdir(parents=True)
            queue.write_text(
                "company_code,company_name,report_year,document_type,source_channel,official_domain,candidate_url,"
                "domain_verification,download_status,next_action,scoring_authorized\n"
                "A001.SZ,甲,2025,annual_report,issuer_official_website,issuer.example.com,,"
                "verified,pending_report_discovery,发现,False\n"
                "A001.SZ,甲,2025,esg_report,issuer_official_website,issuer.example.com,,"
                "verified,pending_report_discovery,发现,False\n",
                encoding="utf-8",
            )
            index = root / "data/raw/all_markets_document_index.csv"
            index.parent.mkdir(parents=True)
            index.write_text(
                "company_code,company_name,report_year,document_type,source_url,local_path,sha256,size\n"
                "A001.SZ,甲,2025,annual_report,https://x/a.pdf,a.pdf,aa,1\n"
                "A001.SZ,甲,2025,esg_report,https://x/e.pdf,e.pdf,bb,1\n",
                encoding="utf-8",
            )
            script = Path("scripts/build_official_website_source_queue.py").read_text(encoding="utf-8")
            script = script.replace(
                "ROOT = Path(__file__).resolve().parents[1]",
                f"ROOT = Path(r'{root}')",
            )
            runner = root / "build_queue.py"
            runner.write_text(script, encoding="utf-8")
            subprocess.check_call([sys.executable, str(runner)], cwd=str(root))
            with queue.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
            self.assertEqual(2, len(rows))
            self.assertTrue(all(row["domain_verification"] == "verified" for row in rows))
            self.assertTrue(all(row["official_domain"] == "issuer.example.com" for row in rows))
            self.assertTrue(all(row["scoring_authorized"] == "False" for row in rows))

    def test_domain_hygiene_rejects_platforms_and_truncation(self):
        from aegis_esg.domain_hygiene import is_plausible_issuer_domain
        self.assertFalse(is_plausible_issuer_domain("cninfo.com"))
        self.assertFalse(is_plausible_issuer_domain("roadshow.sseinfo.com"))
        self.assertFalse(is_plausible_issuer_domain("ir.p5w.net"))
        self.assertFalse(is_plausible_issuer_domain("ir.p"))
        self.assertFalse(is_plausible_issuer_domain("s.com"))
        self.assertFalse(is_plausible_issuer_domain("qyxxpl.ywzh.lnsthj.cn"))
        self.assertFalse(is_plausible_issuer_domain("www-app.gdeei.cn"))
        self.assertFalse(is_plausible_issuer_domain("hkexnews.hk"))
        self.assertTrue(is_plausible_issuer_domain("chinabaoan.com"))
        self.assertTrue(is_plausible_issuer_domain("sanyre.com.cn"))

    def test_ci_research_merge_preview_never_mutates_research(self):
        import importlib.util
        path = Path("scripts/build_ci_research_merge_preview.py")
        spec = importlib.util.spec_from_file_location("ci_research_merge_preview", path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(module)
        research = {
            ("000001.SZ", "2025", "annual"): {
                "company_code": "000001.SZ", "report_year": "2025",
                "document_type": "annual", "sha256": "aaa",
            },
            ("000002.SZ", "2025", "esg"): {
                "company_code": "000002.SZ", "report_year": "2025",
                "document_type": "esg", "sha256": "bbb",
            },
        }
        ci = {
            ("000001.SZ", "2025", "annual"): {
                "company_code": "000001.SZ", "company_name": "A",
                "report_year": "2025", "document_type": "annual",
                "sha256": "aaa", "source_url": "https://ex/a.pdf", "local_path": "a.pdf",
            },
            ("000002.SZ", "2025", "esg"): {
                "company_code": "000002.SZ", "company_name": "B",
                "report_year": "2025", "document_type": "esg",
                "sha256": "ccc", "source_url": "https://ex/b.pdf", "local_path": "b.pdf",
            },
            ("000003.SZ", "2025", "esg"): {
                "company_code": "000003.SZ", "company_name": "C",
                "report_year": "2025", "document_type": "esg",
                "sha256": "ddd", "source_url": "https://ex/c.pdf", "local_path": "c.pdf",
            },
        }
        preview, summary = module.build_merge_preview(research, ci)
        actions = {row["company_code"]: row["action"] for row in preview}
        self.assertNotIn("000001.SZ", actions)
        self.assertEqual("conflict_or_different_hash", actions["000002.SZ"])
        self.assertEqual("would_add", actions["000003.SZ"])
        self.assertEqual(1, summary["would_add"])
        self.assertEqual(1, summary["conflict_or_different_hash"])
        self.assertFalse(summary["research_index_mutated"])
        self.assertFalse(summary["scoring_authorized"])
        self.assertEqual(research[("000001.SZ", "2025", "annual")]["sha256"], "aaa")

    def test_live_collection_status_script_is_side_effect_safe(self):
        script = Path("scripts/refresh_live_collection_status.py").read_text(encoding="utf-8")
        compact = script.replace(" ", "")
        self.assertIn('"scoring_authorized":False', compact)
        self.assertIn("不启动下载", script)
        self.assertIn("collection_lock_held", script)
        self.assertNotIn("collect_batch", script)
        self.assertNotIn("run_scheduled_collection", script)

    def test_collection_failure_classification_prioritizes_partial_timeouts(self):
        import importlib.util
        path = Path("scripts/classify_collection_failures.py")
        spec = importlib.util.spec_from_file_location("classify_collection_failures", path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(module)
        kind, retryable, action = module.classify_error(
            "curl退出码28; 已保留42499461字节分片; Operation timed out after 180005 milliseconds"
        )
        self.assertEqual("timeout_partial_resume", kind)
        self.assertTrue(retryable)
        self.assertEqual("resume_with_longer_budget", action)
        kind, retryable, action = module.classify_error(
            '公开文档不是有效PDF; response="<html><script>\\n var arg1=\'ABC\';"'
        )
        self.assertEqual("exchange_antibot_html", kind)
        self.assertTrue(retryable)

    def test_aegis_locks_reclaim_dead_pid_and_keep_live_hint(self):
        import importlib.util
        import os
        import sys
        import tempfile
        path = Path("scripts/aegis_locks.py")
        spec = importlib.util.spec_from_file_location("aegis_locks", path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        sys.modules[spec.name] = module
        spec.loader.exec_module(module)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            dead = root / "aegisesp-text-extraction.lock"
            self.assertTrue(module.acquire_lock(dead, pid=os.getpid()))
            # Simulate crash: rewrite pid to a non-existent process.
            (dead / "pid").write_text("99999999\n", encoding="utf-8")
            status = module.lock_status(dead)
            self.assertTrue(status.stale)
            self.assertTrue(status.reclaimable)
            self.assertTrue(module.reclaim_lock(dead))
            self.assertFalse(dead.exists())
            # Fresh acquire works after reclaim.
            self.assertTrue(module.acquire_lock(dead))
            module.release_lock(dead)
            self.assertFalse(dead.exists())

    def test_ci_incremental_coverage_packet_never_authorizes_scoring(self):
        script = Path("scripts/build_ci_incremental_coverage_packet.py").read_text(encoding="utf-8")
        self.assertIn('"scoring_authorized": False', script)
        self.assertIn('"formal_publishable": False', script)
        self.assertIn("review_required", script)
        self.assertNotIn("apply-governance-benchmarks", script)
        self.assertNotIn("formal_rank_fixed", script)

    def test_read_document_index_keeps_rows_with_empty_source_url(self):
        from aegis_esg.collector import _read_document_index
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "index.csv"
            path.write_text(
                "company_code,company_name,report_year,document_type,source_url,retrieval_url,local_path,sha256,size\n"
                "A001.SZ,甲,2025,annual_report,,,a.pdf,aa,1\n"
                "A002.SZ,乙,2025,esg_report,,,e.pdf,bb,1\n"
                "A003.SZ,丙,0,esg_report,,,bad.pdf,cc,1\n",
                encoding="utf-8",
            )
            rows = _read_document_index(path)
            self.assertEqual(2, len(rows))
            self.assertIn("local://A001.SZ/2025/annual_report", rows)
            self.assertIn("local://A002.SZ/2025/esg_report", rows)

    def test_collector_http_referers_and_curl_fallback_helpers(self):
        from aegis_esg.collector import (
            _http_curl_max_time, _http_urlopen_timeout, _referer_for, _szse_curl_max_time,
        )
        self.assertIn("hkexnews", _referer_for("https://www1.hkexnews.hk/a.pdf"))
        self.assertIn("sse.com.cn", _referer_for("https://www.sse.com.cn/a.pdf"))
        self.assertIn("szse.cn", _referer_for("https://disc.static.szse.cn/a.pdf"))
        self.assertIn("cninfo.com.cn", _referer_for("https://static.cninfo.com.cn/a.pdf"))
        self.assertGreaterEqual(_http_curl_max_time(), 30)
        self.assertGreaterEqual(_http_urlopen_timeout(), 10)
        self.assertGreaterEqual(_szse_curl_max_time(), 600)

    def test_collection_coverage_reports_identity_gaps(self):
        script = Path("scripts/build_collection_coverage_report.py").read_text(encoding="utf-8")
        self.assertIn("missing_identities", script)
        self.assertIn("identity_coverage_rate", script)
        self.assertIn("redundant_url_gaps", script)
        retry = Path("scripts/build_collection_retry_manifest.py").read_text(encoding="utf-8")
        self.assertIn("skipped_redundant_url_gaps", retry)
        self.assertIn("true identity", retry.lower().replace("真实身份", "true identity"))

    def test_identity_gap_fill_uses_cninfo_alternate_channel(self):
        script = Path("scripts/fill_identity_gaps_alternate_sources.py").read_text(encoding="utf-8")
        self.assertIn("downloaded_from_cninfo", script)
        self.assertIn("find_disclosure_pdf", script)
        module = Path("src/aegis_esg/sources/cninfo.py").read_text(encoding="utf-8")
        self.assertIn("hisAnnouncement/query", module)
        self.assertIn("static.cninfo.com.cn", module)

    def test_source_authority_prefers_exchange_over_issuer_and_other(self):
        from aegis_esg.source_authority import SourceTier, disclosure_authority, prefer, source_tier
        exchange = Observation(
            "A", "甲", 2025, "Q_G_ROE", 10.0, ValueStatus.CONFIRMED,
            "https://www.sse.com.cn/a.pdf", "data/raw/ci_collection/A/2025/annual_report.pdf",
            1, "净资产收益率 10", 0.9,
        )
        issuer = Observation(
            "A", "甲", 2025, "Q_G_ROE", 11.0, ValueStatus.CONFIRMED,
            "https://www.example-energy.com/esg.pdf", "data/raw/issuer_site/A/2025/esg_report.pdf",
            1, "官网披露 净资产收益率 11", 0.9,
        )
        other = Observation(
            "A", "甲", 2025, "Q_G_ROE", 12.0, ValueStatus.CONFIRMED,
            "https://finance.sina.com.cn/x", "data/cache/third_party.txt",
            1, "第三方转载 12", 0.5,
        )
        self.assertEqual(SourceTier.EXCHANGE, source_tier(exchange))
        self.assertEqual(SourceTier.ISSUER_WEBSITE, source_tier(issuer))
        self.assertEqual(SourceTier.OTHER, source_tier(other))
        self.assertLess(disclosure_authority(exchange), disclosure_authority(issuer))
        self.assertLess(disclosure_authority(issuer), disclosure_authority(other))
        self.assertEqual(10.0, prefer(issuer, exchange).value)
        script = Path("scripts/fill_missing_from_authoritative_sources.py").read_text(encoding="utf-8")
        self.assertIn("exchange", script.lower())
        self.assertIn("issuer", script.lower())
        self.assertIn("False", Path("src/aegis_esg/research_qualitative.py").read_text(encoding="utf-8"))

    def test_pdf_algorithm_alignment_and_yoy_soft_check_scripts(self):
        align = Path("scripts/build_pdf_algorithm_alignment_audit.py").read_text(encoding="utf-8")
        self.assertIn("g_sasac_benchmark", align)
        self.assertIn("formal_release_authorized", align)
        yoy = Path("scripts/compare_client_ranking_yoy.py").read_text(encoding="utf-8")
        self.assertIn("client_top50_to_outside_our_top200", yoy)
        self.assertIn("不要求严格对齐", yoy)
        bench = Path("data/methodologies/governance_benchmarks_from_client_report_2024.csv")
        self.assertTrue(bench.is_file())
        research_m = Path("data/methodologies/energy_esg_2025_research_sasac.json")
        if research_m.is_file():
            payload = json.loads(research_m.read_text(encoding="utf-8"))
            self.assertEqual("ENERGY-ESG-2025-RESEARCH-SASAC-v1", payload.get("version"))
            self.assertNotEqual("DLT2971-2025-v1", payload.get("version"))
            g = [i for i in payload["indicators"] if i["code"].startswith("Q_G_")]
            self.assertTrue(all(i.get("benchmark") is not None for i in g))
        report = Path("output/audit/pdf_algorithm_alignment_v1.json")
        if report.is_file():
            body = json.loads(report.read_text(encoding="utf-8"))
            self.assertFalse(body.get("formal_release_authorized"))

    def test_research_ranking_refresh_stays_non_formal(self):
        script = Path("scripts/run_research_ranking_refresh.py").read_text(encoding="utf-8")
        self.assertIn('--mode", "research"', script)
        self.assertIn("scoring_authorized_formal", script)
        self.assertIn("False", script)
        self.assertIn("full_auto_v21_exchange_zero", script)
        self.assertIn("legacy_zero_v1", script)
        self.assertIn("exchange-key-accept-v1", script)
        self.assertIn("prefer_exchange_chinese_direct_then_ci_harvest", script)
        self.assertNotIn("--release", script)
        meta = Path("output/research/2025/full_auto_v21_exchange_zero/ranking_metadata.json")
        if meta.is_file():
            payload = json.loads(meta.read_text(encoding="utf-8"))
            self.assertEqual("research", payload.get("ranking_mode"))
            self.assertFalse(payload.get("official_release"))
            self.assertEqual("legacy_zero_v1", payload.get("missing_strategy_version"))

        import importlib.util
        module_path = Path("scripts/run_research_ranking_refresh.py")
        spec = importlib.util.spec_from_file_location("run_research_ranking_refresh", module_path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(module)
        chinese = Observation(
            "A", "甲", 2025, "Q_G_ROE", 10.0, ValueStatus.CONFIRMED,
            "https://www.sse.com.cn/a.pdf", "data/raw/ci_collection/A/2025/annual_report.pdf",
            1, "加权平均净资产收益率(%) 10.85", 0.9,
        )
        english = Observation(
            "A", "甲", 2025, "Q_G_ROE", 11.0, ValueStatus.CONFIRMED,
            "https://www.sse.com.cn/a.pdf", "data/raw/ci_collection/A/2025/annual_report.pdf",
            1, "English consolidated statements derived: ROE", 0.9,
        )
        self.assertLess(module.disclosure_authority(chinese), module.disclosure_authority(english))
        self.assertTrue(module.is_exchange_source(chinese))
        self.assertEqual(
            MissingStrategy.LEGACY_ZERO_V1,
            validate_ranking_mode("research", [], None),
        )

    def test_truncated_text_repair_and_scan_fallback_packets(self):
        repair = Path("scripts/repair_truncated_ci_text_exports.py").read_text(encoding="utf-8")
        self.assertIn('"scoring_authorized": False', repair)
        self.assertIn('"ocr_authorized": False', repair)
        self.assertIn("page_markers_before", repair)
        extract = Path("scripts/run_ci_text_extraction.py").read_text(encoding="utf-8")
        self.assertIn("repair_truncated_ci_text_exports.py", extract)
        self.assertIn("truncated_repaired", extract)
        swift = Path("scripts/extract_pdf_batch.swift").read_text(encoding="utf-8")
        self.assertIn("resolvingSymlinksInPath", swift)
        self.assertIn("relativePath(of:", swift)
        fallback = Path("scripts/build_scan_esg_annual_fallback_packet.py").read_text(encoding="utf-8")
        self.assertIn('"ocr_authorized": False', fallback)
        self.assertIn("candidates_from_annual", fallback)
        api = Path("src/aegis_esg/api.py").read_text(encoding="utf-8")
        self.assertIn("/demo/scan-esg-annual-fallback", api)

    def test_ci_thin_text_packet_never_authorizes_ocr_or_scoring(self):
        script = Path("scripts/build_ci_thin_text_packet.py").read_text(encoding="utf-8")
        self.assertIn('"scoring_authorized": False', script)
        self.assertIn('"ocr_authorized": False', script)
        self.assertIn('"formal_publishable": False', script)
        self.assertIn("critical_thin", script)
        self.assertIn("prefer_annual_embedded_evidence", script)
        self.assertIn("await_ocr_authorization", script)
        self.assertIn("truncated_export", script)
        self.assertNotIn("apply-governance-benchmarks", script)
        api = Path("src/aegis_esg/api.py").read_text(encoding="utf-8")
        self.assertIn("/demo/ci-thin-text", api)
        live = Path("scripts/refresh_live_collection_status.py").read_text(encoding="utf-8")
        self.assertIn("build_ci_thin_text_packet.py", live)

        import importlib.util
        module_path = Path("scripts/build_ci_thin_text_packet.py")
        spec = importlib.util.spec_from_file_location("build_ci_thin_text_packet", module_path)
        module = importlib.util.module_from_spec(spec)
        assert spec.loader is not None
        spec.loader.exec_module(module)
        self.assertEqual("critical_thin", module.classify(0, 10_000_000, missing=False, page_markers=40))
        self.assertEqual("thin", module.classify(800, 100_000, missing=False, page_markers=10))
        self.assertEqual("truncated_export", module.classify(300, 5_000_000, missing=False, page_markers=1))
        self.assertEqual("large_pdf_low_text", module.classify(3000, 8_000_000, missing=False, page_markers=40))
        self.assertIsNone(module.classify(20_000, 8_000_000, missing=False, page_markers=40))
        self.assertEqual("missing_text", module.classify(0, 0, missing=True))
        self.assertEqual(
            "prefer_annual_embedded_evidence",
            module._triage_action(
                klass="critical_thin", document_type="esg_report",
                sibling_annual_non_ws=50_000, sibling_esg_non_ws=0, candidate_count=12,
            ),
        )
        self.assertEqual(
            "await_ocr_authorization",
            module._triage_action(
                klass="critical_thin", document_type="esg_report",
                sibling_annual_non_ws=100, sibling_esg_non_ws=0, candidate_count=0,
            ),
        )

    def test_cninfo_fallback_helpers_and_collect_batch(self):
        from aegis_esg.sources.cninfo import find_disclosure_pdf, should_try_cninfo_fallback
        self.assertTrue(
            should_try_cninfo_fallback(
                "https://disc.static.szse.cn/a.PDF",
                "公开文档不是有效PDF: <!DOCTYPE html> Page Verification",
            )
        )
        self.assertFalse(
            should_try_cninfo_fallback("https://www.sse.com.cn/a.pdf", "timeout")
        )

        calls = {"n": 0}

        def fake_fetcher(request):
            calls["n"] += 1
            url = request.full_url
            if "topSearch" in url:
                return json.dumps([
                    {"code": "300073", "orgId": "9900011167", "zwjc": "当升科技"},
                ]).encode()
            return json.dumps({
                "announcements": [{
                    "announcementTitle": "2025年度可持续发展报告",
                    "adjunctUrl": "finalpage/2026-03-31/1225057154.PDF",
                }]
            }).encode()

        hit = find_disclosure_pdf("300073.SZ", 2025, "esg_report", fetcher=fake_fetcher, pause_seconds=0)
        self.assertIsNotNone(hit)
        self.assertIn("1225057154.PDF", hit[1])

        body = b"%PDF-1.7\n" + b"z" * 10_000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            index, failures = root / "index.csv", root / "failures.csv"
            write_document_index(index, [])
            manifest = root / "manifest.csv"
            manifest.write_text(
                "company_code,company_name,report_year,document_type,source_url\n"
                "B,乙,2025,esg_report,https://disc.static.szse.cn/need.PDF\n",
                encoding="utf-8",
            )
            downloads: list[str] = []

            def fake_download(url):
                downloads.append(url)
                if "disc.static.szse.cn" in url:
                    raise ValueError("公开文档不是有效PDF: Page Verification")
                return body, url

            with patch("aegis_esg.collector._download_pdf", side_effect=fake_download), patch(
                "aegis_esg.collector.find_disclosure_pdf",
                return_value=("2025年度可持续发展报告", "https://static.cninfo.com.cn/alt.PDF"),
            ):
                rows, errors = collect_batch(
                    manifest, root / "raw", index, failures,
                    delay_seconds=0, reuse_existing=True, workers=1, preserve_index=True,
                )
            self.assertFalse(errors)
            self.assertEqual(
                ["https://disc.static.szse.cn/need.PDF", "https://static.cninfo.com.cn/alt.PDF"],
                downloads,
            )
            self.assertEqual(1, len(rows))
            self.assertIn("cninfo.com.cn", rows[0].source_url)
            self.assertIn("cninfo_fallback", rows[0].retrieval_url)

    def test_governance_benchmark_audit_and_apply_freeze(self):
        from aegis_esg.benchmarks import (
            apply_governance_benchmarks, audit_governance_benchmarks, require_governance_benchmarks,
        )
        audit = audit_governance_benchmarks(self.methodology)
        self.assertEqual(17, audit["governance_indicator_count"])
        self.assertEqual(17, audit["missing_count"])
        self.assertFalse(audit["formal_ready"])
        with self.assertRaisesRegex(ValueError, "17项优秀值"):
            require_governance_benchmarks(self.methodology)
        template = Path("data/methodologies/governance_benchmarks_template_2025.csv")
        self.assertTrue(template.is_file())
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            table = root / "benchmarks.csv"
            with template.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
            for index, row in enumerate(rows):
                row["benchmark"] = str(10 + index)
            with table.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=rows[0].keys(), lineterminator="\n")
                writer.writeheader()
                writer.writerows(rows)
            output = root / "energy_esg_dlt2971_v1.json"
            report = apply_governance_benchmarks(
                "data/methodologies/energy_esg_2025.json", table, output_path=output,
            )
            self.assertTrue(report["formal_ready"])
            self.assertEqual("DLT2971-2025-v1", report["written_version"])
            frozen = load_methodology(output)
            self.assertEqual("DLT2971-2025-v1", frozen.version)
            self.assertEqual(17, audit_governance_benchmarks(frozen)["filled_count"])
            self.assertIsNotNone(frozen.by_code["Q_G_ROE"].benchmark)
            # Incomplete table cannot freeze by default.
            partial = root / "partial.csv"
            partial.write_text(
                "indicator_code,benchmark\nQ_G_ROE,12.5\n", encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "优秀值"):
                apply_governance_benchmarks(
                    "data/methodologies/energy_esg_2025.json", partial,
                    output_path=root / "draft.json",
                )

    def test_dlt_alignment_status_reports_benchmark_blocker(self):
        from aegis_esg.dlt_alignment import build_dlt_alignment_status
        report = build_dlt_alignment_status("data/methodologies/energy_esg_2025.json")
        self.assertEqual("DL/T 2971—2025", report["standard_ref"])
        self.assertEqual(6, report["check_count"])
        self.assertEqual(4, report["ready_count"])
        self.assertFalse(report["aligned"])
        self.assertFalse(report["checks"]["governance_benchmarks"]["ready"])
        self.assertFalse(report["checks"]["formal_methodology_frozen"]["ready"])
        self.assertTrue(report["checks"]["indicator_coverage"]["ready"])
        self.assertTrue(report["checks"]["grade_mapping"]["ready"])

    def test_prepare_governance_benchmark_packet_lists_mapping_risks(self):
        from aegis_esg.benchmarks import prepare_governance_benchmark_packet
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            summary = prepare_governance_benchmark_packet(
                "data/methodologies/energy_esg_2025.json",
                csv_path=root / "intake.csv",
                html_path=root / "packet.html",
                summary_path=root / "summary.json",
            )
            self.assertEqual(17, summary["row_count"])
            self.assertEqual(0, summary["filled_count"])
            self.assertEqual(4, summary["mapping_risk_count"])
            self.assertTrue((root / "intake.csv").is_file())
            html_text = (root / "packet.html").read_text(encoding="utf-8")
            self.assertIn("已获利息倍数", html_text)
            self.assertIn("需口径确认", html_text)
            with (root / "intake.csv").open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
            self.assertEqual(17, len(rows))
            self.assertIn("sasac_name", rows[0])
            self.assertEqual("已获利息倍数", next(r["sasac_name"] for r in rows if r["indicator_code"] == "Q_G_EBITDA_INTEREST"))

    def test_review_impact_prioritizes_conflict_crossing_rank_boundary(self):
        indicators = self.methodology.quantitative[:2]
        tiers = [
            ReviewTier("A", "甲", 2025, indicators[0].code, 2, "1|2", "1|2", .9,
                       "manual_signature_required", "review_conflict_candidates", "conflicting_values"),
            ReviewTier("B", "乙", 2025, indicators[1].code, 1, "3", "3", .8,
                       "single_candidate_review", "manual_spot_check", "not_auto"),
        ]
        sensitivity = {
            "companies": [
                {"company_code": "A", "best_rank": 150, "worst_rank": 250,
                 "rank_span": 100, "ranks": {"indicator_neutral_v1": 190}},
                {"company_code": "B", "best_rank": 400, "worst_rank": 410,
                 "rank_span": 10, "ranks": {"indicator_neutral_v1": 405}},
            ]
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "sensitivity.json"
            path.write_text(json.dumps(sensitivity), encoding="utf-8")
            tasks, summary = prioritize_review_by_impact(tiers, path, self.methodology)
        self.assertEqual("A", tasks[0].company_code)
        self.assertTrue(tasks[0].crosses_top_200)
        self.assertGreater(tasks[0].impact_score, tasks[1].impact_score)
        self.assertEqual(1, summary["crosses_boundary_count"])

    def test_evidence_constraint_graph_is_stable_and_detects_conflicts(self):
        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 1, ValueStatus.PENDING,
                        "https://example.test/a.pdf", "a.pdf", 3, "证据一", .9),
            Observation("A", "甲", 2025, indicator.code, 2, ValueStatus.PENDING,
                        "https://example.test/a.pdf", "a.pdf", 4, "证据二", .9),
        ]
        first, summary = build_evidence_constraint_graph(observations, self.methodology)
        second, _ = build_evidence_constraint_graph(observations, self.methodology)
        self.assertEqual(first, second)
        self.assertEqual(1, summary["conflicting_group_count"])
        self.assertEqual(2, summary["node_kind_counts"]["candidate"])
        self.assertTrue(any(edge["relation"] == "member_of" for edge in first["edges"]))

    def test_evidence_graph_keeps_company_identity_stable_across_name_variants(self):
        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲公司", 2025, indicator.code, 1, source_file="a.pdf"),
            Observation("A", "甲", 2025, indicator.code, 1, source_file="a.pdf"),
        ]
        graph, summary = build_evidence_constraint_graph(observations, self.methodology)
        companies = [item for item in graph["nodes"] if item["kind"] == "company"]
        self.assertEqual(1, len(companies))
        self.assertEqual(1, summary["node_kind_counts"]["company"])

    def test_e1_validation_requires_signed_truth_and_compares_baseline(self):
        indicator = self.methodology.quantitative[0]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 1, ValueStatus.PENDING,
                        source_file="a.pdf", source_page=1, evidence_text="证据", confidence=.9),
            Observation("A", "甲", 2025, indicator.code, 2, ValueStatus.PENDING,
                        source_file="a.pdf", source_page=2, evidence_text="错误", confidence=.9),
            Observation("B", "乙", 2025, indicator.code, 3, ValueStatus.PENDING,
                        source_file="b.pdf", source_page=1, evidence_text="证据", confidence=.9),
        ]
        graph, graph_summary = build_evidence_constraint_graph(observations, self.methodology)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            graph_path = root / "graph.json"
            graph_path.write_text(json.dumps(graph), encoding="utf-8")
            rows, summary = prepare_e1_validation_sample(graph_path, 1)
            validation = root / "validation.csv"
            write_e1_validation_sample(validation, root / "summary.json", rows, summary)
            with self.assertRaisesRegex(ValueError, "缺少true/false"):
                evaluate_e1_validation(validation)
            with validation.open(encoding="utf-8-sig", newline="") as stream:
                raw = list(csv.DictReader(stream))
            for index, row in enumerate(raw):
                row["ground_truth_valid"] = "true" if index == len(raw) - 1 else "false"
                row["reviewer"] = "审核员"
                row["reviewed_at"] = "2026-08-03T12:00:00+08:00"
                row["review_note"] = "人工核验PDF原页"
            with validation.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=raw[0].keys(), lineterminator="\n")
                writer.writeheader(); writer.writerows(raw)
            report = evaluate_e1_validation(validation)
        self.assertTrue(report["applicable"])
        self.assertEqual(len(raw), report["labeled_count"])
        self.assertIn("precision_improvement", report)

    def test_quantitative_auto_decision_validation_is_stratified_signed_and_bound(self):
        indicators = self.methodology.quantitative[:2]
        observations = [
            Observation("A", "甲", 2025, indicators[0].code, 1, ValueStatus.CONFIRMED,
                        source_file="a.pdf", source_page=1, evidence_text="中文跨表派生: 总量与营收", confidence=.9),
            Observation("B", "乙", 2025, indicators[0].code, 2, ValueStatus.CONFIRMED,
                        source_file="b.pdf", source_page=2, evidence_text="表格直接披露", confidence=.95),
            Observation("C", "丙", 2025, indicators[1].code, 3, ValueStatus.CONFIRMED,
                        source_file="c.pdf", source_page=3, evidence_text="direct disclosure", confidence=.95),
        ]
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            confirmed = root / "confirmed.csv"
            write_observations(confirmed, observations)
            rows, summary = prepare_quantitative_validation_sample(confirmed, 1)
            self.assertEqual(3, summary["stratum_count"])
            self.assertEqual(2, summary["indicator_count"])
            sample = root / "sample.csv"
            manifest = root / "sample.json"
            write_quantitative_validation_sample(sample, manifest, rows, summary)
            with self.assertRaisesRegex(ValueError, "缺少true/false"):
                evaluate_quantitative_validation(sample)
            with sample.open(encoding="utf-8-sig", newline="") as stream:
                raw = list(csv.DictReader(stream))
            for row in raw:
                row["ground_truth_valid"] = "true"
                row["ground_truth_value"] = row["value"]
                row["reviewer"] = "真实审核员"
                row["reviewed_at"] = "2026-08-03T15:00:00+08:00"
                row["review_note"] = "核验PDF原页和计算口径"
            with sample.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=raw[0].keys(), lineterminator="\n")
                writer.writeheader(); writer.writerows(raw)
            evaluation = evaluate_quantitative_validation(sample, manifest_path=manifest)
            self.assertTrue(evaluation["sampling_accuracy_passed"])
            evaluation_path = root / "evaluation.json"
            evaluation_path.write_text(json.dumps(evaluation), encoding="utf-8")
            coverage_path = root / "coverage.json"
            coverage_path.write_text(json.dumps({"quantitative_indicator_count": 2}), encoding="utf-8")
            applied = apply_quantitative_validation(coverage_path, evaluation_path, confirmed)
            self.assertTrue(applied["sampling_accuracy_passed"])

    def test_quantitative_validation_rejects_partial_indicator_coverage(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            (root / "coverage.json").write_text(json.dumps({"quantitative_indicator_count": 37}), encoding="utf-8")
            (root / "evaluation.json").write_text(json.dumps({
                "validation_version": "quantitative-auto-decision-validation-v1",
                "applicable": True, "indicator_count": 0, "sampling_accuracy_passed": True,
                "sample_complete": True,
            }), encoding="utf-8")
            confirmed = root / "confirmed.csv"
            confirmed.write_text("x", encoding="utf-8")
            evaluation = json.loads((root / "evaluation.json").read_text())
            evaluation["confirmed_input_sha256"] = hashlib.sha256(confirmed.read_bytes()).hexdigest()
            (root / "evaluation.json").write_text(json.dumps(evaluation), encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "未覆盖自动决定"):
                apply_quantitative_validation(root / "coverage.json", root / "evaluation.json", confirmed)

    def test_dependency_graph_limits_recompute_to_affected_indicator_population(self):
        indicator, other = self.methodology.quantitative[:2]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 1, source_file="a.pdf"),
            Observation("B", "乙", 2025, indicator.code, 2, source_file="b.pdf"),
            Observation("C", "丙", 2025, other.code, 3, source_file="c.pdf"),
        ]
        graph, _ = build_evidence_constraint_graph(observations, self.methodology)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "graph.json"
            path.write_text(json.dumps(graph), encoding="utf-8")
            report = plan_incremental_recompute(path, changed_documents=["a.pdf"])
        self.assertTrue(report["applicable"])
        self.assertEqual(64, len(report["graph_sha256"]))
        self.assertEqual(1, report["seed_candidate_count"])
        self.assertEqual(1, report["affected_indicator_count"])
        self.assertEqual(2, report["affected_company_count"])
        self.assertGreater(report["company_score_reduction_rate"], 0)

    def test_cached_recompute_is_field_equivalent_to_full_scoring(self):
        indicator, other = self.methodology.quantitative[:2]
        observations = [
            Observation("A", "甲", 2025, indicator.code, 1),
            Observation("B", "乙", 2025, indicator.code, 2),
            Observation("A", "甲", 2025, other.code, 3),
            Observation("B", "乙", 2025, other.code, 4),
            Observation("C", "丙", 2025, other.code, 5),
        ]
        engine = ScoringEngine(self.methodology)
        full = engine.evaluate(observations, MissingStrategy.INDICATOR_NEUTRAL_V1)
        cache = engine.build_cache(observations, MissingStrategy.INDICATOR_NEUTRAL_V1)
        incremental = engine.evaluate_from_cache(cache, {indicator.code})
        self.assertEqual(
            [item.to_dict() for item in full],
            [item.to_dict() for item in incremental],
        )

        benchmark = benchmark_incremental_scoring(
            engine, observations, {indicator.code},
            MissingStrategy.INDICATOR_NEUTRAL_V1, repetitions=1,
        )
        self.assertTrue(benchmark["field_equivalent"])
        self.assertEqual(benchmark["full_output_sha256"], benchmark["incremental_output_sha256"])

        changed = replace(observations[0], value=9.5, evidence_text="signed simulation")
        dynamic, audit = engine.apply_cache_changes(cache, [changed])
        expected = engine.evaluate(
            [changed, *observations[1:]], MissingStrategy.INDICATOR_NEUTRAL_V1,
        )
        self.assertTrue(audit["committed"])
        self.assertEqual(
            [item.to_dict() for item in expected],
            [item.to_dict() for item in dynamic],
        )

    def test_qualitative_gap_priority_favors_boundary_high_weight_gap(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            gaps = root / "gaps.csv"; sensitivity = root / "sensitivity.json"; coverage = root / "coverage.csv"
            gaps.write_text(
                "company_code,company_name,report_year,indicator_code,indicator_name,indicator_weight,priority,status,next_action\n"
                "A,甲,2025,X1,高权重,10,1,evidence_missing,collect\n"
                "B,乙,2025,X2,低权重,1,1,evidence_missing,collect\n", encoding="utf-8",
            )
            sensitivity.write_text(json.dumps({"companies": [
                {"company_code": "A", "best_rank": 190, "worst_rank": 210, "rank_span": 20},
                {"company_code": "B", "best_rank": 500, "worst_rank": 510, "rank_span": 10},
            ]}), encoding="utf-8")
            coverage.write_text("stock_code,esg_status\nA,missing\nB,collected\n", encoding="utf-8")
            rows, summary = prioritize_qualitative_gaps(gaps, sensitivity, coverage)
        self.assertEqual("A", rows[0]["company_code"])
        self.assertEqual(1, summary["crosses_boundary_count"])

    def test_embedded_esg_coverage_preserves_completion_contract(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); coverage = root / "coverage.csv"; evidence = root / "evidence.csv"
            coverage.write_text(
                "stock_code,annual_status,esg_status,next_action\nA,collected,missing,discover_esg_report\n",
                encoding="utf-8",
            )
            evidence.write_text("company_code\nA\n", encoding="utf-8")
            rows, summary = recognize_embedded_esg_coverage(coverage, evidence)
        self.assertEqual("embedded_in_annual", rows[0]["esg_status"])
        self.assertEqual(1, summary["annual_coverage_count"])
        self.assertEqual(1, summary["esg_coverage_count"])

    def test_quantitative_gap_priority_favors_key_boundary_gap(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); tasks = root / "tasks.csv"; sensitivity = root / "sensitivity.json"
            key, other = self.methodology.quantitative[:2]
            fields = "company_code,company_name,indicator_code,indicator_name,dimension,key_indicator,candidate_count,report_years,source_pages,max_confidence,status,next_action,priority\n"
            tasks.write_text(fields +
                f"A,甲,{key.code},{key.name},E,True,0,,,0,missing_candidate,extend_extraction_rules,0\n" +
                f"B,乙,{other.code},{other.name},E,False,0,,,0,missing_candidate,extend_extraction_rules,2\n",
                encoding="utf-8")
            sensitivity.write_text(json.dumps({"companies": [
                {"company_code": "A", "best_rank": 190, "worst_rank": 210, "rank_span": 20},
                {"company_code": "B", "best_rank": 500, "worst_rank": 510, "rank_span": 10},
            ]}), encoding="utf-8")
            rows, summary = prioritize_quantitative_gaps(tasks, sensitivity, self.methodology)
        self.assertEqual("A", rows[0]["company_code"])
        self.assertEqual(1, summary["key_indicator_gap_count"])

    def test_thin_population_batch_balances_indicators_below_gate(self):
        with tempfile.TemporaryDirectory() as directory:
            impact = Path(directory) / "impact.csv"
            impact.write_text(
                "impact_rank,impact_score,company_code,company_name,indicator_code,indicator_population,status\n"
                "1,99,A,甲,Q1,1,missing_candidate\n"
                "2,98,B,乙,Q1,1,missing_candidate\n"
                "3,97,C,丙,Q2,19,missing_candidate\n"
                "4,96,D,丁,Q3,20,missing_candidate\n",
                encoding="utf-8",
            )
            rows, summary = build_thin_population_gap_batch(impact, 20, 1)
        self.assertEqual(["Q1", "Q2"], sorted(item["indicator_code"] for item in rows))
        self.assertEqual({"Q1": 1, "Q2": 1}, summary["selected_counts"])
        self.assertEqual(2, summary["thin_indicator_count"])
        self.assertEqual(19, rows[0]["population_deficit"])

    def test_observation_revision_keeps_audit_history(self):
        with tempfile.TemporaryDirectory() as directory:
            repo = SQLiteRepository(Path(directory) / "audit.db")
            repo.initialize()
            first = Observation("A", "甲公司", 2024, "Q_E_GHG_INTENSITY", 10)
            second = Observation("A", "甲公司", 2024, "Q_E_GHG_INTENSITY", 8)
            repo.upsert_observations([first])
            repo.upsert_observations([second])
            latest = repo.confirmed_observations(2024)
            self.assertEqual(1, len(latest))
            self.assertEqual(8, latest[0].value)
            count = repo.connection.execute("SELECT COUNT(*) FROM observation").fetchone()[0]
            self.assertEqual(2, count)

    def test_sse_official_response_classification(self):
        payload = '{"result":[' \
            '{"SECURITY_CODE":"600900","SECURITY_NAME":"长江电力","SSEDATE":"2026-04-30",' \
            '"TITLE":"长江电力2025年年度报告","URL":"/annual.pdf"},' \
            '{"SECURITY_CODE":"600900","SECURITY_NAME":"长江电力","SSEDATE":"2026-04-30",' \
            '"TITLE":"长江电力2025年度环境、社会和公司治理报告","URL":"/esg.pdf"}' \
            ']}'
        parsed = parse_response(payload)
        self.assertEqual("https://www.sse.com.cn/annual.pdf", parsed[0].source_url)
        self.assertEqual("annual_report", classify_title(parsed[0].title, "2025"))
        self.assertEqual("esg_report", classify_title(parsed[1].title, "2025"))
        self.assertIsNone(classify_title("长江电力2025年年度报告披露提示性公告", "2025"))
        self.assertEqual("annual_report", classify_title("中国石油天然气股份有限公司2025年年报", "2025"))
        self.assertEqual("esg_report", classify_title("中国石油天然气股份有限公司2025年度环境、社会和治理报告", "2025"))

        def fetcher(_request):
            return payload.encode()
        reports = discover_reports("600900.SH", 2025, fetcher=fetcher)
        self.assertEqual({"annual_report", "esg_report"}, {item.document_type for item in reports})

    def test_report_title_policy_rejects_non_report_announcements(self):
        annual_lookalikes = [
            "岳阳兴长：关于举办2025年年度报告及2026年一季度报告网上业绩说明会的公告",
            "豪鹏科技：关于举行2025年年度报告网上业绩说明会的通知",
            "思源电气：关于举行2025年年度报告网上说明会的通知",
            "麦格米特：关于举行2025年度报告网上业绩说明会的公告（更正后）",
            "蓝天燃气关于2025年年度报告更正的公告",
            "关于公司自愿披露2025年年度报告（简版）及可持续发展报告英文版本的公告",
            "阳光电源：关于2025年年度报告（英文简版）的自愿性披露公告",
            "某公司2025年半年度报告",
            "某公司2025年年度报告摘要",
        ]
        for title in annual_lookalikes:
            with self.subTest(title=title):
                self.assertIsNone(classify_title(title, "2025"))
                self.assertIsNone(classify_szse_title(title, "2025"))
                self.assertIsNone(classify_disclosure_title(title, "2025"))
        self.assertEqual("annual_report", classify_title("石化油服2025年年度报告全文（修订稿）", "2025"))
        self.assertEqual("annual_report", classify_title("电气风电2025年年度报告（修订版）", "2025"))

    def test_report_selection_prefers_chinese_full_text_over_later_variants(self):
        payload = json.dumps({"data": [
            {"secCode": "002459", "secName": "晶澳科技", "publishTime": "2026-04-29", "title": "晶澳科技：2025年年度报告", "attachPath": "/annual-cn.pdf"},
            {"secCode": "002459", "secName": "晶澳科技", "publishTime": "2026-06-27", "title": "晶澳科技：2025年年度报告英文版（2025 Annual Report）", "attachPath": "/annual-en.pdf"},
            {"secCode": "002459", "secName": "晶澳科技", "publishTime": "2026-05-07", "title": "晶澳科技：关于举行2025年年度报告网上业绩说明会的公告", "attachPath": "/briefing.pdf"},
        ], "announceCount": 3}, ensure_ascii=False)
        reports = discover_szse_reports("002459.SZ", 2025, fetcher=lambda request: payload.encode())
        annual = [item for item in reports if item.document_type == "annual_report"]
        self.assertEqual(1, len(annual))
        self.assertTrue(annual[0].source_url.endswith("annual-cn.pdf"))

    def test_report_selection_prefers_a_share_over_h_share_version(self):
        payload = '{"result":[' \
            '{"SECURITY_CODE":"600011","SECURITY_NAME":"华能国际","SSEDATE":"2026-03-20",' \
            '"TITLE":"华能国际2025年年度报告","URL":"/annual-a.pdf"},' \
            '{"SECURITY_CODE":"600011","SECURITY_NAME":"华能国际","SSEDATE":"2026-04-18",' \
            '"TITLE":"华能国际H股2025年度报告","URL":"/annual-h.pdf"}' \
            ']}'
        reports = discover_reports("600011.SH", 2025, fetcher=lambda request: payload.encode())
        annual = [item for item in reports if item.document_type == "annual_report"]
        self.assertEqual(1, len(annual))
        self.assertTrue(annual[0].source_url.endswith("annual-a.pdf"))

    def test_report_selection_keeps_english_when_no_chinese_version(self):
        payload = '{"result":[' \
            '{"SECURITY_CODE":"688005","SECURITY_NAME":"容百科技","SSEDATE":"2026-04-30",' \
            '"TITLE":"容百科技2025年年度报告（英文版）","URL":"/annual-en.pdf"}' \
            ']}'
        reports = discover_reports("688005.SH", 2025, fetcher=lambda request: payload.encode())
        annual = [item for item in reports if item.document_type == "annual_report"]
        self.assertEqual(1, len(annual))
        self.assertTrue(annual[0].source_url.endswith("annual-en.pdf"))

    def test_szse_official_response_classification(self):
        payload = json.dumps({"data": [
            {"secCode": "000027", "secName": "深圳能源", "publishTime": "2026-04-30 18:00:00", "title": "深圳能源2025年年度报告", "attachPath": "/disc/a.pdf"},
            {"secCode": ["000027"], "secName": ["深圳能源"], "publishTime": "2026-04-30", "title": "深圳能源2025年度可持续发展报告", "attachPath": "/disc/esg.pdf"},
        ]}, ensure_ascii=False)
        rows = parse_szse_response(payload)
        self.assertEqual("https://disc.static.szse.cn/disc/a.pdf", rows[0].source_url)
        self.assertEqual("000027.SZ", rows[1].stock_code)
        self.assertEqual("annual_report", classify_szse_title(rows[0].title, "2025"))
        self.assertIsNone(classify_szse_title("崧盛股份：2025年年度报告披露提示性公告", "2025"))
        self.assertIsNone(classify_szse_title("关于披露2025年度可持续发展报告的提示性公告", "2025"))
        reports = discover_szse_reports("000027.SZ", 2025, fetcher=lambda request: payload.encode())
        self.assertEqual({"annual_report", "esg_report"}, {item.document_type for item in reports})

    def test_szse_discovers_esg_from_general_paginated_channel(self):
        calls = []
        annual = {"data": [{"secCode": ["000027"], "secName": ["深圳能源"], "publishTime": "2026-04-30", "title": "深圳能源2025年年度报告", "attachPath": "/annual.pdf"}], "announceCount": 1}
        general_pages = {
            1: {"data": [{"secCode": ["000027"], "secName": ["深圳能源"], "publishTime": "2026-03-01", "title": "深圳能源2025年度社会责任报告", "attachPath": "/old-esg.pdf"}], "announceCount": 101},
            2: {"data": [{"secCode": ["000027"], "secName": ["深圳能源"], "publishTime": "2026-05-01", "title": "深圳能源2025年度可持续发展报告", "attachPath": "/new-esg.pdf"}], "announceCount": 101},
        }
        def fetcher(request):
            query = json.loads(request.data)
            calls.append(query)
            result = annual if query.get("bigCategoryId") else general_pages[query["pageNum"]]
            return json.dumps(result, ensure_ascii=False).encode()
        reports = discover_szse_reports("000027.SZ", 2025, fetcher=fetcher)
        by_type = {item.document_type: item for item in reports}
        self.assertTrue(by_type["esg_report"].source_url.endswith("new-esg.pdf"))
        self.assertEqual(3, len(calls))

    def test_szse_batch_checkpoints_failures_and_resumes(self):
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "manifest.csv"
            failures = Path(directory) / "failures.csv"
            summary = Path(directory) / "summary.json"
            attempts = {"000002.SZ": 0}
            def discover(code, year):
                if code == "000002.SZ" and attempts[code] == 0:
                    attempts[code] += 1
                    raise TimeoutError("temporary")
                return [SZSEDisclosure(code, code, "2026-04-30", f"{year}年年度报告", "annual_report", f"https://disc/{code}.pdf")]
            _, first_failures, first = discover_szse_batch(
                [("000001.SZ", "甲"), ("000002.SZ", "乙")], 2025,
                output, failures, summary, 0, False, discover,
            )
            self.assertEqual(1, len(first_failures))
            self.assertFalse(first["complete"])
            rows, second_failures, second = discover_szse_batch(
                [("000001.SZ", "甲"), ("000002.SZ", "乙")], 2025,
                output, failures, summary, 0, True, discover,
            )
            self.assertFalse(second_failures)
            self.assertTrue(second["complete"])
            self.assertEqual(2, len(rows))

    def test_financial_derivation(self):
        def fact(code, value):
            return FinancialFact("600900.SH", "长江电力", 2025, code, Decimal(value), "https://official/report.pdf")
        facts = [
            fact("net_profit", "120"), fact("equity_begin", "900"), fact("equity_end", "1100"),
            fact("liabilities_end", "600"), fact("assets_end", "1500"),
            fact("cash_dividend_total", "50"), fact("total_shares", "100"),
            fact("environmental_investment", "2"), fact("revenue", "200"),
        ]
        values = {item.indicator_code: item.value for item in derive_financial_observations(facts)}
        self.assertAlmostEqual(12, values["Q_G_ROE"])
        self.assertAlmostEqual(40, values["Q_G_DEBT_ASSET_RATE"])
        self.assertAlmostEqual(.5, values["Q_S_DIVIDEND_PER_SHARE"])
        self.assertAlmostEqual(1, values["Q_S_ENV_INVEST_RATE"])

    def test_text_extraction_is_pending_and_converts_units(self):
        pages = [PageText(12, "2025年水资源使用强度为0.81立方米/万元，研发投入占比2.85%。")]
        items = extract_indicator_candidates(pages, "600900.SH", "长江电力", 2025, "https://official/esg.pdf", "esg.pdf")
        values = {item.indicator_code: item for item in items}
        self.assertEqual(810, values["Q_E_WATER_INTENSITY"].value)
        self.assertEqual("pending", values["Q_E_WATER_INTENSITY"].status.value)
        self.assertAlmostEqual(2.85, values["Q_S_RD_RATE"].value)

    def test_page_text_export_preserves_page_number(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.txt"
            path.write_text("\n=== PAGE 7 ===\n披露文本\n", encoding="utf-8")
            self.assertEqual([PageText(7, "披露文本\n")], read_page_text_export(path))

    def test_observation_csv_preserves_small_environmental_values(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "observations.csv"
            original = Observation(
                "A", "甲", 2025, "Q_E_CLEAN_ENERGY_INTENSITY", 0.0000218081042751,
                ValueStatus.PENDING, "url", "esg.pdf", 68, "same-table evidence", .95,
            )
            write_observations(path, [original])
            restored = read_observations(path, self.methodology)[0]
            self.assertAlmostEqual(original.value, restored.value, places=15)

    def test_batch_text_extraction_reports_company_coverage(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            text = root / "text/A/2025/esg_report.txt"
            text.parent.mkdir(parents=True)
            text.write_text("\n=== PAGE 2 ===\n资产负债率58.2%\n", encoding="utf-8")
            index = root / "index.csv"
            record = DocumentRecord("A", "甲", 2025, "esg_report", "https://source", "https://retrieval", "data/raw/A/2025/esg_report.pdf", "abc", 12)
            write_document_index(index, [record])
            items, coverage = extract_batch_text_exports(index, root / "text")
            self.assertEqual(1, len(items))
            self.assertEqual(
                {"Q_G_DEBT_ASSET_RATE": {"candidate_count": 1, "company_count": 1}},
                coverage,
            )
            filtered, _ = extract_batch_text_exports(index, root / "text", report_year=2024)
            self.assertFalse(filtered)

    def test_resolve_text_export_path_supports_ci_collection_layout(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            text_root = root / "data/text/ci_collection"
            target = text_root / "002709.SZ/2025/esg_report.txt"
            target.parent.mkdir(parents=True)
            target.write_text("\n=== PAGE 1 ===\nok\n", encoding="utf-8")
            row = {
                "company_code": "002709.SZ",
                "report_year": "2025",
                "local_path": "data/raw/ci_collection/002709.SZ/2025/esg_report.pdf",
            }
            resolved = resolve_text_export_path(text_root, row)
            self.assertIsNotNone(resolved)
            self.assertTrue(target.samefile(resolved))
            abs_row = {
                **row,
                "local_path": str(root / "data/raw/ci_collection/002709.SZ/2025/esg_report.pdf"),
            }
            self.assertTrue(target.samefile(resolve_text_export_path(text_root, abs_row)))
            # text_root may also be the parent data/text directory
            self.assertTrue(target.samefile(resolve_text_export_path(root / "data/text", row)))

    def test_debt_ratio_excludes_guarantee_threshold_and_formula(self):
        pages = [PageText(1, "资产负债率超过70%的被担保对象；资产负债率＝负债/资产×100%；期末资产负债率58.2%。")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "a.pdf")
        self.assertEqual([58.2], [item.value for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_english_debt_asset_ratio_direct_disclosure_is_current_actual_only(self):
        pages = [PageText(1, "The ratio of total liabilities to total assets of the Group was approximately 18.0% (2024: 21.1%).")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual.pdf")
        self.assertEqual([18.0], [item.value for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])
        rejected = extract_indicator_candidates(
            [PageText(2, "The ratio of total liabilities to total assets of the Group is expected to be below 20.0%. The ratio is calculated as total liabilities to total assets.")],
            "A", "甲", 2025, "url", "annual.pdf",
        )
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_intensity_excludes_table_footnote_as_value(self):
        pages = [PageText(1, "温室气体排放强度3 吨二氧化碳当量/万元 0.02 0.03")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_GHG_INTENSITY"])

    def test_standard_annual_report_direct_metrics(self):
        pages = [PageText(8, "加权平均净资产收益率（%）15.90；每10股派息数（元）（含税）10；研发投入总额占营业收入比例（%）2.85")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertEqual(15.9, values["Q_G_ROE"])
        self.assertEqual(1, values["Q_S_DIVIDEND_PER_SHARE"])
        self.assertEqual(2.85, values["Q_S_RD_RATE"])
        no_particle = extract_indicator_candidates(
            [PageText(9, "研发投入占营业收入比例 6.04% 7.18% -1.14%")],
            "A", "甲", 2025, "url", "annual.pdf",
        )
        self.assertEqual([6.04], [item.value for item in no_particle if item.indicator_code == "Q_S_RD_RATE"])

    def test_revenue_growth_uses_main_accounting_table(self):
        pages = [PageText(6, "近三年主要会计数据和财务指标\n（一）主要会计数据\n营业收入 120.00 100.00 20.00\n利润总额 30.00 20.00")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        self.assertEqual([20], [round(item.value, 6) for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"])

    def test_revenue_growth_repairs_wrapped_numbers_without_joining_columns(self):
        text = "近三年主要会计数据\n（一）主要会计数据\n营业收入 50,363,061,03\n0.51\n52,516,934,87\n3.70\n-4.10\n利润总额 1.00"
        items = extract_indicator_candidates([PageText(6, text)], "A", "甲", 2025, "url", "annual_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"]
        self.assertAlmostEqual(-4.100, values[0], places=2)

    def test_balance_sheet_turnover_normalizes_summary_ten_thousand_yuan(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n单位：万元\n营业收入 12.00 10.00\n利润总额 2.00"),
            PageText(80, "合并资产负债表\n资产总计 200000.00 180000.00\n负债合计 80000.00 70000.00"),
            PageText(81, "合并利润表"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(120000 / 190000, values["Q_G_ASSET_TURNOVER"])

    def test_consolidated_balance_sheet_derives_governance_metrics(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 120.00 100.00\n利润总额 20.00"),
            PageText(80, "合并资产负债表\n应收账款 10.00 8.00\n存货 5.00 4.00\n流动资产合计 40.00 30.00\n资产总计 200.00 180.00\n负债合计 80.00 70.00\n股东权益合计 120.00 110.00"),
            PageText(81, "合并利润表"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(40, values["Q_G_DEBT_ASSET_RATE"])
        self.assertAlmostEqual(120 / 190, values["Q_G_ASSET_TURNOVER"])
        self.assertAlmostEqual(120 / 9, values["Q_G_AR_TURNOVER"])
        self.assertAlmostEqual(37.5, values["Q_G_TWO_FUNDS_RATE"])
        self.assertAlmostEqual(100 / 11, values["Q_G_CAPITAL_ACCUMULATION"])

    def test_chinese_balance_sheet_derives_quick_ratio_and_excludes_parent_equity(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 120.00 100.00\n利润总额 20.00"),
            PageText(80, "合并资产负债表\n应收账款 10.00 8.00\n存货 5.00 4.00\n流动资产合计 40.00 30.00 非流动资产合计 160.00 150.00\n"
                         "资产总计 200.00 180.00\n流动负债合计 25.00 20.00 非流动负债合计 55.00 50.00\n负债合计 80.00 70.00\n"
                         "归属于母公司股东权益合计 100.00 90.00\n少数股东权益 20.00 20.00\n股东权益合计 120.00 110.00"),
            PageText(81, "合并利润表"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(140, values["Q_G_QUICK_RATIO"])
        self.assertAlmostEqual(100 / 11, values["Q_G_CAPITAL_ACCUMULATION"])

    def test_chinese_total_liabilities_and_equity_closes_detached_total_assets(self):
        pages = [PageText(10, """1、合并资产负债
表：元
资产总计
流动负债：
负债合计 400.00 350.00
归属于母公
司所有者权益合计 580.00 530.00
少数股东权益 20.00 20.00
所有者权益合计 600.00 550.00
负债和所有者权益总计 1,000.00 900.00""")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual.pdf")
        debt = [item for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"]
        self.assertEqual(1, len(debt))
        self.assertAlmostEqual(40, debt[0].value)
        self.assertIn("资产等价闭合行", debt[0].evidence_text)

        broken = PageText(10, pages[0].text.replace("1,000.00 900.00", "900.00 800.00"))
        rejected = extract_indicator_candidates([broken], "A", "甲", 2025, "url", "annual.pdf")
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_chinese_equivalent_assets_and_equity_recover_detached_liabilities(self):
        pages = [PageText(10, """1、合并资产负债表：元
资产总计
负债合计"""), PageText(11, """所有者权益合计 100.00 90.00
负债和所有者权益总计 500.00 450.00""")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual.pdf")
        debt = [item for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"]
        self.assertEqual(1, len(debt))
        self.assertAlmostEqual(80, debt[0].value)
        self.assertIn("资产负债恒等式派生", debt[0].evidence_text)

        ordinary_assets = PageText(10, """1、合并资产负债表：元
资产总计 500.00 450.00
负债合计
所有者权益合计 100.00 90.00""")
        rejected = extract_indicator_candidates([ordinary_assets], "A", "甲", 2025, "url", "annual.pdf")
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_chinese_ebitda_margin_requires_complete_supplementary_da(self):
        balance = "合并资产负债表\n流动负债合计 25.00 20.00\n资产总计 200.00 180.00\n股东权益合计 120.00 110.00"
        income = (
            "合并利润表\n二、营业总成本 900.00 800.00\n研发费用 36.00 30.00\n财务费用 25.00 20.00\n"
            "其中：利息费用 20.00 18.00\n三、营业利润（亏损以“－”号填列） 300.00 250.00\n"
            "四、利润总额（亏损总额以“－”号填列） 140.00 118.00\n减：所得税费用 20.00 18.00\n"
            "五、净利润（净亏损以“－”号填列） 120.00 100.00\n（一）按经营持续性分类\n"
            "1.持续经营净利润（净亏损以“－”号填列） 120.00 100.00"
        )
        supplement = (
            "现金流量表补充资料\n将净利润调节为经营活动现金流量：\n净利润 120.00 100.00\n"
            "固定资产折旧、油气资产折耗、生产性生物资产折旧 30.00 25.00\n使用权资产折旧 5.00 4.00\n"
            "无形资产摊销 8.00 7.00\n长期待摊费用摊销 2.00 1.00\n"
            "经营活动产生的现金流量净额 200.00 180.00\n2．不涉及现金收支的重大投资和筹资活动："
        )
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 1,200.00 1,000.00\n利润总额 140.00"),
            PageText(80, balance), PageText(84, income), PageText(90, supplement),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(205 / 1200 * 100, values["Q_G_EBITDA_MARGIN"])
        incomplete = [page for page in pages if page.page != 90] + [
            PageText(90, supplement.replace("\n长期待摊费用摊销 2.00 1.00", "")),
        ]
        items = extract_indicator_candidates(incomplete, "A", "甲", 2025, "url", "annual_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_EBITDA_MARGIN"])

    def test_chinese_roe_rd_fallback_used_only_without_disclosed_values(self):
        balance = "合并资产负债表\n流动负债合计 25.00 20.00\n资产总计 200.00 180.00\n股东权益合计 120.00 110.00"
        income = (
            "合并利润表\n研发费用 36.00 30.00\n其中：利息费用 20.00 18.00\n"
            "三、营业利润（亏损以“－”号填列） 300.00 250.00\n四、利润总额（亏损总额以“－”号填列） 140.00 118.00\n"
            "减：所得税费用 20.00 18.00\n五、净利润（净亏损以“－”号填列） 120.00 100.00"
        )
        base_pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 1,200.00 1,000.00\n利润总额 140.00"),
            PageText(80, balance), PageText(84, income),
        ]
        items = extract_indicator_candidates(base_pages, "A", "甲", 2025, "url", "annual_report.pdf")
        fallback = {
            item.indicator_code: item.value for item in items
            if item.evidence_text.startswith("合并报表回退派生: ")
        }
        self.assertAlmostEqual(120 / 115 * 100, fallback["Q_G_ROE"])
        self.assertAlmostEqual(3.0, fallback["Q_S_RD_RATE"])
        disclosed_pages = base_pages + [PageText(
            7, "主要会计数据 加权平均净资产收益率(%) 15.90 研发投入总额占营业收入比例(%) 3.20",
        )]
        items = extract_indicator_candidates(disclosed_pages, "A", "甲", 2025, "url", "annual_report.pdf")
        suppressed = [
            item for item in items
            if item.indicator_code in {"Q_G_ROE", "Q_S_RD_RATE"}
            and item.evidence_text.startswith("合并报表回退派生: ")
        ]
        self.assertEqual([], suppressed)
        values = {
            item.indicator_code: item.value for item in items
            if item.indicator_code in {"Q_G_ROE", "Q_S_RD_RATE"}
        }
        self.assertAlmostEqual(15.90, values["Q_G_ROE"])
        self.assertAlmostEqual(3.20, values["Q_S_RD_RATE"])

    def test_chinese_rd_fallback_rejects_empty_row_adjacent_values(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 2,499.47 2,300.00\n利润总额 140.00"),
            PageText(80, "合并资产负债表\n流动负债合计 25.00 20.00\n资产总计 200.00 180.00\n股东权益合计 120.00 110.00"),
            PageText(84, "合并利润表\n研发费用 财务费用 461.22 508.34\n其中：利息费用 461.93 509.77\n"
                         "三、营业利润（亏损以“－”号填列） 300.00 250.00\n四、利润总额（亏损总额以“－”号填列） 140.00 118.00\n"
                         "减：所得税费用 五、净利润（净亏损以“－”号填列） 120.00 100.00"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        fallback = [
            item for item in items
            if item.evidence_text.startswith("合并报表回退派生: ") and item.indicator_code == "Q_S_RD_RATE"
        ]
        self.assertEqual([], fallback)
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_EBITDA_MARGIN"])
        dash_pages = pages[:2] + [PageText(
            84, "合并利润表\n研发费用- - 财务费用 242.03 139.85\n其中：利息费用 20.00 18.00\n"
                "三、营业利润（亏损以“－”号填列） 300.00 250.00\n四、利润总额（亏损总额以“－”号填列） 140.00 118.00\n"
                "减：所得税费用 - - 五、净利润（净亏损以“－”号填列） 120.00 100.00",
        )]
        items = extract_indicator_candidates(dash_pages, "A", "甲", 2025, "url", "annual_report.pdf")
        fallback = [
            item for item in items
            if item.evidence_text.startswith("合并报表回退派生: ") and item.indicator_code == "Q_S_RD_RATE"
        ]
        self.assertEqual([], fallback)
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_EBITDA_MARGIN"])

    def test_english_consolidated_statement_derives_debt_asset_rate(self):
        pages = [
            PageText(100, "Consolidated Statement of Financial Position\nas at 31 December 2025\n2025 2024\nInventories 100 80\nTrade receivables 200 160\nTotal current assets 800 700\nTotal current liabilities 400 350\nTotal assets 2,000 1,800\nTotal liabilities 800 700\nTotal equity 1,200 1,100"),
            PageText(101, "Consolidated Statement of Profit or Loss\nRevenue 1,200 1,000\nFinance costs (20) (18)\nProfit before income tax 140 118\nIncome tax expense (20) (18)\nProfit for the year 120 100"),
            PageText(102, "Consolidated Statement of Cash Flows\nDepreciation and amortisation 40 35\nNet cash generated from operating activities 240 180"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        debt = [item for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"]
        self.assertEqual(1, len(debt))
        self.assertAlmostEqual(40, debt[0].value)
        self.assertEqual(100, debt[0].source_page)
        self.assertEqual(ValueStatus.PENDING, debt[0].status)
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1200 / 1900, values["Q_G_ASSET_TURNOVER"])
        self.assertAlmostEqual(1200 / 750, values["Q_G_CURRENT_ASSET_TURNOVER"])
        self.assertAlmostEqual(1200 / 180, values["Q_G_AR_TURNOVER"])
        self.assertAlmostEqual(37.5, values["Q_G_TWO_FUNDS_RATE"])
        self.assertAlmostEqual(175, values["Q_G_QUICK_RATIO"])
        self.assertAlmostEqual(100 / 1100 * 100, values["Q_G_CAPITAL_ACCUMULATION"])
        self.assertAlmostEqual(120 / 1150 * 100, values["Q_G_ROE"])
        self.assertAlmostEqual(160 / 1900 * 100, values["Q_G_ROA"])
        self.assertAlmostEqual(8, values["Q_G_EBITDA_INTEREST"])
        self.assertAlmostEqual(200 / 1200 * 100, values["Q_G_EBITDA_MARGIN"])

    def test_english_implicit_section_totals_require_two_year_identity_closure(self):
        valid = PageText(100, """Consolidated Statement of Financial Position
as at 31 December 2025
2025 2024
Current assets
Cash 10 9
22,838 26,839
Current liabilities
Payables (20,000) (21,000)
(38,414) (44,804)
Net current liabilities (15,576) (17,965)
Total assets less current liabilities 200,230 188,909
Equity
117,425 110,118
Non-current liabilities
Borrowings 50,000 49,000
82,805 78,791
Equity and non-current liabilities 200,230 188,909""")
        items = extract_indicator_candidates([valid], "A", "甲", 2025, "url", "annual.pdf")
        debt = [item for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"]
        self.assertEqual(1, len(debt))
        self.assertAlmostEqual((38_414 + 82_805) / 238_644 * 100, debt[0].value)

        broken = PageText(100, valid.text.replace("117,425 110,118", "117,000 110,118"))
        rejected = extract_indicator_candidates([broken], "A", "甲", 2025, "url", "annual.pdf")
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_english_audited_financial_summary_debt_rate_requires_identity(self):
        valid = PageText(20, """FIVE-YEAR FINANCIAL SUMMARY
ASSETS AND LIABILITIES
2025 2024 2023 2022 2021
Total assets 總資產 334,957 310,370 326,301 278,674 229,897
Total liabilities 總負債 (151,600) (131,123) (150,477) (109,648) (81,319)
Net assets 資產淨值 183,357 179,247 175,824 169,026 148,578""")
        items = extract_indicator_candidates([valid], "A", "甲", 2025, "url", "annual.pdf")
        debt = [item for item in items if item.indicator_code == "Q_G_DEBT_ASSET_RATE"]
        self.assertEqual(1, len(debt))
        self.assertAlmostEqual(151_600 / 334_957 * 100, debt[0].value)
        self.assertIn("two-year assets = liabilities + net assets", debt[0].evidence_text)

        broken = PageText(20, valid.text.replace("183,357 179,247", "180,000 170,000"))
        rejected = extract_indicator_candidates([broken], "A", "甲", 2025, "url", "annual.pdf")
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])
        associate = PageText(21, valid.text.replace(
            "FIVE-YEAR FINANCIAL SUMMARY\nASSETS AND LIABILITIES",
            "NOTES TO THE FINANCIAL STATEMENTS\nSummarised information of an associate",
        ))
        rejected = extract_indicator_candidates([associate], "A", "甲", 2025, "url", "annual.pdf")
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_G_DEBT_ASSET_RATE"])

    def test_english_statement_explicit_ascending_year_columns_are_normalized(self):
        pages = [
            PageText(10, """Statement of Profit or Loss
2024 2025
Revenue 1,000 1,200
Profit for the year 100 120"""),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual.pdf")
        growth = [item for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"]
        self.assertEqual(1, len(growth))
        self.assertAlmostEqual(20, growth[0].value)
        self.assertIn("ascending year columns normalized", growth[0].evidence_text)

    def test_english_balance_sheet_rejects_mixed_receivables_for_ar_metrics(self):
        pages = [
            PageText(100, "Consolidated Statement of Financial Position\n2025 2024\nInventories 100 80\nTrade and other receivables 200 160\nTotal current assets 800 700\nTotal current liabilities 400 350\nTotal assets 2,000 1,800\nTotal liabilities 800 700"),
            PageText(101, "Consolidated Statement of Profit or Loss\nRevenue 1,200 1,000"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertNotIn("Q_G_AR_TURNOVER", values)
        self.assertNotIn("Q_G_TWO_FUNDS_RATE", values)
        self.assertAlmostEqual(175, values["Q_G_QUICK_RATIO"])

    def test_english_ebitda_margin_requires_complete_combined_da_line(self):
        pages = [
            PageText(101, "Consolidated Statement of Profit or Loss\nRevenue 1,200 1,000\nFinance costs (20) (18)\nIncome tax expense (20) (18)\nProfit for the year 120 100"),
            PageText(102, "Consolidated Statement of Cash Flows\nDepreciation of property, plant and equipment 30 25\nAmortisation of intangible assets 10 10"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_EBITDA_MARGIN"])

    def test_chinese_transposed_roe_table_layout(self):
        pages = [
            PageText(250, "2、净资产收益率及每股收益\n报告期利润 加权平均净资产收益率 每股收益"),
            PageText(251, "基本每股收益（元/股） 稀释每股收益（元/股）\n归属于公司普通股股东的净 利润 6.08% 0.5414 0.5414\n扣除非经常性损益后归属于 公司普通股股东的净利润 5.93% 0.5274 0.5274"),
        ]
        items = extract_indicator_candidates(pages, "001289.SZ", "龙源电力", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(6.08, roe[0].value)
        self.assertEqual(251, roe[0].source_page)
        self.assertIn("转置表", roe[0].evidence_text)

    def test_chinese_transposed_roe_table_split_label_and_deduction_guard(self):
        pages = [
            PageText(90, "2、净资产收益率及每股收益\n报告期利润 加权平均净资产收益率 每股收益\n基本每股收益（元/股） 稀释每股收益（元/股）\n归属于公司普通股股东 -66.40% -0.199 -0.199 的净利润\n扣除非经常性损益后归 属于公司普通股股东的 -69.01% -0.207 -0.207 净利润"),
        ]
        items = extract_indicator_candidates(pages, "002506.SZ", "协鑫集成", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(-66.40, roe[0].value)

    def test_summary_roe_row_split_label_current_first(self):
        pages = [
            PageText(7, "主要会计数据\n加权平均净资产\n收益率\n1.05% 0.93% 0.12% 1.74%\n2025 年末 2024 年末\n总资产（元） 109,629,492,714.94"),
        ]
        items = extract_indicator_candidates(pages, "000703.SZ", "恒逸石化", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(1.05, roe[0].value)
        self.assertIn("主要会计数据加权平均净资产收益率行", roe[0].evidence_text)

    def test_summary_roe_row_six_value_restated_columns(self):
        pages = [
            PageText(7, "主要会计数据\n加权平均净资\n产收益率-73.17% -40.71% -40.71% -32.46% -15.69% -15.69%\n2024 年末 2025 年末"),
        ]
        items = extract_indicator_candidates(pages, "300051.SZ", "琏升科技", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(-73.17, roe[0].value)

    def test_summary_roe_row_declared_percent_bare_cells(self):
        pages = [
            PageText(7, "主要会计数据和财务指标\n加权平均净资产收益率（%）-155.89 -97.80 不适用-36.50\n扣除非经常性损益后的加权平均\n净资产收益率（%）-159.21 -98.41 不适用-37.09"),
        ]
        items = extract_indicator_candidates(pages, "600405.SH", "动力源", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(-155.89, roe[0].value)

    def test_summary_roe_row_loss_label_parenthesized_negative(self):
        pages = [
            PageText(6, "主要会计数据\n加权平均净资产（亏损）/收益率（%）\n*\n(5.953) 1.270 减少 7.22 个百分点 (5.504)\n扣除非经常性损益后的加权平均净资\n产（亏损）/收益率（%）* (5.910) 1.356 减少 7.27\n个百分点 (5.346)"),
        ]
        items = extract_indicator_candidates(pages, "600688.SH", "上海石化", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(-5.953, roe[0].value)

    def test_summary_roe_row_bse_multiline_basis_note(self):
        pages = [
            PageText(5, "主要财务数据\n加权平均净资产收益率%（依\n据归属于上市公司股东的净利\n润计算）\n7.08% 7.85% - 15.73%\n加权平均净资产收益率%（依\n据归属于上市公司股东的扣除\n非经常性损益后的净利润计\n算）\n5.55% 7.71% - 15.03%"),
        ]
        items = extract_indicator_candidates(pages, "920185.BJ", "贝特瑞", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(7.08, roe[0].value)

    def test_summary_roe_row_not_applicable_current_year_yields_nothing(self):
        pages = [
            PageText(7, "主要会计数据\n加权平均净资产收益\n率 不适用 -114.48% 不适用 -193.36%\n2025 年末 2024 年末"),
        ]
        items = extract_indicator_candidates(pages, "300340.SZ", "科恒股份", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(0, len(roe))

    def test_summary_roe_row_collapsed_year_header_rejected(self):
        pages = [
            PageText(7, "主要会计数据\n加权平均净资产收益率 2024 年 本年比上年增减 2023 年末\n2023 年\n2025 年末 2024 年末"),
        ]
        items = extract_indicator_candidates(pages, "300510.SZ", "金冠股份", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(0, len(roe))

    def test_summary_roe_row_prose_not_matched(self):
        from aegis_esg.extraction import _extract_summary_roe_row
        self.assertIsNone(_extract_summary_roe_row("管理层讨论\n公司加权平均净资产收益率为12.5%，同比提升。"))
        self.assertIsNone(_extract_summary_roe_row("扣除非经常性损益后的加权平均净资产收益率\n2.5% 3.1% 0.6% 1.9%"))

    def test_direct_roe_rule_preserves_negative_sign(self):
        pages = [
            PageText(8, "主要会计数据\n稀释每股收益(元/股) -0.27 -0.16 -68.75% -0.25 加权平均净资产收益 率 -24.66% -12.14% -103.13%"),
        ]
        items = extract_indicator_candidates(pages, "000637.SZ", "茂化实华", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertTrue(roe)
        self.assertTrue(all(item.value < 0 for item in roe))
        self.assertAlmostEqual(-24.66, roe[0].value)

    def test_direct_roe_rule_rejects_change_wording(self):
        pages = [
            PageText(8, "本报告期加权平均净资产收益率较上年同期下降5.20个百分点。"),
        ]
        items = extract_indicator_candidates(pages, "600000.SH", "测试", 2025, "url", "annual_report.pdf")
        roe = [item for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertEqual(0, len(roe))

    def test_direct_roe_rule_ignores_section_number_and_formula_variable(self):
        pages = [
            PageText(8, "2、加权平均净资产收益率的计算过程 (1) 加权平均净资产收益率以归属于本公司普通股股东的合并净利润除以加权平均净资产"),
            PageText(9, "L=D+A/2 4,239,398,052.63 加权平均净资产收益率 M=A/L 14.15% 期初股份总数 N 733,941,062.00"),
        ]
        items = extract_indicator_candidates(pages, "002533.SZ", "金杯电工", 2025, "url", "annual_report.pdf")
        roe = [item.value for item in items if item.indicator_code == "Q_G_ROE"]
        self.assertNotIn(1, roe)
        self.assertIn(14.15, roe)

    def test_chinese_employee_note_derives_per_capita(self):
        pages = [
            PageText(45, "八、公司员工情况\n报告期末母公司在职员工的数量（人） 500\n"
                         "报告期末主要子公司在职员工的数量（人） 4,390\n"
                         "报告期末在职员工的数量合计（人） 4,890"),
            PageText(158, "七、合并财务报表项目注释\n40、应付职工薪酬"),
            PageText(159, "（2） 短期薪酬列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "1、工资、奖金、津贴和补贴 172,535,126.31 597,283,514.77 585,204,289.87 184,614,351.21\n"
                          "2、职工福利费 32,261,755.23 32,261,755.23\n"
                          "3、社会保险费 29,499,254.51 29,496,178.88 3,075.63\n"
                          "其中：医疗保险费 23,751,028.64 23,748,312.74 2,715.90\n"
                          "4、住房公积金 8,737.00 29,292,647.14 29,151,613.14 149,771.00\n"
                          "5、工会经费和职工教育经费 6,080,718.87 5,087,463.95 6,169,927.47 4,998,255.35\n"
                          "合计 178,624,582.18 693,424,635.60 682,283,764.59 189,765,453.19"),
        ]
        items = extract_indicator_candidates(pages, "002533.SZ", "金杯电工", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(597_283_514.77 / 10_000 / 4890, values["Q_S_PAY_PER_EMPLOYEE"], places=6)
        benefit = (32_261_755.23 + 29_499_254.51 + 29_292_647.14) / 10_000 / 4890
        self.assertAlmostEqual(benefit, values["Q_S_BENEFIT_PER_EMPLOYEE"], places=6)
        self.assertAlmostEqual(5_087_463.95 / 10_000 / 4890, values["Q_S_EDU_PER_EMPLOYEE"], places=6)
        evidences = {item.indicator_code: item.evidence_text for item in items}
        self.assertTrue(evidences["Q_S_PAY_PER_EMPLOYEE"].startswith("中文应付职工薪酬附注派生: "))

    def test_chinese_employee_note_rejects_scope_mismatch(self):
        pages = [
            PageText(45, "报告期末母公司在职员工的数量（人） 500\n"
                         "报告期末主要子公司在职员工的数量（人） 4,390\n"
                         "报告期末在职员工的数量合计（人） 9,999"),
            PageText(159, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "1、工资、奖金、津贴和补贴 172,535,126.31 597,283,514.77 585,204,289.87 184,614,351.21\n"
                          "5、工会经费和职工教育经费 6,080,718.87 5,087,463.95 6,169,927.47 4,998,255.35"),
        ]
        items = extract_indicator_candidates(pages, "002533.SZ", "金杯电工", 2025, "url", "annual_report.pdf")
        social = [item for item in items if item.indicator_code.startswith("Q_S_") and "EMPLOYEE" in item.indicator_code]
        self.assertEqual(0, len(social))

    def test_chinese_employee_note_rejects_parent_company_section(self):
        pages = [
            PageText(45, "报告期末在职员工的数量合计（人） 4,890"),
            PageText(200, "七、合并财务报表项目注释\n1、货币资金 100.00 90.00"),
            PageText(260, "十七、母公司财务报表主要项目注释\n（2） 短期薪酬列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "1、工资、奖金、津贴和补贴 10,000.00 99,000.00 99,000.00 10,000.00\n"
                          "5、工会经费和职工教育经费 1,000.00 9,000.00 9,000.00 1,000.00"),
        ]
        items = extract_indicator_candidates(pages, "600000.SH", "测试", 2025, "url", "annual_report.pdf")
        social = [item for item in items if item.indicator_code.startswith("Q_S_") and "EMPLOYEE" in item.indicator_code]
        self.assertEqual(0, len(social))

    def test_chinese_employee_note_rejects_identity_failure(self):
        pages = [
            PageText(45, "报告期末在职员工的数量合计（人） 4,890"),
            PageText(159, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "5、工会经费和职工教育经费 6,080,718.87 9,999,999.99 6,169,927.47 4,998,255.35"),
        ]
        items = extract_indicator_candidates(pages, "002533.SZ", "金杯电工", 2025, "url", "annual_report.pdf")
        edu = [item for item in items if item.indicator_code == "Q_S_EDU_PER_EMPLOYEE"]
        self.assertEqual(0, len(edu))

    def test_chinese_employee_note_bse_employee_total(self):
        pages = [
            PageText(62, "二、 员工情况\n(一) 在职员工（公司及控股子公司）基本情况\n"
                         "按工作性质分类 期初人数 本期新增 本期减少 期末人数\n"
                         "管理人员 276 15 61 230\n员工总计 8,352 1,720 1,930 8,142"),
            PageText(159, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n项目 年初余额 本年增加 本年减少 年末余额\n"
                          "1、工资、奖金、\n津贴和补贴 247,221,035.58 1,280,263,200.97 1,295,257,918.63 232,226,317.92\n"
                          "2、职工福利费 114,233,676.01 114,109,605.72 124,070.29\n"
                          "3、社会保险费 67,500.97 48,166,301.16 48,154,106.57 79,695.56\n"
                          "4、住房公积金 985,001.06 58,718,760.64 58,830,803.14 872,958.56\n"
                          "5、工会经费和\n职工教育经费 2,533,400.81 20,204,747.53 20,112,882.53 2,625,265.81"),
        ]
        items = extract_indicator_candidates(pages, "920185.BJ", "贝特瑞", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1_280_263_200.97 / 10_000 / 8142, values["Q_S_PAY_PER_EMPLOYEE"], places=6)
        benefit = (114_233_676.01 + 48_166_301.16 + 58_718_760.64) / 10_000 / 8142
        self.assertAlmostEqual(benefit, values["Q_S_BENEFIT_PER_EMPLOYEE"], places=6)
        self.assertAlmostEqual(20_204_747.53 / 10_000 / 8142, values["Q_S_EDU_PER_EMPLOYEE"], places=6)

    def test_chinese_employee_note_horizontal_employee_table(self):
        pages = [
            PageText(38, "十、公司员工情况\n1、员工数量、专业构成及教育程度\n"
                         "报告期末母公司在职员工的数量（人） 报告期末主要子公司在职员工的数量（人） "
                         "报告期末在职员工的数量合计（人） 当期领取薪酬员工总人数（人） 审议通过事项\n"
                         "无 无\n38\n1,624\n153\n1,777\n1,777\n39\n专业构成\n专业构成类别 专业构成人数（人）\n"
                         "生产人员 900\n销售人员 200\n技术人员 500\n财务人员 50\n行政人员 127\n合计 1,777"),
            PageText(148, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "5、工会经费和职工教育经费 18,149,074.71 286,516.34 265,126.62 18,170,464.43"),
        ]
        items = extract_indicator_candidates(pages, "300690.SZ", "双一科技", 2025, "url", "annual_report.pdf")
        edu = [item for item in items if item.indicator_code == "Q_S_EDU_PER_EMPLOYEE"]
        self.assertEqual(1, len(edu))
        self.assertAlmostEqual(286_516.34 / 10_000 / 1777, edu[0].value, places=6)

    def test_chinese_employee_note_cross_page_header_binding(self):
        pages = [
            PageText(30, "报告期末母公司在职员工的数量（人） 400\n"
                         "报告期末主要子公司在职员工的数量（人） 2,272\n报告期末在职员工的数量合计（人） 2,672"),
            PageText(163, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n单位：元\n"
                          "项目 期初余额 本期增加 本期减少 期末余额\n163"),
            PageText(164, "某某股份有限公司 2025 年年度报告全文\n"
                          "1、工资、奖金、津贴和补贴 150,687,035.94 685,500,805.16 649,810,762.44 186,377,078.66\n"
                          "2、职工福利费 11,072,640.34 11,072,640.34\n"
                          "3、社会保险费 27,370,582.98 27,319,354.86 51,228.12\n"
                          "4、住房公积金 17,955,372.75 17,794,190.75 161,182.00\n"
                          "5、工会经费和职工教育经费 240.80 121,838.57 111,484.17 10,595.20\n"
                          "合计 150,687,276.74 742,021,239.80 706,108,432.56 186,600,083.98\n"
                          "（3） 设定提存计划列示\n单位：元\n项目 期初余额 本期增加 本期减少 期末余额\n"
                          "1、基本养老保险 32,448,315.93 32,447,712.64 603.29"),
        ]
        items = extract_indicator_candidates(pages, "002121.SZ", "科陆电子", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(685_500_805.16 / 10_000 / 2672, values["Q_S_PAY_PER_EMPLOYEE"], places=6)
        benefit = (11_072_640.34 + 27_370_582.98 + 17_955_372.75) / 10_000 / 2672
        self.assertAlmostEqual(benefit, values["Q_S_BENEFIT_PER_EMPLOYEE"], places=6)

    def test_chinese_employee_note_truncated_header_with_detached_year_end(self):
        pages = [
            PageText(50, "报告期末在职员工的数量合计（人） 305"),
            PageText(151, "七、合并财务报表项目注释\n24. 应付职工薪酬\n（1） 应付职工薪酬分类\n"
                          "项目 年初余额 本年增加 本年减少 年末余额\n"
                          "短期薪酬 16,052,879.47 110,632,039.53 107,188,425.05 19,496,493.95"),
            PageText(152, "（2） 短期薪酬\n项目 年初余额 本年增加 本年减少 工资、奖金、津贴\n"
                          "和补贴 15,180,182.34 89,766,895.18 86,203,229.46 职工福利费 731,700.11 4,986,114.72 "
                          "5,078,811.21 社会保险费 4,680,731.18 4,675,950.39 其中：医疗保险费 3,737,049.09 "
                          "3,733,355.89 住房公积金 9,267,711.08 9,267,711.08\n"
                          "工会经费和职工教\n育经费 140,997.02 1,930,587.37 1,962,722.91 短期带薪缺勤\n"
                          "合计 16,052,879.47 110,632,039.53 107,188,425.05 （3） 设定提存计划\n"
                          "年末余额\n18,743,848.06\n639,003.62\n4,780.79\n3,693.20\n108,861.48\n19,496,493.95"),
        ]
        items = extract_indicator_candidates(pages, "000037.SZ", "深南电", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(89_766_895.18 / 10_000 / 305, values["Q_S_PAY_PER_EMPLOYEE"], places=6)
        benefit = (4_986_114.72 + 4_680_731.18 + 9_267_711.08) / 10_000 / 305
        self.assertAlmostEqual(benefit, values["Q_S_BENEFIT_PER_EMPLOYEE"], places=6)
        self.assertAlmostEqual(1_930_587.37 / 10_000 / 305, values["Q_S_EDU_PER_EMPLOYEE"], places=6)

    def test_chinese_employee_note_rejects_band_split_layout(self):
        pages = [
            PageText(30, "报告期末在职员工的数量合计（人） 2,000"),
            PageText(176, "七、合并财务报表项目注释\n（2） 短期薪酬列示\n单位：元\n176"),
            PageText(177, "项目 期初余额 1、工资、奖金、津贴\n和补贴 160,950,554.56 2、职工福利费 65,611.95 "
                          "5、工会经费和职工教\n育经费 1,643,586.06 合计 165,224,336.00 （3） 设定提存计划列示\n"
                          "本期增加 本期减少 期末余额\n368,011,318.41 316,268,866.91 212,693,006.06\n"
                          "65,611.95 65,611.95 -\n1,930,587.37 1,962,722.91 1,611,450.52"),
        ]
        items = extract_indicator_candidates(pages, "000155.SZ", "川能动力", 2025, "url", "annual_report.pdf")
        social = [item for item in items if item.indicator_code.startswith("Q_S_") and "EMPLOYEE" in item.indicator_code]
        self.assertEqual(0, len(social))

    def test_chinese_numbered_statement_titles_derive_balance_indicators(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 1,000.00 900.00\n利润总额 150.00"),
            PageText(60, "1、合并资产负债表\n流动资产合计 500.00 450.00\n存货 100.00 90.00\n应收账款 200.00 180.00\n资产总计 1,000.00 900.00\n流动负债合计 300.00 250.00\n负债合计 400.00 350.00\n股东权益合计 600.00 550.00"),
            PageText(63, "3、合并利润表\n一、营业总收入 1,000.00 900.00\n三、营业利润（亏损以“－”号填列） 150.00 120.00"),
        ]
        items = extract_indicator_candidates(pages, "000703.SZ", "恒逸石化", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(40, values["Q_G_DEBT_ASSET_RATE"])
        self.assertAlmostEqual((500 - 100) / 300 * 100, values["Q_G_QUICK_RATIO"])

    def test_summary_marker_variant_enables_revenue(self):
        pages = [
            PageText(6, "六、主要会计数据和财务指标\n单位：元\n项目 2025年 2024年\n营业收入 8,633,599,405.13 7,903,721,308.59\n归属于上市公司股东的净利润 1,024,495,080.91 705,541,276.56"),
            PageText(60, "合并资产负债表\n资产总计 20,000,000,000.00 18,000,000,000.00\n负债合计 7,000,000,000.00 6,000,000,000.00\n股东权益合计 13,000,000,000.00 12,000,000,000.00"),
            PageText(63, "合并利润表\n五、净利润（净亏损以“－”号填列） 1,024,495,080.91 705,541,276.56\n减：所得税费用 300,000,000.00 200,000,000.00"),
        ]
        items = extract_indicator_candidates(pages, "000690.SZ", "宝新能源", 2025, "url", "annual_report.pdf")
        roe = [
            item for item in items
            if item.indicator_code == "Q_G_ROE" and item.evidence_text.startswith("合并报表回退派生: ")
        ]
        self.assertEqual(1, len(roe))
        self.assertAlmostEqual(1024495080.91 / ((13_000_000_000 + 12_000_000_000) / 2) * 100, roe[0].value)

    def test_cas_english_statements_derive_indicators(self):
        pages = [
            PageText(112, "Consolidated balance sheet\nDecember 31, 2025\nUnit:Yuan Currency:CNY\nTotal current assets 66,238,762,666.65 66,192,918,621.75\nInventories 14,880,555,929.26 12,633,286,216.02\nAccounts receivable 6,863,715,377.32 6,706,810,000.00"),
            PageText(113, "Total assets 187,779,256,397.63 195,916,763,061.99\nTotal current liabilities 55,766,509,445.26 56,538,075,361.44"),
            PageText(114, "Total liabilities 136,391,138,210.41 137,997,611,563.26\nTotal owners’ equity (or shareholders'\nequity) 51,388,118,187.22 57,919,151,498.73\nTotal liabilities and owners’ equity 187,779,256,397.63 195,916,763,061.99"),
            PageText(116, "Consolidated Profit Statement\nI. Total operating revenue 84,128,281,703.14 91,994,404,333.54\nII. Total operating cost\nIncluding: Operating cost 81,856,406,510.25 86,117,213,124.73\nR&D cost 1,106,164,996.59 1,510,114,124.23\nIncluding: Interest expense 2,895,890,520.93 2,259,805,051.33\nIII. Operating profit (“\n-\n” for loss) -11,525,989,535.68 -8,418,172,185.07"),
            PageText(117, "IV: Total profit (“\n-\n” for loss) -11,670,943,654.43 -8,683,316,454.96\nLess: Income tax expense -770,060,689.82 -574,532,383.25\nV . Net profit (“\n-\n” for net loss) -10,900,882,964.61 -8,108,784,071.71\n1. Net profit attributable to shareholders of\nthe parent company -9,553,425,884.06 -7,038,757,392.54"),
        ]
        items = extract_indicator_candidates(pages, "600438.SH", "通威股份", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(136391138210.41 / 187779256397.63 * 100, values["Q_G_DEBT_ASSET_RATE"])
        self.assertAlmostEqual((51388118187.22 - 57919151498.73) / 57919151498.73 * 100, values["Q_G_CAPITAL_ACCUMULATION"])
        self.assertAlmostEqual(-10900882964.61 / ((51388118187.22 + 57919151498.73) / 2) * 100, values["Q_G_ROE"])
        self.assertAlmostEqual(1106164996.59 / 84128281703.14 * 100, values["Q_S_RD_RATE"])
        self.assertAlmostEqual(-8.55, values["Q_G_REVENUE_GROWTH"], places=1)
        self.assertAlmostEqual((66238762666.65 - 14880555929.26) / 55766509445.26 * 100, values["Q_G_QUICK_RATIO"])
        self.assertNotIn("Q_G_COST_REVENUE_RATE", values)

    def test_balance_derivation_rejected_when_accounting_identity_breaks(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 60,000.00 55,000.00\n利润总额 150.00"),
            PageText(100, "合并资产负债表\n流动资产合计 500.00 450.00\n存货 100.00 90.00\n流动负债合计 25.00 20.00\n资产总计 200.00 180.00\n负债合计 184,139,748.17 300,497,387.53\n股东权益合计 1,484,256,563.32 1,473,230,420.64"),
        ]
        items = extract_indicator_candidates(pages, "001331.SZ", "胜通能源", 2025, "url", "annual_report.pdf")
        derived = [item for item in items if "合并报表自动派生: " in item.evidence_text]
        self.assertEqual([], derived)

    def test_statement_fact_rejects_values_detached_behind_another_label(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 6,000,000,000.00 5,500,000,000.00\n利润总额 150.00"),
            PageText(100, "合并资产负债表\n资产总计 流动负债： 短期借款\n17,174,396.04 29,825,551.92\n股东权益合计 1,484,256,563.32 1,473,230,420.64"),
        ]
        items = extract_indicator_candidates(pages, "001331.SZ", "胜通能源", 2025, "url", "annual_report.pdf")
        derived = {item.indicator_code: item.value for item in items if "合并报表自动派生: " in item.evidence_text}
        self.assertNotIn("Q_G_DEBT_ASSET_RATE", derived)
        self.assertNotIn("Q_G_ASSET_TURNOVER", derived)
        self.assertAlmostEqual((1484256563.32 - 1473230420.64) / 1473230420.64 * 100, derived["Q_G_CAPITAL_ACCUMULATION"])

    def test_balance_derivation_rejected_for_implausibly_small_assets(self):
        pages = [
            PageText(68, "合并资产负债表\n资产总计 3,218,767.51 3,849,642.78\n负债合计 48,109,096.25 47,438,919.35"),
        ]
        items = extract_indicator_candidates(pages, "920237.BJ", "力佳科技", 2025, "url", "annual_report.pdf")
        derived = [item for item in items if "合并报表自动派生: " in item.evidence_text]
        self.assertEqual([], derived)

    def test_english_consolidated_income_derives_revenue_growth_and_skips_note_column(self):
        pages = [
            PageText(120, "Consolidated Statement of Profit or Loss\nfor the year ended 31 December 2025\nRevenue 5 1,200,000 1,000,000\nTotal operating expenses 9 (900,000) (800,000)\nOperating profit 8 240,000 200,000\nCost of sales (800,000) (700,000)"),
            PageText(121, "Consolidated Statement of Changes in Equity"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        growth = [item for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"]
        self.assertEqual(1, len(growth))
        self.assertAlmostEqual(20, growth[0].value)
        self.assertEqual(120, growth[0].source_page)
        self.assertEqual(ValueStatus.PENDING, growth[0].status)
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(20, values["Q_G_OPERATING_MARGIN"])
        self.assertAlmostEqual(20, values["Q_G_OPERATING_PROFIT_GROWTH"])
        self.assertAlmostEqual(75, values["Q_G_COST_REVENUE_RATE"])

    def test_english_cost_revenue_rate_rejects_cost_of_sales_only(self):
        pages = [PageText(
            120,
            "Consolidated Statement of Profit or Loss\nRevenue 1,200 1,000\n"
            "Cost of sales (800) (700)",
        )]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_COST_REVENUE_RATE"])

    def test_english_cashflow_derives_cash_governance_metrics(self):
        pages = [
            PageText(90, "Consolidated Statement of Financial Position\n2025 2024\nTotal current liabilities 400 350\nTotal assets 2,000 1,800\nTotal liabilities 800 700"),
            PageText(91, "Consolidated Statement of Profit or Loss\n2025 2024\nRevenue 1,200 1,000"),
            PageText(92, "Consolidated Statement of Cash Flows\n2025 2024\nReceipts from customers 1,260 1,050\nNet cash generated from operating activities 240 180"),
            PageText(93, "Consolidated Statement of Changes in Equity"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(105, values["Q_G_CASH_REALIZATION"])
        self.assertAlmostEqual(60, values["Q_G_CASH_CURRENT_LIABILITY"])

    def test_english_cashflow_rejects_contaminated_customer_receipts(self):
        pages = [
            PageText(91, "Consolidated Statement of Profit or Loss\nRevenue 1,200 1,000"),
            PageText(92, "Consolidated Statement of Cash Flows\nReceipts from customers Payments to suppliers 1,260 (800) 1,050 (700)"),
        ]
        items = extract_indicator_candidates(pages, "00001.HK", "甲", 2025, "url", "annual_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_CASH_REALIZATION"])

    def test_english_direct_method_cashflow_accepts_sales_and_services_receipts(self):
        pages = [
            PageText(
                84,
                "Consolidated Statement of Profit or Loss\n2025 2024\n"
                "I. Total operating revenue V.33 5,000 4,000\n"
                "II. Total operating cost 4,000 3,300\n"
                "R&D expenses V.46 100 80\nIII. Operating profit 500 400",
            ),
            PageText(
                87,
                "Consolidated Cash Flow Statement\n2025 2024\n"
                "Cash received from sales of goods or rendering of\n"
                "services 5,750 4,200\nCash received relating to other operating activities 30 40",
            ),
        ]
        items = extract_indicator_candidates(
            pages, "01713.HK", "SICHUAN ENERGY", 2025, "url", "annual_report.pdf",
        )
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(115, values["Q_G_CASH_REALIZATION"])
        self.assertAlmostEqual(25, values["Q_G_REVENUE_GROWTH"])
        self.assertAlmostEqual(10, values["Q_G_OPERATING_MARGIN"])
        self.assertAlmostEqual(25, values["Q_G_OPERATING_PROFIT_GROWTH"])
        self.assertAlmostEqual(80, values["Q_G_COST_REVENUE_RATE"])
        self.assertAlmostEqual(2, values["Q_S_RD_RATE"])

    def test_english_revenue_intensities_convert_to_methodology_units(self):
        pages = [PageText(
            20,
            "Comprehensive energy consumption intensity\n"
            "tonnes of coal equivalent / RMB million of revenue 0.60\n"
            "Water consumption intensity tonnes / RMB million of revenue 34.95",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(6, values["Q_E_ENERGY_INTENSITY"])
        self.assertAlmostEqual(349.5, values["Q_E_WATER_INTENSITY"])

    def test_english_energy_intensity_accepts_explicit_per_unit_revenue_sentence(self):
        pages = [PageText(22,
            "The comprehensive energy intensity per unit of revenue was 75.30 tonnes of "
            "standard coal equivalent per RMB 100 million of revenue."
        )]
        items = extract_indicator_candidates(pages, "02386.HK", "SINOPEC SEG", 2025, "url", "esg_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_ENERGY_INTENSITY"]
        self.assertEqual([7.53], values)

    def test_english_revenue_intensities_reject_non_revenue_denominators(self):
        pages = [PageText(
            20,
            "Total Energy Consumption Intensity MWh/employee 1.31\n"
            "Water consumption intensity m3/m2 total area 0.005",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_E_ENERGY_INTENSITY", "Q_E_WATER_INTENSITY",
        }])

    def test_english_revenue_intensity_does_not_consume_later_percentage(self):
        pages = [PageText(
            20,
            "Comprehensive energy consumption intensity: 0.6 tonnes of coal equivalent "
            "per RMB million of revenue 100% of employees received training",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_ENERGY_INTENSITY"]
        self.assertEqual([6], values)

    def test_english_waste_intensity_accepts_parenthesized_revenue_unit(self):
        pages = [PageText(11,
            "Hazardous waste intensity (tonnes/RMB million revenue) 0.19\n"
            "Non-hazardous waste intensityNote1 (tonnes/RMB million revenue) 0.61"
        )]
        items = extract_indicator_candidates(pages, "03303.HK", "JUTAL", 2025, "url", "esg_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1.9, values["Q_E_HAZ_WASTE_INTENSITY"])
        self.assertAlmostEqual(6.1, values["Q_E_SOLID_WASTE_INTENSITY"])

    def test_english_ghg_revenue_intensity_converts_rmb_scales(self):
        pages = [
            PageText(20, "GHG emissions intensity tCO2e/RMB million of revenue 1.24"),
            PageText(21, "Greenhouse gas emissions intensity Tonne CO2e/100 million RMB 198.48"),
            PageText(22, "Total GHG emission intensity tCO2e/CNY 10k revenue 0.06"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_INTENSITY"]
        self.assertEqual([12.4, 19.848, 60], values)

    def test_english_ghg_reduction_uses_explicit_current_previous_header(self):
        pages = [PageText(
            20,
            "Environmental Indicators\nIndicator Unit 2025 2024\n"
            "Total GHG emissions (Scope 1 and Scope 2) tCO2e 80,000 100,000",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"]
        self.assertEqual([20], values)

    def test_chinese_ghg_reduction_rejects_first_time_scope3_boundary_change(self):
        pages = [PageText(
            65,
            "指标名称 指标单位 2024 年数值 2025 年数值\n"
            "温室气体排放总量 吨二氧化碳当量 435,889.44 4,465,680.936\n"
            "注：2025 年温室气体排放总量首次增加了范围三的统计。",
        )]
        items = extract_indicator_candidates(
            pages, "600481.SH", "双良节能", 2025, "https://source", "esg.pdf",
        )
        self.assertFalse([
            item for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"
        ])

    def test_strict_chinese_hazardous_waste_intensity_is_auto_confirmed(self):
        pages = [PageText(
            31,
            "指标 单位 2025 年\n危险废弃物产生强度 吨 / 万元营收 0.0006",
        )]
        items = extract_indicator_candidates(
            pages, "000531.SZ", "穗恒运A", 2025, "https://source", "esg_report.pdf",
        )
        confirmed, unresolved, decisions = resolve_pending_candidates(items)
        hazardous = [
            item for item in confirmed if item.indicator_code == "Q_E_HAZ_WASTE_INTENSITY"
        ]
        self.assertEqual([0.6], [item.value for item in hazardous])
        self.assertFalse([
            item for item in unresolved if item.indicator_code == "Q_E_HAZ_WASTE_INTENSITY"
        ])
        self.assertEqual(
            ["auto_confirmed"],
            [item.decision for item in decisions if item.indicator_code == "Q_E_HAZ_WASTE_INTENSITY"],
        )

    def test_chinese_rd_rate_table_maps_explicit_ascending_years(self):
        pages = [PageText(45, """披露项 单位 2023 年 2024 年 2025 年
研发费用占营业收入比例 % 0.92 1.29 0.97""")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_RD_RATE"]
        self.assertEqual([0.97], values)
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertEqual([0.97], [item.value for item in confirmed if item.indicator_code == "Q_S_RD_RATE"])
        self.assertFalse([item for item in unresolved if item.indicator_code == "Q_S_RD_RATE"])
        rejected = extract_indicator_candidates(
            [PageText(46, "2023-2025 年研发绩效\n研发费用占营业收入比例 % 0.92 1.29 0.97")],
            "A", "甲", 2025, "url", "esg.pdf",
        )
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_S_RD_RATE"])

    def test_english_ghg_reduction_normalizes_explicit_ascending_year_order(self):
        pages = [PageText(
            20,
            "Environmental Indicators\nIndicator Unit 2024 2025\n"
            "Total GHG emissions tCO2e 100,000 80,000",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"]
        self.assertEqual([20], values)

    def test_english_total_ghg_yoy_narrative_requires_group_actual_and_report_year(self):
        pages = [PageText(20, "In 2025, the Group’s total GHG emissions declined by 9.7% year-on-year to 45,783 kilotonnes CO2e.")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"]
        self.assertEqual([9.7], values)
        rejected = extract_indicator_candidates(
            [PageText(21, "By 2030, the Group’s total GHG emissions are expected to decrease by 20% year-on-year.")],
            "A", "甲", 2025, "https://source", "esg.pdf",
        )
        self.assertFalse([item for item in rejected if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"])

    def test_english_ghg_reduction_ignores_trailing_percentage_change_column(self):
        pages = [PageText(20, """GHG Emissions Units 2024 2025
Percentage change
Total GHG emissions tCO2e 37.95 26.95 (28.99%)""")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"]
        self.assertEqual(1, len(values))
        self.assertAlmostEqual((37.95 - 26.95) / 37.95 * 100, values[0])

    def test_english_ghg_reduction_accepts_equivalent_tonnes_unit_order(self):
        pages = [PageText(124, """Indicator Note 2 Unit 2025 2024
Total GHG Emissions (Scope 1 & 2) Equivalent of carbon dioxide in tonnes 2,058.07 2,069.83""")]
        items = extract_indicator_candidates(pages, "00206.HK", "CM ENERGY", 2025, "https://source", "annual.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_GHG_REDUCTION_RATE"]
        self.assertEqual(1, len(values))
        self.assertAlmostEqual((2069.83 - 2058.07) / 2069.83 * 100, values[0])

    def test_english_ghg_intensity_rejects_non_rmb_denominators(self):
        pages = [PageText(
            20,
            "GHG emissions intensity 0.50 kg CO2e/kWh. "
            "Total GHG emissions intensity tCO2e / USD million revenue 3,130.41",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_GHG_INTENSITY"])

    def test_english_nox_and_solid_waste_intensities_convert_units(self):
        pages = [
            PageText(20, "Intensity of NOx emissions Tonnes/RMB billion in revenue 0.09"),
            PageText(21, "Non-hazardous waste intensity Tonnes/CNY 10k revenue 0.03"),
            PageText(22, "Non-hazardous waste intensity Tons per Million Yuan of Revenue 0.05"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [(item.indicator_code, item.value) for item in items]
        self.assertEqual([
            ("Q_E_NOX_INTENSITY", .9),
            ("Q_E_SOLID_WASTE_INTENSITY", 30),
            ("Q_E_SOLID_WASTE_INTENSITY", .5),
        ], values)

    def test_english_inverted_air_emission_labels_use_rmb_revenue_units(self):
        pages = [PageText(
            57,
            "Type of Emissions Unit 2025 2024 2023\n"
            "Intensity of SO2 emissions Kg/RMB10,000 0.44 0.24 0.18\n"
            "Intensity of particulate emissions Kg/RMB10,000 0.11 0.06 0.05",
        )]
        items = extract_indicator_candidates(
            pages, "06885.HK", "JINMA ENERGY", 2025, "https://source", "annual_report.pdf",
        )
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(440, values["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(110, values["Q_E_PM_INTENSITY"])

    def test_english_additional_environmental_intensities_convert_units(self):
        pages = [PageText(
            20,
            "SO2 emissions intensity Tonnes per Million Yuan of Revenue 0.05\n"
            "PM emissions intensity kg/RMB million of revenue 0.20\n"
            "Wastewater discharge intensity Tons per Million Yuan of Revenue 4.97\n"
            "Hazardous waste intensity Tonnes/CNY 10k revenue 0.01",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(500, values["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(2, values["Q_E_PM_INTENSITY"])
        self.assertAlmostEqual(49.7, values["Q_E_WASTEWATER_INTENSITY"])
        self.assertAlmostEqual(10, values["Q_E_HAZ_WASTE_INTENSITY"])

    def test_english_additional_environmental_intensities_reject_production_denominators(self):
        pages = [PageText(
            20,
            "SOx emission intensity 1.73 kg/MW of wafers\n"
            "Particulate matter emission intensity 0.72 kg/MW of wafers\n"
            "Intensity of wastewater discharge 2.10 tonnes per barrel of crude oil\n"
            "Hazardous waste intensity 0.4 tonnes per employee",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_E_SO2_INTENSITY", "Q_E_PM_INTENSITY", "Q_E_WASTEWATER_INTENSITY",
            "Q_E_HAZ_WASTE_INTENSITY",
        }])

    def test_hazardous_waste_rule_does_not_match_non_hazardous_label(self):
        pages = [PageText(
            20, "Non-hazardous waste intensity Tonnes/CNY 10k revenue 0.03",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_HAZ_WASTE_INTENSITY"])

    def test_english_recycled_water_proportion_accepts_explicit_share(self):
        pages = [PageText(20, "Proportion of the recycled water consumption 8.53 %")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"]
        self.assertEqual([8.53], values)

    def test_english_recycled_water_proportion_rejects_utilisation_and_over_100(self):
        pages = [PageText(
            20,
            "Utilisation rate of recycled water (%) 96.8\n"
            "Proportion of recycled water 292.7%",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_ALTERNATIVE_WATER_RATE"])

    def test_english_clean_energy_intensity_and_environmental_investment_rate(self):
        pages = [PageText(
            20,
            "Clean energy consumption intensity tonnes of standard coal equivalent "
            "per RMB million of revenue 0.60\n"
            "Proportion of environmental protection investment16 Unit 2024 % / 2025 0.18",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(6, values["Q_E_CLEAN_ENERGY_INTENSITY"])
        self.assertAlmostEqual(.18, values["Q_S_ENV_INVEST_RATE"])

    def test_english_clean_energy_and_environmental_investment_reject_wrong_context(self):
        pages = [PageText(
            20,
            "Renewable Energy Consumption intensity kWh/Production unit (’000) 208.77\n"
            "Proportion of Environmental Protection Investment (%) = "
            "Environmental Protection Investment/Revenue",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_E_CLEAN_ENERGY_INTENSITY", "Q_S_ENV_INVEST_RATE",
        }])

    def test_english_clean_energy_intensity_derives_from_current_first_same_table(self):
        pages = [PageText(20,
            "Environmental\nIndicator Unit 2025 2024 2023\n"
            "Renewable energy consumption Tonnes of standard coal 2,440.43 1,422.36 1,251.75\n"
            "Comprehensive energy consumption Tonnes of standard coal 38,926.06 71,336.26 85,106.43\n"
            "Intensity of comprehensive energy consumption\n"
            "Tonnes of standard coal /RMB billion in revenue 347.85 649.38 747.48\n"
            "Intensity of water consumption Tonnes/RMB billion in revenue 6,091.30 25,839.34 16,307.71\n"
            "Intensity of hazardous waste generation Tonnes/RMB billion in revenue 0.61 0.21 0.41\n"
            "Intensity of non-hazardous waste generation Tonnes/RMB billion in revenue 7.48 24.59 22.05\n"
            "Total SO2 emissions Tonnes 1.49 45.84 21.06\n"
            "Total wastewater discharge Tonnes 403,550.20 2,412,749.27 1,578,248.77\n"
        )]
        items = extract_indicator_candidates(pages, "02688.HK", "ENN ENERGY", 2025, "url", "esg_report.pdf")
        by_code = {item.indicator_code: item for item in items}
        matches = [by_code["Q_E_CLEAN_ENERGY_INTENSITY"]]
        self.assertEqual(1, len(matches))
        self.assertAlmostEqual(0.21808104, matches[0].value, places=8)
        self.assertTrue(matches[0].evidence_text.startswith("English same-table renewable energy intensity derived:"))
        self.assertAlmostEqual(3.4785, by_code["Q_E_ENERGY_INTENSITY"].value)
        self.assertAlmostEqual(60.913, by_code["Q_E_WATER_INTENSITY"].value)
        self.assertAlmostEqual(.0061, by_code["Q_E_HAZ_WASTE_INTENSITY"].value)
        self.assertAlmostEqual(.0748, by_code["Q_E_SOLID_WASTE_INTENSITY"].value)
        self.assertAlmostEqual(.13314898, by_code["Q_E_SO2_INTENSITY"].value)
        self.assertAlmostEqual(36.06194335, by_code["Q_E_WASTEWATER_INTENSITY"].value)

    def test_english_clean_energy_intensity_rejects_unordered_or_non_revenue_table(self):
        pages = [PageText(20,
            "Indicator Unit 2024 2025 2023\n"
            "Renewable energy consumption Tonnes of standard coal 2 3 1\n"
            "Comprehensive energy consumption Tonnes of standard coal 10 12 8\n"
            "Intensity of comprehensive energy consumption Tonnes of standard coal / tonne output 1 2 3\n"
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "esg_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_CLEAN_ENERGY_INTENSITY"])

    def test_english_current_first_direct_rows_convert_rmb_million_units(self):
        pages = [PageText(164,
            "Environmental\nIndicator Unit 2025 2024 2023\n"
            "Intensity of Scope 1&2 GHG emissions Ton/RMB million 11.20 9.95 8.58\n"
            "Intensity of non-hazardous waste kg/RMB million 38.27 34.84 35.43\n"
            "Intensity of hazardous waste kg/RMB million 15.06 13.85 14.12\n"
            "Intensity of total water consumption Ton/RMB million 16.72 15.36 14.78\n"
            "Total Environmental protection investment\n% 0.56 1.13 0.93\nas % of revenue\n"
        )]
        items = extract_indicator_candidates(pages, "01798.HK", "DATANG RENEW", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(112, values["Q_E_GHG_INTENSITY"])
        self.assertAlmostEqual(.3827, values["Q_E_SOLID_WASTE_INTENSITY"])
        self.assertAlmostEqual(.1506, values["Q_E_HAZ_WASTE_INTENSITY"])
        self.assertAlmostEqual(167.2, values["Q_E_WATER_INTENSITY"])
        self.assertAlmostEqual(.56, values["Q_S_ENV_INVEST_RATE"])

    def test_english_current_year_interleaved_waste_row_uses_intensity_not_total(self):
        pages = [PageText(90,
            "Non-hazardous waste generation data Unit Year 2025\n"
            "Total discharge of non-hazardous waste Tons Intensity Ton/ten thousand RMB\n"
            "13.87\nrevenue 0.000072"
        )]
        items = extract_indicator_candidates(pages, "06828.HK", "BEIJING GAS BLUE SKY", 2025, "url", "esg_report.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_E_SOLID_WASTE_INTENSITY"]
        self.assertAlmostEqual(.072, values[0])

    def test_english_current_first_million_yuan_rows_map_to_methodology_units(self):
        pages = [PageText(34,
            "Environmental Performance Indicators\nIndicator Name Unit 2025 2024 2023\n"
            "GHG emissions intensity (Scope 1 and Scope 2) Tons of Carbon Dioxide Equivalent per "
            "Million Yuan of Revenue 3.23 3.84 4.14\n"
            "Particulate emissions intensity Kg per Million Yuan of Revenue 0.0029 0.0039 0.0130\n"
            "Water consumption intensity Tons per Million Yuan of Revenue 9.96 10.58 11.84\n"
        )]
        items = extract_indicator_candidates(pages, "01600.HK", "TIAN LUN GAS", 2025, "url", "esg_report.pdf")
        by_code = {}
        for item in items:
            by_code.setdefault(item.indicator_code, []).append(item.value)
        self.assertEqual([32.3], by_code["Q_E_GHG_INTENSITY"])
        self.assertEqual(1, len(by_code["Q_E_PM_INTENSITY"]))
        self.assertAlmostEqual(.029, by_code["Q_E_PM_INTENSITY"][0])
        self.assertAlmostEqual(99.6, by_code["Q_E_WATER_INTENSITY"][0])

    def test_english_current_first_resource_table_converts_tce_and_freshwater(self):
        pages = [PageText(66,
            "2023-2025 Resource Usage Data\nType of resources Unit 2025 2024 2023\n"
            "Total volume of integrated Ton of standard coal 1,651,569.72 1,490,274.14 1,331,963.35\n"
            "energy consumption\n"
            "Intensity of integrated energy\nTon of standard coal/\n2.0 1.3 1.10\n"
            "consumption\nRMB10,000\n"
            "Total volume of freshwater\nMillion ton 4.39 5.42 4.85\nconsumption\n"
            "Intensity of freshwater\nconsumption\nRecycling rate of water for\nindustrial use\n"
            "Ton/RMB10,000 5.41 4.67 4.02\n% 98.28 98.28 98.28\n"
            "The intensity data above is calculated by dividing consumption volume by revenue.\n"
        )]
        items = extract_indicator_candidates(
            pages, "06885.HK", "JINMA ENERGY", 2025, "url", "annual_report.pdf",
        )
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(2_000, values["Q_E_ENERGY_INTENSITY"])
        self.assertAlmostEqual(5_410, values["Q_E_WATER_INTENSITY"])

    def test_english_current_first_resource_table_requires_explicit_revenue_note(self):
        pages = [PageText(66,
            "Type of resources Unit 2025 2024 2023\n"
            "Intensity of integrated energy Ton of standard coal/ 2.0 1.3 1.10 "
            "consumption RMB10,000\n"
            "Intensity of freshwater consumption Recycling rate of water for industrial use "
            "Ton/RMB10,000 5.41 4.67 4.02 % 98.28 98.28 98.28\n"
        )]
        items = extract_indicator_candidates(
            pages, "A", "甲", 2025, "url", "annual_report.pdf",
        )
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_E_ENERGY_INTENSITY", "Q_E_WATER_INTENSITY",
        }])

    def test_english_statement_titles_allow_line_breaks_and_bilingual_row_prefixes(self):
        pages = [
            PageText(70,
                "合并财务状况表 Consolidated Statement of\nFinancial Position\n"
                "资产总额 Total assets 1,000,000 900,000\n"
                "负债总额 Total liabilities 400,000 360,000\n"
                "流动资产 Total current assets 300,000 270,000\n"
                "流动负债 Total current liabilities 120,000 100,000\n"
                "存货 Inventories 30,000 25,000\n"
                "应收账款 Trade receivables 50,000 45,000\n"
                "权益总额 Total equity 600,000 540,000\n"),
            PageText(72,
                "合并损益表 Consolidated Statement of\nProfit or Loss\n"
                "收入 Revenue 500,000 400,000\n"
                "经营利润 Operating profit 80,000 60,000\n"),
        ]
        items = extract_indicator_candidates(
            pages, "01733.HK", "E-COMMODITIES", 2025, "url", "annual_report.pdf",
        )
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(40, values["Q_G_DEBT_ASSET_RATE"])
        self.assertAlmostEqual(25, values["Q_G_REVENUE_GROWTH"])
        self.assertAlmostEqual(16, values["Q_G_OPERATING_MARGIN"])

    def test_collapsed_english_income_row_recovers_only_deterministic_revenue_growth(self):
        pages = [PageText(73,
            "Consolidated Statement of Profit or Loss and Other Comprehensive Income\n"
            "Revenue 7 Cost of sales and services provided 11 Gross profit "
            "20,364,482 (15,883,625) 23,147,916 (19,654,507) "
            "4,480,857 3,493,409\n"
        )]
        items = extract_indicator_candidates(
            pages, "00607.HK", "FULLSHARE", 2025, "url", "annual_report.pdf",
        )
        matches = [item for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"]
        self.assertEqual(1, len(matches))
        self.assertAlmostEqual(-12.0245554719, matches[0].value)
        self.assertTrue(
            matches[0].evidence_text.startswith("English collapsed statement row derived:"),
        )

    def test_collapsed_english_income_row_rejects_missing_statement_title(self):
        pages = [PageText(73,
            "Revenue 7 Cost of sales and services provided 11 Gross profit "
            "20,364,482 (15,883,625) 23,147,916 (19,654,507) "
            "4,480,857 3,493,409\n"
        )]
        items = extract_indicator_candidates(
            pages, "A", "甲", 2025, "url", "annual_report.pdf",
        )
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"])

    def test_english_donation_rate_accepts_explicit_revenue_share(self):
        pages = [PageText(20, "Proportion of donation total in revenue 0.03 %")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_DONATION_RATE"]
        self.assertEqual([.03], values)

    def test_english_donation_rate_rejects_amount_and_formula_only(self):
        pages = [PageText(
            20,
            "Corporate donation totalled RMB 60,000. "
            "Proportion of donation total in revenue = donation/revenue.",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_S_DONATION_RATE"])

    def test_english_pay_per_employee_uses_same_group_rmb_paragraph(self):
        pages = [PageText(
            20,
            "As at 31 December 2025, the Group had 12,685 full-time employees, "
            "staff cost of the Group for the Reporting Period amounted to RMB2,267.04 million.",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_PAY_PER_EMPLOYEE"]
        self.assertAlmostEqual(2267.04 * 100 / 12685, values[0])

    def test_english_pay_per_employee_rejects_hkd_and_mismatched_period(self):
        pages = [
            PageText(20, "As at 31 December 2025, the Group had 21,677 employees. Total staff costs amounted to HK$1,459.247 million."),
            PageText(21, "As at 31 December 2024, the Group had 457 employees. Total staff costs was RMB68.8 million."),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_S_PAY_PER_EMPLOYEE"])

    def test_english_employee_benefit_note_derives_rmb_per_capita_metrics(self):
        pages = [
            PageText(20, "As at 31 December 2025, the Group had a total of 2,669 full-time employees."),
            PageText(151, "Notes (Expressed in RMB unless otherwise indicated)\nEmployee benefits payable\n2025 Opening balance Increase during the year Decrease during the year Closing balance\nStaff welfare — 15,459,263.55 15,459,263.55 —\nSocial insurance 22,480.63 27,477,202.96 27,477,202.92 22,480.67\nHousing provident fund (6,271.80) 39,187,497.36 39,116,330.36 64,895.20\nLabour union operating funds and\nstaff education funds 7,021,581.31 11,407,746.42 11,121,875.15 7,307,452.58"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        values = {item.indicator_code: item.value for item in items}
        expected_benefit = (15_459_263.55 + 27_477_202.96 + 39_187_497.36) / 10_000 / 2669
        self.assertAlmostEqual(expected_benefit, values["Q_S_BENEFIT_PER_EMPLOYEE"])
        self.assertAlmostEqual(11_407_746.42 / 10_000 / 2669, values["Q_S_EDU_PER_EMPLOYEE"])

    def test_english_employee_benefit_note_requires_rmb_and_increase_header(self):
        pages = [
            PageText(20, "As at 31 December 2025, the Group had a total of 2,669 full-time employees."),
            PageText(151, "Notes expressed in HKD\nEmployee benefits payable\nStaff welfare 15,459,263.55"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_S_BENEFIT_PER_EMPLOYEE", "Q_S_EDU_PER_EMPLOYEE",
        }])

    def test_english_pollutant_and_waste_intensities_reject_production_and_usd(self):
        pages = [PageText(
            20,
            "NOx emission intensity 2.13 kg/MW of wafers. "
            "Non-hazardous waste generation intensity tonnes/USD million revenue 1.21",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        self.assertFalse([item for item in items if item.indicator_code in {
            "Q_E_NOX_INTENSITY", "Q_E_SOLID_WASTE_INTENSITY",
        }])

    def test_english_rd_intensity_direct_disclosure(self):
        pages = [PageText(
            20,
            "R&D investment intensity was 5.2%, a year-on-year increase of 0.64 percentage point. "
            "The annual R&D investment intensity will be no less than 3% from 2026 to 2030.",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_RD_RATE"]
        self.assertEqual([5.2], values)

    def test_english_income_statement_derives_rd_rate(self):
        pages = [PageText(
            30,
            "Consolidated Statement of Profit or Loss\n"
            "Revenue 10,000 8,000\n"
            "Research and development expenses (800) (600)\n"
            "Operating profit 1,000 900",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_RD_RATE"]
        self.assertEqual([8], values)

    def test_english_full_year_rmb_dividend_per_share(self):
        pages = [
            PageText(20, "Dividend per share (RMB cent) 31.58 31.58"),
            PageText(21, "Dividend per share (RMB) 0.168 0.212 0.132"),
            PageText(22, "The total dividend per share for the whole year amounts to RMB0.358."),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_DIVIDEND_PER_SHARE"]
        self.assertEqual(3, len(values))
        for actual, expected in zip(values, [.3158, .168, .358]):
            self.assertAlmostEqual(expected, actual)

    def test_english_work_safety_investment_rate(self):
        pages = [
            PageText(20, "Proportion of work safety investment to operating revenue % 0.12"),
            PageText(21, "Work safety investment as % of % 4.41 3.63 2.48 revenue"),
            PageText(22, "Proportion of Safety Production Investment (%) = Safety Production Investment/Revenue."),
            PageText(23, "Proportion of safety production investment38 % / 1.77"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "esg.pdf")
        values = [item.value for item in items if item.indicator_code == "Q_S_SAFETY_INVEST_RATE"]
        self.assertEqual([.12, 4.41, 1.77], values)

    def test_english_dividend_rejects_hkd_and_final_only(self):
        pages = [PageText(
            20,
            "Dividend per share (HK$) 3.20. Proposed final dividend per share (RMB) 0.12. "
            "A final dividend of RMB0.12 per share was proposed.",
        )]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "https://source", "annual.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_S_DIVIDEND_PER_SHARE"])

    def test_english_income_rejects_missing_period_and_contaminated_extra_columns(self):
        for text in (
            "Consolidated Statement of Profit or Loss\nRevenue 19 - 29,317",
            "Consolidated Statement of Profit or Loss\nRevenue 5 135,362 104,024 125,294",
        ):
            items = extract_indicator_candidates([PageText(120, text)], "A", "甲", 2025, "url", "annual.pdf")
            self.assertFalse([item for item in items if item.indicator_code == "Q_G_REVENUE_GROWTH"])

    def test_income_and_cashflow_statements_derive_governance_metrics(self):
        pages = [
            PageText(5, "近三年主要会计数据\n（一）主要会计数据\n营业收入 120.00 100.00\n利润总额 20.00"),
            PageText(80, "合并资产负债表\n流动负债合计 50.00 40.00\n资产总计 200.00 180.00"),
            PageText(81, "合并利润表\n营业总成本 70.00 65.00\n其中：利息费用 5.00 4.00\n营业利润 30.00 20.00\n利润总额 25.00 18.00"),
            PageText(82, "母公司利润表"),
            PageText(83, "合并现金流量表\n经营活动现金流入小计 130.00 110.00\n经营活动产生的\n现金流量净额 20.00 18.00"),
            PageText(84, "母公司现金流量表"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(25, values["Q_G_OPERATING_MARGIN"])
        self.assertAlmostEqual(50, values["Q_G_OPERATING_PROFIT_GROWTH"])
        self.assertAlmostEqual(130 / 120 * 100, values["Q_G_CASH_REALIZATION"])
        self.assertAlmostEqual(40, values["Q_G_CASH_CURRENT_LIABILITY"])
        self.assertAlmostEqual(30 / 190 * 100, values["Q_G_ROA"])

    def test_operating_profit_growth_rejects_next_line_amount_as_prior_year(self):
        pages = [
            PageText(5, "近三年主要会计数据\n营业收入 120.00 100.00\n利润总额 20.00"),
            PageText(80, "合并资产负债表\n资产总计 200.00 180.00"),
            PageText(81, "合并利润表\n营业利润 3000000.00\n加：营业外收入 1000.00\n利润总额 3001000.00"),
            PageText(82, "母公司利润表"),
        ]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "url", "annual_report.pdf")
        self.assertFalse([item for item in items if item.indicator_code == "Q_G_OPERATING_PROFIT_GROWTH"])

    def test_review_summary_only_recommends_consistent_values(self):
        items = [
            Observation("A", "甲", 2025, "Q_G_ROE", 10, source_page=1),
            Observation("A", "甲", 2025, "Q_G_ROE", 9, source_page=2),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 40, source_page=3),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 40, source_page=4),
        ]
        summaries = {item.indicator_code: item for item in summarize_review_candidates(items)}
        self.assertEqual("", summaries["Q_G_ROE"].recommended_value)
        self.assertEqual("40", summaries["Q_G_DEBT_ASSET_RATE"].recommended_value)

    def test_pending_resolver_confirms_only_consistent_annual_direct_values(self):
        annual = Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING, source_file="annual_report.pdf", confidence=.92)
        duplicate = Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING, source_file="annual_report.pdf", confidence=.92)
        environmental = Observation("A", "甲", 2025, "Q_E_GHG_INTENSITY", 2, ValueStatus.PENDING, source_file="esg_report.pdf", confidence=.9)
        confirmed, unresolved, decisions = resolve_pending_candidates([annual, duplicate, environmental])
        self.assertEqual(2, len(confirmed))
        self.assertTrue(all(item.status == ValueStatus.CONFIRMED for item in confirmed))
        self.assertFalse(unresolved)
        self.assertEqual({"auto_confirmed"}, {item.decision for item in decisions})

    def test_pending_resolver_accepts_debt_ratio_rounding_tolerance(self):
        items = [
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 54.22, ValueStatus.PENDING, source_file="annual_report.pdf", confidence=.9),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 54.2174, ValueStatus.PENDING, source_file="annual_report.pdf", confidence=.94),
        ]
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertEqual([54.2174], [item.value for item in confirmed])
        self.assertFalse(unresolved)

    def test_pending_resolver_prefers_statement_derived_year_end_debt_ratio(self):
        items = [
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 53.76, ValueStatus.PENDING, source_file="annual_report.pdf", evidence_text="期初", confidence=.9),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 57.86, ValueStatus.PENDING, source_file="annual_report.pdf", evidence_text="期末", confidence=.9),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 57.8641, ValueStatus.PENDING, source_file="annual_report.pdf", evidence_text="合并报表自动派生: 负债/资产", confidence=.94),
        ]
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertEqual([57.8641], [item.value for item in confirmed])
        self.assertFalse(unresolved)

    def test_pending_resolver_confirms_consistent_official_esg_intensity(self):
        item = Observation("A", "甲", 2025, "Q_E_GHG_INTENSITY", 6.68, ValueStatus.PENDING, source_file="esg_report.pdf", confidence=.82)
        confirmed, unresolved, _ = resolve_pending_candidates([item])
        self.assertEqual([6.68], [value.value for value in confirmed])
        self.assertFalse(unresolved)

    def test_review_tiers_separate_auto_conflict_and_manual_spot_check(self):
        items = [
            Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING,
                        source_file="annual_report.pdf", source_page=1, confidence=.92),
            Observation("B", "乙", 2025, "Q_S_SAFETY_INVEST_RATE", 2, ValueStatus.PENDING,
                        source_file="esg_report.pdf", source_page=2, confidence=.94),
            Observation("B", "乙", 2025, "Q_S_SAFETY_INVEST_RATE", 3, ValueStatus.PENDING,
                        source_file="esg_report.pdf", source_page=3, confidence=.94),
            Observation("C", "丙", 2025, "Q_S_ENV_INVEST_RATE", 1, ValueStatus.PENDING,
                        source_file="esg_report.pdf", source_page=4, confidence=.94),
        ]
        rows, summary = plan_review_tiers(items)
        tiers = {(item.company_code, item.tier) for item in rows}
        self.assertIn(("A", "auto_policy_eligible"), tiers)
        self.assertIn(("B", "manual_signature_required"), tiers)
        self.assertIn(("C", "single_candidate_review"), tiers)
        self.assertEqual(3, summary["candidate_group_count"])
        self.assertFalse(summary["applicable"])

    def test_pending_resolver_v6_requires_strict_evidence_prefix_for_new_metrics(self):
        eligible = Observation(
            "A", "甲", 2025, "Q_G_QUICK_RATIO", 150, ValueStatus.PENDING,
            source_file="annual_report.pdf", confidence=.94,
            evidence_text="English consolidated statement derived: current assets | inventory | liabilities",
        )
        unverified = Observation(
            "B", "乙", 2025, "Q_G_QUICK_RATIO", 160, ValueStatus.PENDING,
            source_file="annual_report.pdf", confidence=.94,
            evidence_text="management discussion ratio",
        )
        confirmed, unresolved, decisions = resolve_pending_candidates([eligible, unverified])
        self.assertEqual([150], [item.value for item in confirmed])
        self.assertEqual([160], [item.value for item in unresolved])
        self.assertEqual("public-disclosure-v6", decisions[0].policy_version)
        self.assertEqual("strict_extraction_evidence_consistent", decisions[0].reason)

    def test_pending_resolver_accepts_strict_annual_resource_intensity_row(self):
        energy = Observation(
            "A", "甲", 2025, "Q_E_ENERGY_INTENSITY", 2_000, ValueStatus.PENDING,
            source_file="annual_report.pdf", confidence=.96,
            evidence_text="English current-first revenue resource row: "
                          "Intensity of integrated energy Ton of standard coal/ 2.0 1.3 1.10",
        )
        unverified = Observation(
            "B", "乙", 2025, "Q_E_ENERGY_INTENSITY", 3_000, ValueStatus.PENDING,
            source_file="annual_report.pdf", confidence=.96,
            evidence_text="Energy intensity 3.0",
        )
        confirmed, unresolved, _ = resolve_pending_candidates([energy, unverified])
        self.assertEqual([2_000], [item.value for item in confirmed])
        self.assertEqual([3_000], [item.value for item in unresolved])

    def test_resolution_preview_audit_closes_auto_and_manual_groups(self):
        auto = Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING,
                           source_file="annual_report.pdf", confidence=.92)
        manual = [
            Observation("B", "乙", 2025, "Q_S_SAFETY_INVEST_RATE", value, ValueStatus.PENDING,
                        source_file="esg_report.pdf", source_page=page, confidence=.94)
            for value, page in ((2, 2), (3, 3))
        ]
        confirmed, unresolved, decisions = resolve_pending_candidates([auto, *manual])
        report = audit_resolution_preview([auto, *manual], confirmed, unresolved, decisions)
        self.assertTrue(report["valid"])
        self.assertFalse(report["freeze_ready"])
        self.assertEqual(1, report["auto_confirmed_group_count"])
        self.assertEqual(1, report["manual_required_group_count"])

    def test_resolution_preview_audit_rejects_selected_value_tamper(self):
        candidate = Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING,
                                source_file="annual_report.pdf", confidence=.92)
        confirmed, unresolved, decisions = resolve_pending_candidates([candidate])
        decision = decisions[0]
        tampered = ResolutionDecision(**{**vars(decision), "selected_value": "11"})
        with self.assertRaisesRegex(ValueError, "selected value drift"):
            audit_resolution_preview([candidate], confirmed, unresolved, [tampered])

    def test_manual_review_selection_uses_complete_frozen_tiers(self):
        candidates = [
            Observation("A", "甲", 2025, "Q_G_ROE", 10, ValueStatus.PENDING,
                        source_file="annual_report.pdf", confidence=.92),
            Observation("B", "乙", 2025, "Q_S_SAFETY_INVEST_RATE", 2),
            Observation("B", "乙", 2025, "Q_S_SAFETY_INVEST_RATE", 3),
        ]
        tiers, _ = plan_review_tiers(candidates)
        selected = select_manual_review_candidates(candidates, tiers)
        self.assertEqual([2, 3], [item.value for item in selected])
        with self.assertRaisesRegex(ValueError, "exactly match"):
            select_manual_review_candidates(candidates, tiers[:1])

    def test_manual_review_requires_selected_candidate(self):
        candidate = Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 40, ValueStatus.PENDING, source_file="annual.pdf", source_page=3)
        instruction = ReviewInstruction("A", 2025, "Q_G_DEBT_ASSET_RATE", "confirm", "40", "reviewer", "2026-07-28T12:00:00+08:00", "核对年末值")
        confirmed, unresolved = apply_review_instructions([candidate], [instruction])
        self.assertEqual(1, len(confirmed))
        self.assertFalse(unresolved)
        self.assertIn("manual-review:reviewer", confirmed[0].evidence_text)

    def test_conflict_review_confirms_or_rejects_with_audit(self):
        conflict = [
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 40, ValueStatus.PENDING, source_page=3, confidence=.9),
            Observation("A", "甲", 2025, "Q_G_DEBT_ASSET_RATE", 42, ValueStatus.PENDING, source_page=4, confidence=.95),
        ]
        other = Observation("B", "乙", 2025, "Q_G_DEBT_ASSET_RATE", 30, ValueStatus.PENDING)
        signed = ReviewInstruction("A", 2025, "Q_G_DEBT_ASSET_RATE", "confirm", "42", "alice", "2026-07-30T09:00:00+08:00", "核对报表")
        confirmed, unresolved, audits = apply_conflict_review_instructions(conflict + [other], [signed])
        self.assertEqual([42], [item.value for item in confirmed])
        self.assertEqual(["B"], [item.company_code for item in unresolved])
        self.assertEqual(("confirm", "40|42", "3|4"),
                         (audits[0].action, audits[0].candidate_values, audits[0].source_pages))
        rejected = ReviewInstruction("A", 2025, "Q_G_DEBT_ASSET_RATE", "reject", "", "bob", "2026-07-30T10:00:00+08:00", "口径均不适用")
        confirmed, unresolved, audits = apply_conflict_review_instructions(conflict, [rejected])
        self.assertFalse(confirmed)
        self.assertFalse(unresolved)
        self.assertEqual("reject", audits[0].action)

    def test_review_instruction_requires_timezone_and_note(self):
        header = "company_code,report_year,indicator_code,action,selected_value,reviewer,reviewed_at,note\n"
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "review.csv"
            path.write_text(header + "A,2025,Q_G_ROE,confirm,10,alice,2026-07-30T09:00:00,核对\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "必须包含时区"):
                read_review_instructions(path)
            path.write_text(header + "A,2025,Q_G_ROE,reject,,alice,2026-07-30T09:00:00+08:00,\n", encoding="utf-8")
            with self.assertRaisesRegex(ValueError, "必须填写"):
                read_review_instructions(path)

    def test_quality_gate_rejects_incomplete_dataset(self):
        item = Observation("A", "甲公司", 2025, "Q_G_ROE", 10)
        report = evaluate_quality([item], self.methodology, expected_companies=2)
        self.assertFalse(report.publishable)
        self.assertIn("UNIVERSE_MISMATCH", {issue.code for issue in report.issues})

    def test_universe_audit_deduplicates_entities_and_blocks_partial_release(self):
        companies = [
            UniverseCompany("600001.SH", "甲", "SSE", "电力", entity_id="ENTITY-A"),
            UniverseCompany("00001.HK", "甲", "HKEX", "电力", False, "A/H去重", "ENTITY-A"),
            UniverseCompany("000002.SZ", "乙", "SZSE", "能源设备", entity_id="ENTITY-B"),
        ]
        report = audit_universe(companies, 632, ["600001.SH"])
        self.assertEqual(3, report.security_count)
        self.assertEqual(2, report.included_company_count)
        self.assertEqual(1, report.completed_company_count)
        self.assertAlmostEqual(1 / 632, report.completed_coverage_rate)
        self.assertEqual(("BSE", "HKEX"), report.missing_exchanges)
        self.assertEqual(2, report.missing_source_count)
        self.assertEqual(2, report.missing_date_count)
        self.assertFalse(report.publishable)

    def test_universe_audit_requires_provenance_and_consistent_decisions(self):
        companies = [
            UniverseCompany(
                "600001.SH", "甲", "SSE", "电力", entity_id="ENTITY-A",
                source_url="https://official/sse", as_of_date="2026-07-29",
            ),
            UniverseCompany(
                "00001.HK", "甲H", "HKEX", "电力", entity_id="ENTITY-A",
                source_url="https://official/hkex", as_of_date="2026-07-29",
            ),
            UniverseCompany("000002.SZ", "乙", "SZSE", "电力", exclusion_reason="不应存在"),
            UniverseCompany("920001.BJ", "丙", "BSE", "电力", False),
        ]
        report = audit_universe(companies, 2, ["600001.SH", "00001.HK"])
        self.assertEqual(1, report.duplicate_included_entity_count)
        self.assertEqual(1, report.included_with_exclusion_reason_count)
        self.assertEqual(1, report.excluded_without_reason_count)
        self.assertFalse(report.publishable)

    def test_universe_audit_allows_complete_evidenced_release(self):
        companies = [
            UniverseCompany(
                "600001.SH", "甲", "SSE", "电力", entity_id="A",
                source_url="https://official/sse", as_of_date="2026-07-29",
            ),
            UniverseCompany(
                "000002.SZ", "乙", "SZSE", "电力", entity_id="B",
                source_url="https://official/szse", as_of_date="2026-07-29",
            ),
            UniverseCompany(
                "920001.BJ", "丙", "BSE", "电力", entity_id="C",
                source_url="https://official/bse", as_of_date="2026-07-29",
            ),
            UniverseCompany(
                "00001.HK", "丁", "HKEX", "电力", entity_id="D",
                source_url="https://official/hkex", as_of_date="2026-07-29",
            ),
        ]
        report = audit_universe(companies, 4, [item.stock_code for item in companies])
        self.assertTrue(report.publishable)

    def test_universe_audit_treats_review_placeholders_as_unclassified(self):
        company = UniverseCompany(
            "600001.SH", "甲", "SSE", "历史能源样本待复核", entity_id="A",
            source_url="https://official", as_of_date="2026-07-29",
        )
        report = audit_universe([company], 1, [company.stock_code], required_exchanges=("SSE",))
        self.assertEqual(1, report.unclassified_count)
        self.assertFalse(report.publishable)
        seed = UniverseCompany(
            "600002.SH", "乙", "SSE", "参考榜单能源种子待行业复核", entity_id="B",
            source_url="https://official", as_of_date="2026-07-29",
        )
        self.assertEqual(1, audit_universe([seed], 1, [], required_exchanges=("SSE",)).unclassified_count)

    def test_universe_builder_filters_st_non_energy_and_ah_duplicates(self):
        rows = [
            ExchangeSecurity("00001.HK", "甲电力", "HKEX", "电力", "ENTITY-A"),
            ExchangeSecurity("600001.SH", "甲电力", "SSE", "电力", "ENTITY-A"),
            ExchangeSecurity("000002.SZ", "*ST乙", "SZSE", "能源设备", "ENTITY-B"),
            ExchangeSecurity("920001.BJ", "丙软件", "BSE", "软件", "ENTITY-C"),
            ExchangeSecurity("920002.BJ", "丁储能", "BSE", "其他", "ENTITY-D", energy_eligible="true"),
        ]
        decisions = build_energy_universe(rows)
        self.assertEqual([False, True, False, False, True], [item.included for item in decisions])
        self.assertIn("保留600001.SH", decisions[0].exclusion_reason)
        self.assertEqual("ST/*ST排除", decisions[2].exclusion_reason)
        self.assertEqual("行业待复核:软件", decisions[3].exclusion_reason)

    def test_exchange_snapshot_round_trip_to_universe(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "snapshot.csv"
            snapshot.write_text(
                "stock_code,company_name,exchange,industry,entity_id,st_status,listing_status,energy_eligible,source_url,as_of_date\n"
                "600900.SH,长江电力,SSE,电力,ENTITY-A,,上市,true,https://official,2026-07-28\n",
                encoding="utf-8",
            )
            output = Path(directory) / "universe.csv"
            write_universe(output, build_energy_universe(read_exchange_snapshot(snapshot)))
            companies = __import__("aegis_esg.universe", fromlist=["read_universe"]).read_universe(output)
            self.assertEqual("ENTITY-A", companies[0].entity_id)
            self.assertTrue(companies[0].included)

    def test_normalize_chinese_exchange_export_and_codes(self):
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "official.csv"
            source.write_text(
                "证券代码,证券简称,所属行业,上市状态,ST状态\n"
                "9001,测试能源,电力,上市,\n",
                encoding="utf-8",
            )
            rows = normalize_exchange_export(source, "BSE", "https://www.bse.cn/list", "2026-07-28")
            self.assertEqual("009001.BJ", rows[0].stock_code)
            self.assertEqual("电力", rows[0].industry)
            self.assertTrue(audit_snapshot(rows).valid)

    def test_stock_code_normalization_for_four_exchanges(self):
        self.assertEqual("600900.SH", normalize_stock_code("600900.ss", "SSE"))
        self.assertEqual("000001.SZ", normalize_stock_code("1", "SZSE"))
        self.assertEqual("920001.BJ", normalize_stock_code("920001", "BSE"))
        self.assertEqual("02688.HK", normalize_stock_code("2688.HK", "HKEX"))
        with self.assertRaises(ValueError):
            normalize_stock_code("ABC", "SSE")

    def test_snapshot_quality_requires_provenance(self):
        row = ExchangeSecurity("600900.SH", "长江电力", "SSE", "电力", "ENTITY")
        quality = audit_snapshot([row])
        self.assertFalse(quality.valid)
        self.assertEqual(1, quality.missing_source_count)
        self.assertEqual(1, quality.missing_date_count)

    def test_snapshot_quality_rejects_duplicates_and_invalid_provenance(self):
        rows = [
            ExchangeSecurity("600900.SH", "甲", "SSE", "电力", "A", source_url="ftp://bad", as_of_date="2026/07/28"),
            ExchangeSecurity("600900.SH", "甲", "SSE", "电力", "A", source_url="https://official", as_of_date="2026-07-28"),
        ]
        quality = audit_snapshot(rows)
        self.assertEqual(1, quality.duplicate_code_count)
        self.assertEqual(1, quality.invalid_source_count)
        self.assertEqual(1, quality.invalid_date_count)
        self.assertFalse(quality.valid)

    def test_official_listing_pagination_is_complete(self):
        payloads = {
            1: '{"data":[{"zqdm":"920001","zqjc":"甲电力","hymc":"电力"}],"pageNo":1,"pageCount":2,"total":2}',
            2: '{"data":[{"zqdm":"920002","zqjc":"乙设备","hymc":"能源设备"}],"pageNo":2,"pageCount":2,"total":2}',
        }
        rows = collect_listing_pages("BSE", "https://official/list", "2026-07-28", payloads.__getitem__)
        self.assertEqual(["920001.BJ", "920002.BJ"], [item.stock_code for item in rows])
        self.assertEqual("电力", rows[0].industry)

    def test_hkex_full_list_filters_non_company_securities(self):
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "ListOfSecurities.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["List of Securities"])
            sheet.append(["Updated as at 29/07/2026"])
            sheet.append(["Stock Code", "Name of Securities", "Category", "Sub-Category"])
            sheet.append(["1", "CKH HOLDINGS", "Equity", "Equity Securities (Main Board)"])
            sheet.append(["8100", "GET NICE", "Equity", "Equity Securities (GEM)"])
            sheet.append(["2800", "TRACKER FUND", "Exchange Traded Products", "Exchange Traded Funds"])
            sheet.append(["10001", "DERIVATIVE", "Derivative Warrants", None])
            workbook.save(source)
            rows, as_of_date = import_hkex_securities(
                source, "https://www.hkex.com.hk/ListOfSecurities.xlsx", "2026-07-29",
            )
            self.assertEqual("2026-07-29", as_of_date)
            self.assertEqual(["00001.HK", "08100.HK"], [item.stock_code for item in rows])
            self.assertTrue(audit_snapshot(rows).valid)

    def test_hkex_full_list_rejects_date_mismatch(self):
        from openpyxl import Workbook
        with tempfile.TemporaryDirectory() as directory:
            source = Path(directory) / "ListOfSecurities.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["List of Securities"])
            sheet.append(["Updated as at 28/07/2026"])
            sheet.append(["Stock Code", "Name of Securities", "Category", "Sub-Category"])
            sheet.append(["1", "CKH HOLDINGS", "Equity", "Equity Securities (Main Board)"])
            workbook.save(source)
            with self.assertRaisesRegex(ValueError, "日期不一致"):
                import_hkex_securities(source, "https://official", "2026-07-29")

    def test_hkex_profile_parser_validates_token_code_and_evidence(self):
        html = (
            "<script>LabCI.getToken = function () {\n"
            "//return \"Base64-AES-Encrypted-Token\";\n"
            "return \"abcdefghijklmnopqrstuvwxyz123456\";\n};</script>"
        )
        self.assertEqual("abcdefghijklmnopqrstuvwxyz123456", parse_hkex_access_token(html))
        encoded_html = "LabCI.getToken = function () { return 'abcdefghijklmnopqrstuv%2B1234'; };"
        self.assertEqual("abcdefghijklmnopqrstuv+1234", parse_hkex_access_token(encoded_html))
        payload = (
            'cb({"data":{"responsecode":"000","quote":{"sym":"2688","nm":"新奧能源控股有限公司",'
            '"nm_s":"新奧能源","summary":"主要从事销售及分销管道燃气。",'
            '"hsic_ind_classification":"公用事業 - 公用事業",'
            '"hsic_sub_sector_classification":"燃氣供應","db_updatetime":"2026年7月29日09:48"}}})'
        )
        profile = parse_hkex_quote_payload(payload, "02688.HK")
        self.assertEqual("新奧能源控股有限公司", profile.chinese_name)
        self.assertEqual("燃氣供應", profile.hsic_sub_sector)
        self.assertEqual("candidate", profile.evidence_status)
        with self.assertRaises(ValueError):
            parse_hkex_quote_payload(payload, "00003.HK")

    def test_hkex_profile_collector_reuses_page_token_and_preserves_raw_payload(self):
        html = "LabCI.getToken = function () { return 'abcdefghijklmnopqrstuvwxyz123456'; };"
        payloads = {
            "2688": 'aegisHKEX({"data":{"responsecode":"000","quote":{"sym":"2688","nm":"新奧能源",'
                    '"summary":"燃气业务","hsic_sub_sector_classification":"燃氣供應"}}})',
            "135": 'aegisHKEX({"data":{"responsecode":"000","quote":{"sym":"135","nm":"昆侖能源",'
                   '"summary":"能源业务","hsic_sub_sector_classification":"油氣生產商"}}})',
        }
        calls = []

        def fetch(url):
            calls.append(url)
            if "getequityquote" not in url:
                return html
            query = urllib.parse.parse_qs(urllib.parse.urlparse(url).query)
            return payloads[query["sym"][0]]

        profiles, raw = collect_hkex_issuer_profiles(["02688.HK", "00135.HK"], fetch)
        self.assertEqual(["02688.HK", "00135.HK"], [item.stock_code for item in profiles])
        self.assertEqual(3, len(calls))
        self.assertEqual({"02688.HK", "00135.HK"}, set(raw))

    def test_hkex_disclosure_parses_exact_stock_and_relevant_documents(self):
        lookup = 'callback({"stockInfo":[{"stockId":42,"code":"00042","name":"NEE"}]});'
        self.assertEqual(("42", "NEE"), parse_stock_lookup(lookup, "00042.HK"))
        page = """<html><title>Listed Company Information Title Search</title>
        <div>Total records found: 2</div><table><tbody>
        <tr><td class='release-time'>Release Time: 30/04/2026 18:01</td>
        <td class='stock-short-code'>Stock Code: 00042</td><td class='stock-short-name'>NEE</td>
        <td><div class='headline'>Financial Statements/ESG Information</div><div class='doc-link'>
        <a href='/listedco/annual.pdf'>2025 ANNUAL REPORT</a></div></td></tr>
        <tr><td class='release-time'>Release Time: 01/03/2026 09:00</td>
        <td class='stock-short-code'>00042</td><td class='stock-short-name'>NEE</td>
        <td><div class='headline'>Monthly Returns</div><div class='doc-link'>
        <a href='/listedco/monthly.pdf'>MONTHLY RETURN</a></div></td></tr>
        </tbody></table></html>"""
        rows = parse_title_search(page, "00042.HK", "42")
        self.assertEqual(1, len(rows))
        self.assertEqual("annual_report", rows[0].document_type)
        self.assertEqual(2025, rows[0].report_year)
        self.assertEqual("2026-04-30", rows[0].published_date)

    def test_hkex_disclosure_rejects_incomplete_page_and_classifies_listing(self):
        page = "<html><title>Listed Company Information Title Search</title><div>Total records found: 1</div></html>"
        with self.assertRaisesRegex(ValueError, "分页不完整"):
            parse_title_search(page, "00600.HK", "1")
        self.assertEqual(("listing_document", 0), classify_continuity_document("GLOBAL OFFERING", "Listing Documents"))
        self.assertEqual(("annual_report", 2025), classify_continuity_document(
            "2024/25 ANNUAL REPORT", "Financial Statements/ESG Information",
        ))
        self.assertEqual(("annual_report", 2026), classify_continuity_document(
            "Annual Report 2025/2026", "Financial Statements/ESG Information",
        ))
        self.assertEqual((None, 0), classify_continuity_document(
            "Letter to shareholders - publication of Annual Report 2025", "Circulars - [Other]",
        ))

    def test_hkex_disclosure_bisects_incomplete_date_range(self):
        lookup = b'callback({"stockInfo":[{"stockId":42,"code":"00042","name":"NEE"}]});'

        def page(total, suffix):
            row = "" if total == 101 else f"""<tr><td class='release-time'>01/01/2026</td>
            <td class='stock-short-code'>00042</td><td class='stock-short-name'>NEE</td>
            <td><div class='headline'>Listing Documents</div><div class='doc-link'>
            <a href='/listedco/{suffix}.pdf'>GLOBAL OFFERING</a></div></td></tr>"""
            return f"<html><title>Listed Company Information Title Search</title><div>Total records found: {total}</div><table>{row}</table></html>".encode()

        calls = []

        def fetch(request):
            if request.data is None:
                return lookup
            calls.append(urllib.parse.parse_qs(request.data.decode()))
            if len(calls) == 1:
                return page(101, "full")
            return page(1, str(len(calls)))

        rows, raw = discover_hkex_continuity_documents("00042.HK", "2026-01-01", "2026-01-10", fetch)
        self.assertEqual(2, len(rows))
        self.assertEqual(3, len(calls))
        self.assertEqual(3, len(raw["title_search_ranges"]))

    def test_hkex_continuity_download_selection_avoids_overwrite(self):
        def row(year, kind, url, date="2026-01-01", headline="Financial Statements"):
            return HKEXDisclosure("00600.HK", "AXERA", year, kind, url, date, "title", headline, "1")

        selected, summary = select_continuity_downloads([
            row(2024, "annual_report", "https://hkex/old.pdf", "2025-04-01"),
            row(2025, "annual_report", "https://hkex/new.pdf", "2026-04-01"),
            row(0, "listing_document", "https://hkex/notice.pdf", headline="Formal Notice"),
            row(0, "listing_document", "https://hkex/prospectus.pdf", headline="Listing Documents"),
        ])
        self.assertEqual(["annual_report", "listing_document"], [item.document_type for item in selected])
        self.assertEqual("https://hkex/new.pdf", selected[0].source_url)
        self.assertEqual("https://hkex/prospectus.pdf", selected[1].source_url)
        self.assertEqual(0, summary["duplicate_target_count"])
        target, _ = select_continuity_downloads([
            row(2024, "annual_report", "https://hkex/old.pdf"),
            row(2025, "annual_report", "https://hkex/new.pdf"),
        ], report_year=2024)
        self.assertEqual("https://hkex/old.pdf", target[0].source_url)

    def test_hkex_discovery_batch_checkpoints_and_resumes_failures(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            output, raw, failures = root / "found.csv", root / "raw.json.gz", root / "failures.csv"
            calls = []

            def discover(code, start, end):
                calls.append(code)
                if code == "00002.HK" and calls.count(code) == 1:
                    raise RuntimeError("temporary")
                row = HKEXDisclosure(code, code, 2025, "annual_report", f"https://hkex/{code}.pdf", "2026-04-01", "Annual Report 2025", "Financial Statements", code)
                return [row], {"stock_lookup": code, "title_search_ranges": ["html"]}

            _, first_failures, first_summary = discover_hkex_continuity_batch(
                ["00001.HK", "00002.HK"], "2025-01-01", "2026-07-29",
                output, raw, failures, 0, False, discover,
            )
            self.assertEqual(1, len(first_failures))
            self.assertEqual(1, first_summary["completed_count"])
            rows, second_failures, second_summary = discover_hkex_continuity_batch(
                ["00001.HK", "00002.HK"], "2025-01-01", "2026-07-29",
                output, raw, failures, 0, True, discover,
            )
            self.assertEqual(["00001.HK", "00002.HK", "00002.HK"], calls)
            self.assertEqual(2, len(rows))
            self.assertFalse(second_failures)
            self.assertTrue(second_summary["complete"])

    def test_hkex_fetch_failure_preserves_underlying_error(self):
        request = urllib.request.Request("https://www1.hkexnews.hk/search/prefix.do")
        with patch("aegis_esg.sources.hkex_disclosure.urllib.request.urlopen", side_effect=TimeoutError("TLS timeout")), patch(
            "aegis_esg.sources.hkex_disclosure.time.sleep",
        ):
            with self.assertRaisesRegex(RuntimeError, "TimeoutError: TLS timeout"):
                _fetch(request)

    def test_extract_hkex_continuity_evidence_keeps_page_and_pending_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            index = root / "index.csv"
            text_root = root / "text"
            text_path = text_root / "00042.HK/2025/annual_report.txt"
            text_path.parent.mkdir(parents=True)
            index.write_text(
                "company_code,company_name,report_year,document_type,source_url,local_path\n"
                "00042.HK,NEE,2025,annual_report,https://hkex/report.pdf,data/raw/00042.HK/2025/annual_report.pdf\n",
                encoding="utf-8",
            )
            text_path.write_text(
                "\n=== PAGE 1 ===\nThe Company was formerly known as Old Name Limited.\n"
                "\n=== PAGE 2 ===\nPrincipal activities include energy equipment. The H shares are listed in Hong Kong.\n",
                encoding="utf-8",
            )
            rows, summary = extract_continuity_evidence_candidates(index, text_root)
            self.assertEqual({"issuer_history", "principal_business", "ah_identity"}, {item.evidence_category for item in rows})
            self.assertTrue(all(item.review_status == "pending" for item in rows))
            self.assertEqual({1, 2}, {item.source_page for item in rows})
            self.assertFalse(summary["applicable"])
            self.assertTrue(all(item.candidate_id.startswith("HKCE-00042-HK-") for item in rows))

    def test_prepare_continuity_review_packets_stays_unsigned(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            tasks = root / "tasks.csv"
            candidates = root / "candidates.csv"
            tasks.write_text(
                "task_id,stock_code,priority,next_action,historical_name,current_chinese_name,profile_evidence_url\n"
                "T1,00042.HK,0,review_issuer_identity_and_industry,旧名,新名,https://hkex/profile\n",
                encoding="utf-8",
            )
            candidates.write_text(
                "candidate_id,company_code,evidence_category,source_page\n"
                "C1,00042.HK,issuer_history,143\nC2,00042.HK,principal_business,18\n",
                encoding="utf-8",
            )
            packets, summary = prepare_continuity_review_packets(tasks, candidates)
            self.assertEqual("C1", packets[0].issuer_history_candidate_ids)
            self.assertEqual("143", packets[0].issuer_history_pages)
            self.assertEqual("", packets[0].outcome)
            self.assertEqual("unsigned", packets[0].review_status)
            self.assertFalse(summary["applicable"])

    def test_finalize_continuity_review_validates_candidate_ownership_and_signature(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            packets = root / "packets.csv"
            candidates = root / "candidates.csv"
            header = (
                "decision_id,stock_code,outcome,related_a_code,entity_id,selected_candidate_ids,"
                "evidence_url,evidence_date,reviewer,reviewed_at,rationale,review_status\n"
            )
            packets.write_text(
                header + "D1,00042.HK,same_issuer,,,C1,https://hkex/report.pdf,2026-04-29,alice,"
                "2026-07-29T10:00:00+08:00,官方年报确认,signed\n",
                encoding="utf-8",
            )
            candidates.write_text(
                "candidate_id,company_code,evidence_category,source_url\n"
                "C1,00042.HK,issuer_history,https://hkex/report.pdf\n"
                "C2,00600.HK,issuer_history,https://hkex/other.pdf\n",
                encoding="utf-8",
            )
            decisions, audits, summary = finalize_continuity_reviews(packets, candidates)
            self.assertEqual("same_issuer", decisions[0].outcome)
            self.assertEqual("issuer_history", audits[0].selected_categories)
            self.assertTrue(summary["complete"])
            packets.write_text(
                header + "D1,00042.HK,same_issuer,,,C2,https://hkex/other.pdf,2026-04-29,alice,"
                "2026-07-29T10:00:00+08:00,跨证券候选,signed\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "候选ID未匹配本证券"):
                finalize_continuity_reviews(packets, candidates)

    def test_render_continuity_review_guide_keeps_review_unsigned(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            packets = root / "packets.csv"
            candidates = root / "candidates.csv"
            packets.write_text(
                "task_id,stock_code,priority,next_action,historical_name,current_chinese_name,profile_evidence_url,review_status\n"
                "T1,00042.HK,0,review_identity,旧名,新名,https://hkex/profile,unsigned\n",
                encoding="utf-8",
            )
            candidates.write_text(
                "candidate_id,company_code,evidence_category,document_type,report_year,source_url,source_page,evidence_text,confidence\n"
                "C1,00042.HK,issuer_history,annual_report,2025,https://hkex/report.pdf,143,Formerly known as Old Name,0.94\n",
                encoding="utf-8",
            )
            guide, summary = render_continuity_review_guide(packets, candidates)
            self.assertIn("## 00042.HK", guide)
            self.assertIn("`C1`", guide)
            self.assertIn("Formerly known as Old Name", guide)
            self.assertEqual(0, summary["signed_count"])
            self.assertFalse(summary["applicable"])

    def test_select_continuity_review_batch_filters_packets_and_candidates(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            packets = root / "packets.csv"
            candidates = root / "candidates.csv"
            packets.write_text(
                "stock_code,priority,review_status\n"
                "00042.HK,0,unsigned\n00002.HK,2,unsigned\n",
                encoding="utf-8",
            )
            candidates.write_text(
                "candidate_id,company_code,evidence_category\n"
                "C1,00042.HK,issuer_history\nC2,00002.HK,principal_business\n",
                encoding="utf-8",
            )
            selected_packets, selected_candidates, summary = select_continuity_review_batch(
                packets, candidates, 0,
            )
            self.assertEqual(["00042.HK"], [row["stock_code"] for row in selected_packets])
            self.assertEqual(["C1"], [row["candidate_id"] for row in selected_candidates])
            self.assertEqual(1, summary["packet_count"])
            self.assertEqual(0, summary["signed_count"])
            self.assertFalse(summary["applicable"])

    def test_prepare_hkex_evidence_drafts_uses_exact_mapping_and_stays_unsigned(self):
        header = (
            "stock_code,chinese_name,chinese_short_name,company_summary,hsic_industry,hsic_sub_sector,"
            "csic_classification,listing_category,primary_market,incorporation_place,profile_updated_at,"
            "source_url,evidence_status\n"
        )
        with tempfile.TemporaryDirectory() as directory:
            profiles = Path(directory) / "profiles.csv"
            mapping = Path(directory) / "mapping.json"
            profiles.write_text(
                header
                + "02688.HK,新奧能源控股有限公司,新奧能源,燃气业务,公用事業,燃氣供應,,主要上市,,开曼,2026年7月29日,https://hkex/2688,candidate\n"
                + "00001.HK,甲公司,甲,工业设备,工業,工業零件及器材,,主要上市,,香港,2026年7月29日,https://hkex/1,candidate\n",
                encoding="utf-8",
            )
            mapping.write_text(
                '{"version":"v1","exact_sub_sector_mappings":{"燃氣供應":"燃气"}}', encoding="utf-8",
            )
            companies = [
                UniverseCompany("02688.HK", "ENN ENERGY", "HKEX", "待分类", entity_id="02688.HK"),
                UniverseCompany("00001.HK", "甲", "HKEX", "待分类", entity_id="00001.HK"),
            ]
            drafts, summary = prepare_hkex_evidence_drafts(
                profiles, companies, mapping, "2026-07-29",
            )
            by_code = {item.stock_code: item for item in drafts}
            self.assertEqual("燃气", by_code["02688.HK"].sub_industry)
            self.assertEqual("include", by_code["02688.HK"].decision)
            self.assertEqual("", by_code["02688.HK"].reviewer)
            self.assertEqual("manual_review", by_code["00001.HK"].review_status)
            self.assertEqual("", by_code["00001.HK"].decision)
            self.assertEqual(1, summary["proposed_count"])
            self.assertEqual(1, summary["manual_review_count"])
            self.assertFalse(summary["applicable"])

    def test_hkex_issuer_continuity_never_fuzzy_merges_and_uses_signed_alias(self):
        with tempfile.TemporaryDirectory() as directory:
            historical = Path(directory) / "historical.csv"
            profiles = Path(directory) / "profiles.csv"
            drafts = Path(directory) / "drafts.csv"
            aliases = Path(directory) / "aliases.csv"
            historical.write_text(
                "stock_code,company_name,company_abbr,exchange\n"
                "00001.HK,甲能源有限公司,甲能源,HKEX\n"
                "00002.HK,乙能源有限公司,乙能源,HKEX\n"
                "#N/A,中国港能智慧能源集团有限公司,中国港能,HKEX\n",
                encoding="utf-8",
            )
            profiles.write_text(
                "stock_code,chinese_name,chinese_short_name,hsic_sub_sector,source_url\n"
                "00001.HK,甲能源有限公司 - H股,甲能源,燃氣供應,https://hkex/1\n"
                "00002.HK,全新发行人有限公司,全新发行人,半導體,https://hkex/2\n"
                "00931.HK,中國港能智慧能源集團有限公司,中國港能,常規電力,https://hkex/931\n",
                encoding="utf-8",
            )
            drafts.write_text(
                "stock_code,review_status\n00001.HK,proposed\n00002.HK,manual_review\n00931.HK,proposed\n",
                encoding="utf-8",
            )
            aliases.write_text(
                "old_code,new_code,evidence_url\n#N/A,00931.HK,https://hkexnews/signed.pdf\n",
                encoding="utf-8",
            )
            rows, summary = audit_hkex_issuer_continuity(
                historical, profiles, drafts, [aliases],
            )
            by_code = {item.stock_code: item for item in rows}
            self.assertEqual("exact_name", by_code["00001.HK"].continuity_status)
            self.assertTrue(by_code["00001.HK"].h_share_clue)
            self.assertEqual("name_difference", by_code["00002.HK"].continuity_status)
            self.assertEqual(0, by_code["00002.HK"].priority)
            self.assertEqual("#N/A", by_code["00931.HK"].historical_stock_code)
            self.assertEqual("signed_code_resolution", by_code["00931.HK"].continuity_status)
            self.assertEqual("https://hkexnews/signed.pdf", by_code["00931.HK"].resolution_evidence_url)
            self.assertEqual(0, summary["auto_merged_count"])
            self.assertFalse(summary["complete"])

    def test_apply_continuity_decisions_excludes_new_issuer_and_deduplicates_ah(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = Path(directory) / "audit.csv"
            decisions = Path(directory) / "decisions.csv"
            audit.write_text(
                "stock_code,next_action\n00600.HK,review_issuer_identity_and_industry\n"
                "00042.HK,review_ah_relationship\n",
                encoding="utf-8",
            )
            decisions.write_text(
                "decision_id,stock_code,outcome,related_a_code,entity_id,evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
                "C1,00600.HK,new_issuer,,,https://hkex/600,2026-07-29,alice,2026-07-29T10:00:00+08:00,当前发行人已经变更\n"
                "C2,00042.HK,ah_same_entity,000585.SZ,ENTITY-NEE,https://hkex/42,2026-07-29,bob,2026-07-29T10:01:00+08:00,官方年报确认A/H同主体\n",
                encoding="utf-8",
            )
            companies = [
                UniverseCompany("00600.HK", "旧候选", "HKEX", "待分类", entity_id="00600.HK"),
                UniverseCompany("00042.HK", "东北电气H", "HKEX", "能源设备", entity_id="00042.HK"),
                UniverseCompany("000585.SZ", "东北电气", "SZSE", "能源设备", entity_id="000585.SZ"),
            ]
            rows, applied, summary = apply_issuer_continuity_decisions(companies, audit, decisions)
            by_code = {item.stock_code: item for item in rows}
            self.assertFalse(by_code["00600.HK"].included)
            self.assertIn("发行人变更", by_code["00600.HK"].exclusion_reason)
            self.assertFalse(by_code["00042.HK"].included)
            self.assertEqual("ENTITY-NEE", by_code["000585.SZ"].entity_id)
            self.assertEqual(0, summary["unresolved_review_count"])
            self.assertTrue(summary["complete"])
            self.assertEqual(2, len(applied))

    def test_plan_continuity_evidence_tasks_skips_completed_and_prioritizes_identity(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = Path(directory) / "audit.csv"
            audit.write_text(
                "stock_code,historical_name,current_chinese_name,priority,next_action,evidence_url,resolution_evidence_url\n"
                "00001.HK,甲能源,甲能源,9,continuity_name_check_complete,https://hkex/1,\n"
                "00600.HK,中国基建投资,爱芯元智,0,review_issuer_identity_and_industry,https://hkex/600,\n"
                "00042.HK,东北电气,东北电气H股,3,review_ah_relationship,https://hkex/42,\n",
                encoding="utf-8",
            )
            tasks, summary = plan_continuity_evidence_tasks(audit)
            self.assertEqual(["00600.HK", "00042.HK"], [item.stock_code for item in tasks])
            self.assertIn("主营业务", tasks[0].evidence_requirements)
            self.assertEqual("pending", tasks[0].task_status)
            self.assertEqual(2, summary["task_count"])
            self.assertEqual(1, summary["complete_without_review_count"])
            self.assertFalse(summary["complete"])

    def test_plan_continuity_evidence_tasks_rejects_unknown_action(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = Path(directory) / "audit.csv"
            audit.write_text(
                "stock_code,historical_name,current_chinese_name,priority,next_action,evidence_url,resolution_evidence_url\n"
                "00001.HK,甲,乙,1,guess_by_name,https://hkex/1,\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "next_action无效"):
                plan_continuity_evidence_tasks(audit)

    def test_apply_continuity_decision_rejects_unsigned_ah_mapping(self):
        with tempfile.TemporaryDirectory() as directory:
            audit = Path(directory) / "audit.csv"
            decisions = Path(directory) / "decisions.csv"
            audit.write_text("stock_code,next_action\n00042.HK,review_ah_relationship\n", encoding="utf-8")
            decisions.write_text(
                "decision_id,stock_code,outcome,related_a_code,entity_id,evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
                "C1,00042.HK,ah_same_entity,000585.SZ,,https://hkex,2026-07-29,,2026-07-29T10:00:00,\n",
                encoding="utf-8",
            )
            with self.assertRaises(ValueError):
                apply_issuer_continuity_decisions(
                    [
                        UniverseCompany("00042.HK", "H", "HKEX", "待分类"),
                        UniverseCompany("000585.SZ", "A", "SZSE", "能源设备"),
                    ], audit, decisions,
                )

    def test_bse_jsonp_zero_based_pagination(self):
        def payload(page, code):
            return 'null([{"content":[{"xxzqdm":"%s","xxzqjc":"测试%s","xxhyzl":"电气机械和器材制造业","xxjsrq":"20260729"}],"number":%s,"totalPages":2,"totalElements":2}])' % (code, page, page)
        pages = {0: payload(0, "920001"), 1: payload(1, "920002")}
        rows, as_of_date, raw = collect_bse_listings(pages.__getitem__, expected_as_of_date="2026-07-29")
        self.assertEqual(["920001.BJ", "920002.BJ"], [item.stock_code for item in rows])
        self.assertEqual("2026-07-29", as_of_date)
        self.assertEqual(2, len(raw))
        self.assertTrue(audit_snapshot(rows).valid)

    def test_bse_rejects_page_mismatch_and_incomplete_result(self):
        wrong = 'null([{"content":[],"number":1,"totalPages":1,"totalElements":0}])'
        with self.assertRaisesRegex(ValueError, "分页响应错位"):
            parse_bse_page(wrong, 0)
        incomplete = 'null([{"content":[],"number":0,"totalPages":1,"totalElements":1}])'
        with self.assertRaisesRegex(ValueError, "分页不完整"):
            collect_bse_listings(lambda _: incomplete)

    def test_bse_official_code_mapping_parses_collisions(self):
        html = '<table><tr><td>1</td><td>许昌智能</td><td>2024/1/26</td><td>831396</td><td>920496</td></tr></table>'
        rows = parse_bse_code_mapping(html)
        self.assertEqual("831396.BJ", rows[0].old_code)
        self.assertEqual("920496.BJ", rows[0].new_code)

    def test_bse_official_disclosure_response_and_strict_titles(self):
        payload = json.dumps({"data": {"content": [{"disclosures": [
            {"companyCd": "920110", "companyName": "雷特科技", "publishDate": "2026-04-10",
             "disclosureTitle": "[定期报告]雷特科技:2025年年度报告", "disclosurePostTitle": "",
             "destFilePath": "/disclosure/2026/annual.pdf"},
            {"companyCd": "920110", "companyName": "雷特科技", "publishDate": "2026-04-10",
             "disclosureTitle": "雷特科技:2025年年度报告摘要", "disclosurePostTitle": "",
             "destFilePath": "/disclosure/2026/summary.pdf"},
            {"companyCd": "920110", "companyName": "雷特科技", "publishDate": "2026-04-11",
             "disclosureTitle": "雷特科技:2025年度社会责任报告", "disclosurePostTitle": "",
             "destFilePath": "/disclosure/2026/esg.pdf"},
        ]}], "number": 0, "totalPages": 3}}, ensure_ascii=False)
        rows, page, pages = parse_bse_disclosures(payload, 2025)
        self.assertEqual((0, 3), (page, pages))
        self.assertEqual({"annual_report", "esg_report"}, {item.document_type for item in rows})
        self.assertEqual("920110.BJ", rows[0].stock_code)
        self.assertTrue(rows[0].source_url.startswith("https://www.bse.cn/disclosure/"))
        self.assertIsNone(classify_disclosure_title("关于2025年年度报告问询函的回复", "2025"))

    def test_bse_disclosure_discovery_validates_pages_and_selects_latest(self):
        def page(number, title, path, total=2):
            return json.dumps({"data": {"content": [{"disclosures": [{
                "companyCd": "920110", "companyName": "雷特科技",
                "publishDate": f"2026-04-{10 + number:02d}", "disclosureTitle": title,
                "destFilePath": path,
            }]}], "number": number, "totalPages": total}}, ensure_ascii=False)
        payloads = {
            0: page(0, "雷特科技:2025年年度报告", "/old.pdf"),
            1: page(1, "雷特科技:2025年年度报告", "/new.pdf"),
        }
        rows = discover_bse_reports("920110.BJ", 2025, payloads.__getitem__)
        self.assertEqual(1, len(rows))
        self.assertTrue(rows[0].source_url.endswith("new.pdf"))
        with self.assertRaisesRegex(ValueError, "分页响应错位"):
            discover_bse_reports("920110.BJ", 2025, lambda _: page(1, "雷特科技:2025年年度报告", "/a.pdf"))

    def test_bse_annual_report_request_uses_official_category(self):
        payload = 'null([{"listInfo":{"content":[{"companyCd":"920110","companyName":"雷特科技","publishDate":"2026-04-10","disclosureTitle":"[定期报告]雷特科技:2025年年度报告","destFilePath":"/disclosure/2026/a.pdf"}],"number":0,"totalPages":1}}])'
        requests = []
        def fetcher(request):
            requests.append(request)
            return payload.encode()
        rows = discover_bse_annual_report("920110.BJ", 2025, fetcher)
        self.assertEqual(1, len(rows))
        query = request_data = requests[0].data.decode()
        self.assertIn("disclosureSubtype%5B%5D=9503-1001", query)
        self.assertIn("companyCd=920110", request_data)

    def test_official_listing_rejects_html_and_incomplete_pages(self):
        with self.assertRaises(ValueError):
            parse_listing_page("<html>blocked</html>")
        payloads = {
            1: '{"rows":[{"code":"1","name":"甲"}],"page":1,"pageCount":2,"total":3}',
            2: '{"rows":[{"code":"2","name":"乙"}],"page":2,"pageCount":2,"total":3}',
        }
        with self.assertRaisesRegex(ValueError, "分页不完整"):
            collect_listing_pages("SZSE", "https://official/list", "2026-07-28", payloads.__getitem__)

    def test_szse_metadata_and_html_name_are_parsed(self):
        payload = '[{"metadata":{"pageno":1,"pagecount":1,"recordcount":1},"data":[' \
                  '{"zqdm":"000027","gsjc":"<a><u>深圳能源</u></a>","sshymc":"D 电力、热力、燃气及水生产和供应业"}]}]'
        page = parse_listing_page(payload)
        self.assertEqual(1, page.total_count)
        rows = collect_listing_pages("SZSE", "https://www.szse.cn/market/stock/company/", "2026-07-28", lambda _: payload)
        self.assertEqual("深圳能源", rows[0].company_name)
        self.assertTrue(rows[0].industry.startswith("D "))

    def test_sse_pagehelp_response_is_parsed(self):
        payload = '{"pageHelp":{"pageNo":1,"pageCount":1,"total":1,"data":[' \
                  '{"A_STOCK_CODE":"688001","COMPANY_ABBR":"华兴源创","CSRC_CODE_DESC":"制造业"}]},' \
                  '"result":[{"A_STOCK_CODE":"688001","COMPANY_ABBR":"华兴源创","CSRC_CODE_DESC":"制造业"}]}'
        rows = collect_listing_pages("SSE", "https://www.sse.com.cn/assortment/stock/home/", "2026-07-28", lambda _: payload)
        self.assertEqual("688001.SH", rows[0].stock_code)
        self.assertEqual("华兴源创", rows[0].company_name)

    def test_reference_ocr_codes_are_unique_and_snapshot_matched(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "ocr.txt"
            path.write_text("600900.SH 00956.HK 600900.SH 835185.BJ", encoding="utf-8")
            snapshot = [UniverseCompany("600900.SH", "长江电力", "SSE", "电力")]
            rows = extract_reference_securities(path, snapshot, "67-72")
            self.assertEqual(["600900.SH", "00956.HK", "835185.BJ"], [item.stock_code for item in rows])
            self.assertEqual("长江电力", rows[0].company_name)
            self.assertTrue(rows[0].matched_snapshot)
            self.assertFalse(rows[1].matched_snapshot)

    def test_reference_old_bse_code_matches_through_official_alias(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "ocr.txt"
            path.write_text("835185.BJ", encoding="utf-8")
            snapshots = [UniverseCompany("920185.BJ", "贝特瑞", "BSE", "电池制造")]
            rows = extract_reference_securities(path, snapshots, "67", {"835185.BJ": "920185.BJ"})
            self.assertTrue(rows[0].matched_snapshot)
            self.assertEqual("920185.BJ", rows[0].current_stock_code)

    def test_registry_reconciliation_exact_normalized_and_unmatched(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.csv"
            path.write_text(
                "证券代码,企业名称\n600900.SH,长江电力\n,深圳能源集团股份有限公司\n,不存在能源\n",
                encoding="utf-8",
            )
            snapshots = [
                UniverseCompany("600900.SH", "长江电力", "SSE", "电力"),
                UniverseCompany("000027.SZ", "深圳能源", "SZSE", "电力"),
            ]
            rows = reconcile_registry(path, snapshots, "用户名录", "", "2026-07-28")
            self.assertEqual(["matched", "matched", "unmatched"], [item.match_status for item in rows])
            self.assertEqual("name_normalized", rows[1].match_method)
            self.assertEqual("深圳能源", normalize_company_name("深圳能源集团股份有限公司"))

    def test_registry_code_name_conflict_requires_review(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.csv"
            path.write_text("stock_code,company_name\n600900.SH,错误名称\n", encoding="utf-8")
            rows = reconcile_registry(path, [UniverseCompany("600900.SH", "长江电力", "SSE", "电力")], "x", "", "2026-07-28")
            self.assertEqual("review", rows[0].match_status)

    def test_collection_plan_prioritizes_missing_annual_report(self):
        companies = [
            UniverseCompany("A", "甲", "SSE", "电力"),
            UniverseCompany("B", "乙", "SZSE", "电力"),
            UniverseCompany("C", "丙", "HKEX", "电力", included=False),
        ]
        records = [DocumentRecord("B", "乙", 2025, "annual_report", "u1", "u1", "a.pdf", "h", 1)]
        tasks = plan_collection(companies, records, 2025)
        self.assertEqual(["A", "B"], [item.stock_code for item in tasks])
        self.assertEqual("discover_annual_and_esg", tasks[0].next_action)
        self.assertEqual("discover_esg_or_scan_annual", tasks[1].next_action)
        summary = collection_summary(tasks)
        self.assertEqual(1, summary["annual_collected"])
        self.assertEqual(0, summary["ready_for_extraction"])

    def test_import_historical_two_row_workbook(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "history.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["历史评价表"])
            sheet.append(["序号", "", "", "联系人", "城市", "证券代码", "公司简称", "公司名称", "数值类别", *[name for _, name in __import__('aegis_esg.historical', fromlist=['INDICATORS']).INDICATORS], "ESG分数"])
            sheet.append([1, "", "", "张三", "北京", "600900.SH", "长江电力", "中国长江电力股份有限公司", "指标数值", *range(1, 11), 88.5])
            sheet.append([1, "", "", "", "", "600900.SH", "长江电力", "中国长江电力股份有限公司", "指标分值", *range(11, 21), ""])
            workbook.save(path)
            companies, observations, audit = import_historical_workbook(
                path, [UniverseCompany("600900.SH", "长江电力", "SSE", "电力")], 2024, 2023,
            )
            self.assertEqual(1, len(companies))
            self.assertEqual(10, len(observations))
            self.assertEqual("matched", companies[0].current_snapshot_status)
            self.assertEqual("1", observations[0].raw_value)
            self.assertEqual(10, audit["observation_count"])

    def test_import_historical_rejects_missing_score_row(self):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest("openpyxl not installed")
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "broken.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["序号", "", "", "联系人", "城市", "证券代码", "公司简称", "公司名称", "数值类别", "温室气体排放强度", "综合能源消耗强度", "NOx", "SO2", "新鲜水资源消耗强度", "一般固废排放强度", "环保/安全生产投入占比", "研发费用占比", "现金分红", "资产负债率", "ESG分数"])
            sheet.append([1, "", "", "", "", "600900.SH", "长江电力", "长江电力", "指标数值"])
            workbook.save(path)
            with self.assertRaisesRegex(ValueError, "缺少紧随其后的指标分值行"):
                import_historical_workbook(path, [], 2024, 2023)

    def test_historical_migration_separates_candidates_and_review(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.csv"
            path.write_text(
                "stock_code,company_name,company_abbr,exchange,st_flag,historical_rank,historical_esg_score\n"
                "600900.SH,长江电力,长江电力,SSE,False,1,70\n"
                "600112.SH,天成控股,*ST天成,SSE,True,2,20\n"
                "00001.HK,港股甲,港股甲,HKEX,False,3,10\n"
                "#N/A,未知公司,未知公司,UNKNOWN,False,4,5\n",
                encoding="utf-8",
            )
            rows, audit = plan_historical_migration(
                path, [UniverseCompany("600900.SH", "长江电力", "SSE", "电力")],
            )
            self.assertEqual(
                ["provisional_include", "exclude", "pending_snapshot", "manual_review"],
                [item.decision for item in rows],
            )
            self.assertEqual(1, audit["provisional_company_count"])
            self.assertEqual(2, audit["requires_review_count"])
            universe = Path(directory) / "universe.csv"
            write_candidate_universe(universe, rows)
            imported = __import__('aegis_esg.universe', fromlist=['read_universe']).read_universe(universe)
            self.assertEqual([True, False, False, False], [item.included for item in imported])

    def test_historical_migration_uses_available_hkex_snapshot(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.csv"
            path.write_text(
                "stock_code,company_name,company_abbr,exchange,st_flag\n"
                "00001.HK,港股甲,港股甲,HKEX,False\n",
                encoding="utf-8",
            )
            rows, audit = plan_historical_migration(
                path, [UniverseCompany("00001.HK", "CKH HOLDINGS", "HKEX", "待分类")],
            )
            self.assertEqual("provisional_include", rows[0].decision)
            self.assertEqual("CKH HOLDINGS", rows[0].current_name)
            self.assertEqual(["HKEX"], audit["snapshot_exchanges"])

    def test_historical_missing_code_can_use_signed_resolution(self):
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "registry.csv"
            path.write_text(
                "stock_code,company_name,company_abbr,exchange,st_flag\n"
                "#N/A,中国港能智慧能源集团有限公司,中国港能,UNKNOWN,False\n",
                encoding="utf-8",
            )
            rows, _ = plan_historical_migration(
                path, [UniverseCompany("00931.HK", "CHINA HK POWER", "HKEX", "待分类")],
                {"#N/A": "00931.HK"},
            )
            self.assertEqual("provisional_include", rows[0].decision)
            self.assertEqual("00931.HK", rows[0].stock_code)

    def test_augment_universe_requires_snapshot_and_evidence(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "snapshot.csv"
            snapshot.write_text(
                "stock_code,company_name,exchange,industry,entity_id,source_url,as_of_date\n"
                "600925.SH,苏能股份,SSE,煤炭,600925.SH,https://official,2026-07-29\n",
                encoding="utf-8",
            )
            additions = Path(directory) / "add.csv"
            additions.write_text(
                "stock_code,evidence_url,reason\n600925.SH,evidence.txt,参考前200\n",
                encoding="utf-8",
            )
            rows = augment_candidate_universe([], additions, snapshot)
            self.assertEqual("600925.SH", rows[0].stock_code)
            self.assertTrue(rows[0].included)

    def test_bind_snapshot_provenance_fills_only_missing_fields_by_exact_code(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "snapshot.csv"
            snapshot.write_text(
                "stock_code,company_name,exchange,industry,entity_id,source_url,as_of_date\n"
                "600900.SH,长江电力,SSE,电力,600900.SH,https://official,2026-07-29\n",
                encoding="utf-8",
            )
            companies = [
                UniverseCompany("600900.SH", "中国长江电力", "SSE", "电力"),
                UniverseCompany(
                    "000001.SZ", "甲", "SZSE", "电力", source_url="https://evidence",
                ),
            ]
            rows, bindings, summary = bind_snapshot_provenance(companies, snapshot)
            self.assertEqual("https://official", rows[0].source_url)
            self.assertEqual("2026-07-29", rows[0].as_of_date)
            self.assertEqual("name_difference", bindings[0].status)
            self.assertEqual("https://evidence", rows[1].source_url)
            self.assertEqual("unmatched", bindings[1].status)
            self.assertTrue(bindings[1].included)
            self.assertEqual(1, summary["source_filled_count"])
            self.assertEqual(1, summary["included_unmatched_count"])
            self.assertFalse(summary["complete"])

    def test_bind_snapshot_provenance_flags_explicit_entity_conflict(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "snapshot.csv"
            snapshot.write_text(
                "stock_code,company_name,exchange,industry,entity_id,source_url,as_of_date\n"
                "00001.HK,甲,HKEX,电力,ENTITY-NEW,https://official,2026-07-29\n",
                encoding="utf-8",
            )
            company = UniverseCompany("00001.HK", "甲", "HKEX", "电力", entity_id="ENTITY-OLD")
            rows, bindings, summary = bind_snapshot_provenance([company], snapshot)
            self.assertEqual("ENTITY-OLD", rows[0].entity_id)
            self.assertEqual("entity_conflict", bindings[0].status)
            self.assertEqual(1, summary["entity_conflict_count"])
            self.assertFalse(summary["complete"])

    def test_universe_evidence_plan_prioritizes_hkex_and_tracks_snapshot_industry(self):
        with tempfile.TemporaryDirectory() as directory:
            snapshot = Path(directory) / "snapshot.csv"
            snapshot.write_text(
                "stock_code,company_name,exchange,industry,entity_id,source_url,as_of_date\n"
                "00001.HK,甲,HKEX,待分类,00001.HK,https://hkex,2026-07-29\n"
                "600001.SH,乙,SSE,制造业,600001.SH,https://sse,2026-07-29\n",
                encoding="utf-8",
            )
            companies = [
                UniverseCompany("600001.SH", "乙", "SSE", "历史能源样本待复核"),
                UniverseCompany("00001.HK", "甲", "HKEX", "待分类"),
            ]
            tasks, summary = plan_universe_evidence(companies, snapshot)
            self.assertEqual("00001.HK", tasks[0].stock_code)
            self.assertEqual("collect_hkex_industry_and_chinese_name_evidence", tasks[0].next_action)
            self.assertEqual("制造业", tasks[1].snapshot_industry)
            self.assertEqual(2, summary["pending_industry_count"])
            self.assertFalse(summary["publishable"])
            hk_tasks, hk_summary = plan_universe_evidence(companies, snapshot, ("HKEX",))
            self.assertEqual(["00001.HK"], [item.stock_code for item in hk_tasks])
            self.assertEqual(["HKEX"], hk_summary["exchange_filter"])

    def test_apply_signed_universe_evidence_classifies_and_deduplicates_ah(self):
        with tempfile.TemporaryDirectory() as directory:
            decisions = Path(directory) / "decisions.csv"
            decisions.write_text(
                "stock_code,decision,sub_industry,entity_id,evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
                "600001.SH,include,电力,ENTITY-A,https://official/a,2026-07-29,alice,2026-07-29T10:00:00+08:00,发行人年报确认\n"
                "00001.HK,include,电力,ENTITY-A,https://official/h,2026-07-29,bob,2026-07-29T10:01:00+08:00,港交所披露确认\n",
                encoding="utf-8",
            )
            companies = [
                UniverseCompany("00001.HK", "甲H", "HKEX", "待分类", entity_id="00001.HK", source_url="https://hkex", as_of_date="2026-07-29"),
                UniverseCompany("600001.SH", "甲", "SSE", "待分类", entity_id="600001.SH", source_url="https://sse", as_of_date="2026-07-29"),
            ]
            rows, audit_rows, summary = apply_universe_evidence(companies, decisions)
            by_code = {item.stock_code: item for item in rows}
            self.assertTrue(by_code["600001.SH"].included)
            self.assertFalse(by_code["00001.HK"].included)
            self.assertIn("保留600001.SH", by_code["00001.HK"].exclusion_reason)
            self.assertEqual("https://hkex", by_code["00001.HK"].source_url)
            self.assertEqual(1, summary["ah_duplicate_excluded_count"])
            self.assertEqual("ah_duplicate_exclude", next(item for item in audit_rows if item.stock_code == "00001.HK").action)

    def test_apply_universe_evidence_rejects_unsigned_or_placeholder_decision(self):
        with tempfile.TemporaryDirectory() as directory:
            decisions = Path(directory) / "decisions.csv"
            decisions.write_text(
                "stock_code,decision,sub_industry,entity_id,evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
                "600001.SH,include,仍待复核,600001.SH,https://official,2026-07-29,,2026-07-29T10:00:00,\n",
                encoding="utf-8",
            )
            with self.assertRaises(ValueError):
                apply_universe_evidence(
                    [UniverseCompany("600001.SH", "甲", "SSE", "待分类")], decisions,
                )

    def test_merge_universe_evidence_batches_tracks_supersede_and_revoke(self):
        header = (
            "decision_id,batch_id,operation,supersedes,stock_code,decision,sub_industry,entity_id,"
            "evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
        )
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "first.csv"
            second = Path(directory) / "second.csv"
            third = Path(directory) / "third.csv"
            first.write_text(
                header + "D1,B1,upsert,,600001.SH,include,电力,A,https://official/1,2026-07-29,alice,2026-07-29T10:00:00+08:00,初审\n",
                encoding="utf-8",
            )
            second.write_text(
                header + "D2,B2,upsert,D1,600001.SH,include,新能源,A,https://official/2,2026-07-30,bob,2026-07-30T10:00:00+08:00,更正行业\n",
                encoding="utf-8",
            )
            third.write_text(
                header + "D3,B3,revoke,D2,600001.SH,,,,https://official/3,2026-07-31,carol,2026-07-31T10:00:00+08:00,撤销待重审\n",
                encoding="utf-8",
            )
            active, ledger, summary = merge_universe_evidence_batches([first, second])
            self.assertEqual("新能源", active[0]["sub_industry"])
            self.assertEqual(["superseded", "active"], [item["ledger_state"] for item in ledger])
            self.assertEqual(1, summary["active_decision_count"])
            active, ledger, summary = merge_universe_evidence_batches([first, second, third])
            self.assertFalse(active)
            self.assertEqual(0, summary["active_decision_count"])
            self.assertEqual("revoked", ledger[1]["ledger_state"])

    def test_merge_universe_evidence_rejects_version_fork(self):
        header = (
            "decision_id,batch_id,operation,supersedes,stock_code,decision,sub_industry,entity_id,"
            "evidence_url,evidence_date,reviewer,reviewed_at,rationale\n"
        )
        with tempfile.TemporaryDirectory() as directory:
            first = Path(directory) / "first.csv"
            fork = Path(directory) / "fork.csv"
            first.write_text(
                header + "D1,B1,upsert,,600001.SH,include,电力,A,https://official/1,2026-07-29,alice,2026-07-29T10:00:00+08:00,初审\n",
                encoding="utf-8",
            )
            fork.write_text(
                header + "D2,B2,upsert,WRONG,600001.SH,include,新能源,A,https://official/2,2026-07-30,bob,2026-07-30T10:00:00+08:00,冲突版本\n",
                encoding="utf-8",
            )
            with self.assertRaises(ValueError):
                merge_universe_evidence_batches([first, fork])

    def test_indicator_plan_builds_complete_company_indicator_matrix(self):
        companies = [
            UniverseCompany("A", "甲", "SSE", "电力"),
            UniverseCompany("B", "乙", "SZSE", "电力"),
        ]
        first = self.methodology.indicators[0]
        observations = [Observation("A", "甲", 2025, first.code, 1)]
        tasks, summary = plan_indicator_tasks(companies, observations, self.methodology, 2025)
        self.assertEqual(160, len(tasks))
        self.assertEqual(1, summary["confirmed_count"])
        self.assertEqual(1, summary["empty_companies"])
        self.assertFalse(summary["publishable"])
        confirmed = next(item for item in tasks if item.company_code == "A" and item.indicator_code == first.code)
        self.assertEqual("complete", confirmed.next_action)

    def test_candidate_coverage_builds_quantitative_company_matrix(self):
        with tempfile.TemporaryDirectory() as directory:
            companies = Path(directory) / "companies.csv"
            companies.write_text(
                "stock_code,chinese_name,included\nA,甲,True\nB,乙,True\nC,丙,False\n",
                encoding="utf-8",
            )
            indicator = self.methodology.quantitative[0]
            observations = [
                Observation("A", "甲", 2025, indicator.code, 1, ValueStatus.PENDING,
                            source_page=8, confidence=.91),
            ]
            tasks, summary = plan_candidate_coverage(companies, observations, self.methodology)
        self.assertEqual(2 * len(self.methodology.quantitative), len(tasks))
        self.assertEqual(37, summary["quantitative_indicator_count"])
        self.assertEqual(1, summary["candidate_task_count"])
        self.assertEqual(1, summary["candidate_observation_count"])
        self.assertEqual(73, summary["missing_task_count"])
        self.assertEqual(36, summary["zero_coverage_indicator_count"])
        self.assertNotIn(indicator.code, summary["zero_coverage_indicator_codes"])
        self.assertEqual(20, summary["minimum_population_threshold"])
        self.assertEqual(37, summary["below_minimum_population_indicator_count"])
        self.assertFalse(summary["minimum_population_gate_passed"])
        self.assertFalse(summary["complete"])
        self.assertFalse(summary["applicable"])
        available = next(
            item for item in tasks if item.company_code == "A" and item.indicator_code == indicator.code
        )
        missing = next(
            item for item in tasks if item.company_code == "B" and item.indicator_code == indicator.code
        )
        self.assertEqual(("candidate_available", "review_candidates", "8"),
                         (available.status, available.next_action, available.source_pages))
        self.assertEqual(("missing_candidate", "extend_extraction_rules"),
                         (missing.status, missing.next_action))

    def test_progress_dashboard_loads_coverage_and_conflict_evidence(self):
        data = load_progress_dashboard(
            ROOT / "output/audit/hkex_quantitative_candidate_tasks_summary_2026-07-29.json",
            ROOT / "output/audit/hkex_quantitative_candidate_tasks_2026-07-29.csv",
            ROOT / "data/review/hkex_indicator_candidates_review_2026-07-29.csv",
            ROOT / "data/review/hkex_indicator_candidates_2026-07-29.csv",
            self.methodology,
            ROOT / "output/audit/hkex_candidate_review_tiers_summary_2026-07-29.json",
            ROOT / "output/audit/hkex_candidate_review_tiers_2026-07-29.csv",
            ROOT / "output/audit/hkex_resolution_preview_freeze_audit_2026-07-29.json",
        )
        self.assertEqual(3404, data["overview"]["task_count"])
        self.assertEqual(422, data["overview"]["candidate_task_count"])
        self.assertEqual(3, data["overview"]["conflict_count"])
        self.assertEqual(
            {"00196.HK", "00600.HK", "01205.HK"},
            {item["company_code"] for item in data["conflicts"]},
        )
        self.assertTrue(all(item["candidates"] for item in data["conflicts"]))
        self.assertEqual(415, data["review_tiers"]["summary"]["tier_counts"]["auto_policy_eligible"])
        self.assertEqual(7, len(data["review_tiers"]["manual_items"]))
        self.assertTrue(data["resolution_freeze_audit"]["valid"])
        self.assertFalse(data["resolution_freeze_audit"]["freeze_ready"])
        rendered = render_progress_dashboard(data)
        self.assertIn("AegisESP 开发进度", rendered)
        self.assertIn("候选数据不等于正式评分", rendered)
        self.assertIn("下一批关键缺口", rendered)
        self.assertIn("indicator-search", rendered)
        self.assertIn("审核分层", rendered)
        self.assertIn("v4可自动确认", rendered)
        self.assertIn("待人工审核", rendered)
        review_template = render_conflict_review_template(data)
        self.assertTrue(review_template.startswith("\ufeffcompany_code,"))
        self.assertIn("00196.HK", review_template)
        self.assertNotIn(",confirm,", review_template)

    def test_system_demo_page_is_distinct_from_progress_dashboard(self):
        data = {
            "overview": {"company_count": 2, "candidate_observation_count": 4,
                          "candidate_task_count": 3, "conflict_count": 1},
            "review_tiers": {"summary": {"tier_counts": {"auto_policy_eligible": 2}}},
            "resolution_freeze_audit": {"freeze_ready": False},
        }
        from aegis_esg.dashboard import render_system_demo
        rendered = render_system_demo(data)
        self.assertIn("系统演示总览", rendered)
        self.assertIn("/demo/ranking", rendered)
        self.assertIn("/demo/review-workbench", rendered)
        self.assertNotIn("AegisESP 开发进度</h1>", rendered)

    def test_download_validation_decompresses_and_rejects_html(self):
        pdf = b"%PDF-1.7\n" + b"x" * 10_000
        self.assertEqual(pdf, _decode_document(gzip.compress(pdf), "gzip", "https://official/a.pdf"))
        with self.assertRaises(ValueError):
            _decode_document(b"<html>blocked</html>", "", "https://official/a.pdf")

    def test_sse_download_has_official_big5_fallback(self):
        url = "https://www.sse.com.cn/disclosure/listedinfo/announcement/a.pdf"
        self.assertEqual(
            [
                url,
                "https://big5.sse.com.cn/site/cht/www.sse.com.cn/disclosure/listedinfo/announcement/a.pdf",
            ],
            _download_candidates(url),
        )
        self.assertEqual(
            ["https://example.test/a.pdf"],
            _download_candidates("https://example.test/a.pdf"),
        )

    def test_document_index_round_trip(self):
        record = DocumentRecord("A", "甲", 2025, "annual_report", "https://source", "https://retrieval", "a.pdf", "abc", 12)
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "index.csv"
            write_document_index(path, [record])
            self.assertEqual(record, _read_document_index(path)[record.source_url])

    def test_resumable_collection_never_rebinds_existing_path_to_new_url(self):
        old_body = b"%PDF-1.7\n" + b"o" * 10_000
        new_body = b"%PDF-1.7\n" + b"n" * 10_000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "raw/A/2025/annual_report.pdf"
            target.parent.mkdir(parents=True)
            target.write_bytes(old_body)
            manifest = root / "manifest.csv"
            manifest.write_text(
                "company_code,company_name,report_year,document_type,source_url\n"
                "A,甲,2025,annual_report,https://official/new.pdf\n",
                encoding="utf-8",
            )
            index, failures = root / "index.csv", root / "failures.csv"
            write_document_index(index, [DocumentRecord(
                "A", "甲", 2025, "annual_report", "https://official/old.pdf",
                "https://official/old.pdf", str(target), "unused", len(old_body),
            )])
            with patch("aegis_esg.collector._download_pdf", return_value=(new_body, "https://official/new.pdf")) as download:
                rows, errors = collect_batch(manifest, root / "raw", index, failures, 0, True)
            self.assertFalse(errors)
            self.assertEqual(new_body, target.read_bytes())
            self.assertEqual("https://official/new.pdf", rows[0].source_url)
            download.assert_called_once_with("https://official/new.pdf")

    def test_incremental_collection_preserves_main_index(self):
        body = b"%PDF-1.7\n" + b"x" * 10_000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            existing = DocumentRecord("A", "甲", 2025, "annual_report", "https://old", "https://old", str(root / "old.pdf"), "abc", 12)
            index, failures = root / "index.csv", root / "failures.csv"
            write_document_index(index, [existing])
            manifest = root / "retry.csv"
            manifest.write_text(
                "company_code,company_name,report_year,document_type,source_url,error\n"
                "B,乙,2025,esg_report,https://new,timeout\n", encoding="utf-8",
            )
            with patch("aegis_esg.collector._download_pdf", return_value=(body, "https://new")):
                rows, errors = collect_batch(manifest, root / "raw", index, failures, 0, True, 1, None, True)
            self.assertFalse(errors)
            self.assertEqual({"https://old", "https://new"}, {item.source_url for item in rows})

    def test_soft_time_budget_stops_after_first_download_and_flushes_index(self):
        body = b"%PDF-1.7\n" + b"y" * 10_000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            index, failures = root / "index.csv", root / "failures.csv"
            manifest = root / "budget.csv"
            manifest.write_text(
                "company_code,company_name,report_year,document_type,source_url\n"
                "A,甲,2025,esg_report,https://a\n"
                "B,乙,2025,esg_report,https://b\n"
                "C,丙,2025,esg_report,https://c\n",
                encoding="utf-8",
            )
            calls: list[str] = []

            def fake_download(url):
                calls.append(url)
                return (body, url)

            # started=0.0; first launch budget check sees 0.0; after the first
            # completion the next budget check sees 9999.0 and stops launching.
            with patch("aegis_esg.collector._clock", side_effect=[0.0, 0.0, 9999.0, 9999.0, 9999.0]), \
                 patch("aegis_esg.collector._download_pdf", side_effect=fake_download):
                rows, errors = collect_batch(
                    manifest, root / "raw", index, failures,
                    delay_seconds=0, reuse_existing=False, workers=1,
                    max_minutes=1,
                )
            self.assertGreaterEqual(len(calls), 1)
            self.assertLess(len(calls), 3)  # budget stopped before all three ran
            self.assertEqual(1, len(rows))  # only the first was checkpointed
            self.assertEqual("https://a", rows[0].source_url)
            self.assertFalse(errors)
            indexed = _read_document_index(index)
            self.assertEqual({"https://a"}, set(indexed))

    def test_collect_batch_prefers_esg_and_dedupes_identity(self):
        body = b"%PDF-1.7\n" + b"z" * 10_000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            index, failures = root / "index.csv", root / "failures.csv"
            # Seed a duplicate/invalid identity that should be compacted away after upsert.
            write_document_index(index, [
                DocumentRecord("A", "甲", 0, "annual_report", "https://bad", "https://bad", str(root / "bad.pdf"), "x", 1),
                DocumentRecord("A", "甲", 2025, "annual_report", "https://old-annual", "https://old-annual", str(root / "old.pdf"), "y", 2),
            ])
            manifest = root / "manifest.csv"
            manifest.write_text(
                "company_code,company_name,report_year,document_type,source_url\n"
                "A,甲,2025,annual_report,https://annual\n"
                "B,乙,2025,esg_report,https://esg\n",
                encoding="utf-8",
            )
            order: list[str] = []

            def fake_download(url):
                order.append(url)
                return (body, url)

            with patch("aegis_esg.collector._download_pdf", side_effect=fake_download):
                rows, errors = collect_batch(
                    manifest, root / "raw", index, failures,
                    delay_seconds=0, reuse_existing=True, workers=1,
                    preserve_index=True, document_priority="esg",
                )
            self.assertFalse(errors)
            self.assertEqual(["https://esg", "https://annual"], order)
            self.assertEqual(2, len(dedupe_document_records(rows)))
            identities = {(item.company_code, item.report_year, item.document_type) for item in rows}
            self.assertIn(("B", 2025, "esg_report"), identities)
            self.assertIn(("A", 2025, "annual_report"), identities)
            self.assertNotIn(("A", 0, "annual_report"), identities)

    def test_supersede_documents_archives_file_and_writes_ledger(self):
        body = b"%PDF-1.4\n" + b"z" * 5000
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "raw/A/2025/annual_report.pdf"
            target.parent.mkdir(parents=True)
            target.write_bytes(body)
            digest = hashlib.sha256(body).hexdigest()
            record = DocumentRecord(
                "A", "甲", 2025, "annual_report", "https://official/briefing.pdf",
                "https://official/briefing.pdf", str(target), digest, len(body),
            )
            keep = DocumentRecord("B", "乙", 2025, "annual_report", "https://two", "https://two", "b.pdf", "def", 13)
            index = root / "index.csv"
            write_document_index(index, [record, keep])
            requests = root / "requests.csv"
            requests.write_text(
                "company_code,report_year,document_type,reason\n"
                "A,2025,annual_report,业绩说明会公告误登记为年报\n",
                encoding="utf-8",
            )
            ledger, summary_path = root / "ledger.csv", root / "summary.json"
            records, rows, summary = supersede_documents(
                index, requests, root / "archive", ledger, summary_path,
            )
            self.assertEqual([keep], records)
            self.assertEqual(1, summary["superseded_count"])
            self.assertFalse(target.exists())
            archived = root / "archive/A/2025" / f"annual_report_{digest[:8]}.pdf"
            self.assertEqual(body, archived.read_bytes())
            self.assertEqual(digest, rows[0]["sha256"])
            self.assertEqual("业绩说明会公告误登记为年报", rows[0]["reason"])
            with self.assertRaisesRegex(ValueError, "匹配0条"):
                supersede_documents(index, requests, root / "archive", ledger, summary_path)

    def test_supersede_documents_rejects_hash_mismatch(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            target = root / "a.pdf"
            target.write_bytes(b"%PDF-1.4\n" + b"z" * 100)
            record = DocumentRecord(
                "A", "甲", 2025, "annual_report", "https://one", "https://one",
                str(target), "0" * 64, 500,
            )
            index = root / "index.csv"
            write_document_index(index, [record])
            requests = root / "requests.csv"
            requests.write_text(
                "company_code,report_year,document_type,reason\nA,2025,annual_report,测试\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "Hash不一致"):
                supersede_documents(
                    index, requests, root / "archive", root / "ledger.csv", root / "summary.json",
                )

    def test_merge_document_indexes_deduplicates_exact_and_rejects_path_conflicts(self):
        first = DocumentRecord("A", "甲", 2025, "annual_report", "https://one", "https://one", "a.pdf", "abc", 12)
        second = DocumentRecord("B", "乙", 2025, "annual_report", "https://two", "https://two", "b.pdf", "def", 13)
        conflict = DocumentRecord("C", "丙", 2025, "annual_report", "https://three", "https://three", "a.pdf", "ghi", 14)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            one, two, bad = root / "one.csv", root / "two.csv", root / "bad.csv"
            write_document_index(one, [first])
            write_document_index(two, [first, second])
            write_document_index(bad, [conflict])
            rows, summary = merge_document_indexes([one, two])
            self.assertEqual([first, second], rows)
            self.assertEqual(1, summary["duplicate_count"])
            self.assertEqual(2, summary["company_count"])
            with self.assertRaisesRegex(ValueError, "本地路径冲突"):
                merge_document_indexes([one, bad])

    def test_merge_document_indexes_allows_hash_identical_metadata_correction(self):
        old = DocumentRecord("A", "甲", 0, "annual_report", "https://same", "https://same", "data/raw/A/0/annual_report.pdf", "abc", 12)
        corrected = DocumentRecord("A", "甲", 2025, "annual_report", "https://same", "https://same", "data/raw/A/2025/annual_report.pdf", "abc", 12)
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory); first = root / "first.csv"; second = root / "second.csv"
            write_document_index(first, [old]); write_document_index(second, [corrected])
            with self.assertRaisesRegex(ValueError, "URL元数据冲突"):
                merge_document_indexes([first, second])
            rows, summary = merge_document_indexes([first, second], allow_metadata_corrections=True)
            self.assertEqual([corrected], rows)
            self.assertEqual(1, summary["metadata_correction_count"])

    def test_document_coverage_distinguishes_missing_esg_and_annual(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            companies, index = root / "companies.csv", root / "index.csv"
            companies.write_text(
                "stock_code,chinese_name,included\nA,甲,true\nB,乙,yes\nC,丙,1\nD,丁,false\n", encoding="utf-8",
            )
            write_document_index(index, [
                DocumentRecord("A", "甲", 2025, "annual_report", "https://a/annual", "https://a/annual", "a.pdf", "a", 1),
                DocumentRecord("A", "甲", 2025, "esg_report", "https://a/esg", "https://a/esg", "e.pdf", "e", 1),
                DocumentRecord("B", "乙", 2025, "annual_report", "https://b/annual", "https://b/annual", "b.pdf", "b", 1),
                DocumentRecord("C", "丙", 2024, "annual_report", "https://c/old", "https://c/old", "c.pdf", "c", 1),
            ])
            rows, summary = audit_document_coverage(companies, index, 2025)
            by_code = {row.stock_code: row for row in rows}
            self.assertEqual("ready_for_extraction", by_code["A"].next_action)
            self.assertEqual("scan_annual_for_esg", by_code["B"].next_action)
            self.assertEqual("discover_annual_report", by_code["C"].next_action)
            self.assertEqual(["C"], summary["missing_annual_codes"])
            self.assertEqual(3, summary["company_count"])
            self.assertEqual(2025, summary["report_year"])
            self.assertNotIn("D", by_code)
            self.assertFalse(summary["complete"])

    def test_annual_esg_scan_preserves_page_and_pending_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            coverage, index = root / "coverage.csv", root / "index.csv"
            coverage.write_text(
                "stock_code,company_name,next_action\nA,甲,scan_annual_for_esg\nB,乙,ready_for_extraction\n",
                encoding="utf-8",
            )
            write_document_index(index, [DocumentRecord(
                "A", "甲", 2025, "annual_report", "https://official/a.pdf",
                "https://official/a.pdf", "data/raw/A/2025/annual_report.pdf", "abc", 12,
            )])
            text = root / "text/A/2025/annual_report.txt"
            text.parent.mkdir(parents=True)
            text.write_text(
                "\n=== PAGE 1 ===\nFinancial statements.\n"
                "\n=== PAGE 12 ===\nEnvironmental, Social and Governance information is set out below.\n",
                encoding="utf-8",
            )
            rows, summary = scan_annual_esg_disclosure(coverage, index, root / "text")
            self.assertEqual(1, len(rows))
            self.assertEqual(12, rows[0].source_page)
            self.assertEqual("pending", rows[0].review_status)
            self.assertEqual(1, summary["candidate_company_count"])
            self.assertFalse(summary["applicable"])

    def test_collect_annual_qualitative_evidence_maps_terms_without_scoring(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            coverage, index = root / "coverage.csv", root / "index.csv"
            coverage.write_text(
                "stock_code,company_name,annual_status\nA,甲,collected\nB,乙,missing\n",
                encoding="utf-8",
            )
            write_document_index(index, [DocumentRecord(
                "A", "甲", 2025, "annual_report", "https://official/a.pdf",
                "https://official/a.pdf", "data/raw/A/2025/annual_report.pdf", "abc", 12,
            )])
            text = root / "text/A/2025/annual_report.txt"
            text.parent.mkdir(parents=True)
            text.write_text(
                "\n=== PAGE 9 ===\n公司完善环境管理体系\x00，并加强职业健康管理。\n",
                encoding="utf-8",
            )
            rows, summary = collect_annual_qualitative_evidence(
                coverage, index, root / "text", self.methodology, 2025,
            )
            self.assertEqual({"X_E_ENV_SYSTEM", "X_S_OCCUPATIONAL_HEALTH"}, {row.indicator_code for row in rows})
            self.assertTrue(all(row.review_status == "pending" for row in rows))
            self.assertFalse(summary["scoring_authorized"])
            self.assertEqual(1, summary["annual_document_count"])
            self.assertTrue(all("\x00" not in row.evidence_text for row in rows))

    def test_qualitative_review_plan_suggests_but_never_confirms(self):
        with tempfile.TemporaryDirectory() as directory:
            coverage = Path(directory) / "coverage.csv"
            coverage.write_text(
                "stock_code,company_name,annual_status\nA,甲,collected\nB,乙,missing\n",
                encoding="utf-8",
            )
            candidates = [QualitativeEvidenceCandidate(
                "A", "甲", 2025, "X_E_ENV_SYSTEM", "环保体系", "https://a", "a.pdf", 9,
                "环境管理体系", "公司建立环境管理体系，制定年度目标，实施培训并实现减排目标。", .75,
            )]
            packets, gaps, summary = plan_qualitative_review(
                candidates, coverage, self.methodology, 2025,
            )
            self.assertEqual(1, len(packets))
            self.assertEqual(80, packets[0].suggested_score)
            self.assertEqual("pending", packets[0].review_status)
            self.assertFalse(packets[0].scoring_authorized)
            self.assertEqual(42, len(gaps))
            self.assertEqual(0, summary["auto_confirmed_count"])
            self.assertFalse(summary["scoring_authorized"])

            template = Path(directory) / "review.csv"
            self.assertEqual(1, write_qualitative_review_template(template, packets, priority=1, limit=10))
            with template.open(encoding="utf-8-sig") as stream:
                row = next(csv.DictReader(stream))
            self.assertFalse(row["action"] or row["selected_score"] or row["reviewer"])

            decision = QualitativeReviewDecision(
                "A", 2025, "X_E_ENV_SYSTEM", "confirm", "80", "alice",
                "2026-07-30T16:30:00+08:00", "核对制度、目标、行动及年度成效",
            )
            confirmed, unresolved, audits = apply_qualitative_review_decisions(packets, [decision])
            self.assertEqual(80, confirmed[0].value)
            self.assertEqual("confirmed", confirmed[0].status.value)
            self.assertEqual("https://a", confirmed[0].source_url)
            self.assertEqual("a.pdf", confirmed[0].source_file)
            self.assertFalse(unresolved)
            self.assertEqual("alice", audits[0].reviewer)

    def test_qualitative_review_decision_requires_timezone_and_leading_evidence_for_100(self):
        with tempfile.TemporaryDirectory() as directory:
            decisions = Path(directory) / "decisions.csv"
            decisions.write_text(
                "company_code,report_year,indicator_code,action,selected_score,reviewer,reviewed_at,note\n"
                "A,2025,X_E_ENV_SYSTEM,confirm,100,alice,2026-07-30T16:30:00+08:00,材料完整\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "领先或标杆"):
                read_qualitative_review_decisions(decisions)

    def _planned_packets(self, directory):
        root = Path(directory)
        coverage = root / "coverage.csv"
        coverage.write_text(
            "stock_code,company_name,annual_status,esg_status\n"
            "A,甲,collected,collected\nB,乙,missing,missing\n",
            encoding="utf-8",
        )
        candidates = [
            QualitativeEvidenceCandidate(
                "A", "甲", 2025, "X_E_ENV_SYSTEM", "环保体系", "https://a", "a.pdf", 9,
                "环境管理体系", "公司建立环境管理体系，制定年度目标，实施培训并实现减排目标。", .8,
            ),
            QualitativeEvidenceCandidate(
                "A", "甲", 2025, "X_S_OCCUPATIONAL_HEALTH", "职业健康", "https://a", "a.pdf", 10,
                "职业健康", "公司关注职业健康。", .6,
            ),
        ]
        packets, gaps, summary = plan_qualitative_review(candidates, coverage, self.methodology, 2025)
        packet_path = root / "packets.csv"
        write_qualitative_review_plan(packet_path, root / "gaps.csv", root / "summary.json", packets, gaps, summary)
        return read_qualitative_review_packets(packet_path), packet_path

    def test_qualitative_review_batch_ledger_gates(self):
        with tempfile.TemporaryDirectory() as directory:
            packets, packet_path = self._planned_packets(directory)
            root = Path(directory)
            ledger_path = root / "ledger.csv"
            batch_path = root / "batch01.csv"
            batch = create_review_batch(
                packets, batch_path, ledger_path, packet_path, label="首批", priority=1, limit=1,
            )
            self.assertEqual(1, batch.group_count)
            self.assertEqual("open", batch.status)
            self.assertTrue(batch.batch_id.startswith("QRB-"))
            self.assertEqual(1, len(read_batch_ledger(ledger_path)))
            with self.assertRaisesRegex(ValueError, "重复分配"):
                create_review_batch(packets, root / "batch02.csv", ledger_path, packet_path, priority=1, limit=1)

            batch_id, rows = read_batch_rows(batch_path)
            self.assertEqual(batch.batch_id, batch_id)
            tampered = [dict(row) for row in rows]
            tampered[0]["indicator_code"] = "X_E_EMERGENCY"
            with self.assertRaisesRegex(ValueError, "哈希不一致"):
                apply_review_batch(packets, read_batch_ledger(ledger_path), batch_id, tampered, [])

            ledger = read_batch_ledger(ledger_path)
            rows[0].update({
                "action": "confirm", "selected_score": rows[0]["suggested_score"],
                "reviewer": "alice", "reviewed_at": "2026-07-31T10:00:00+08:00", "note": "核对证据充分",
            })
            confirmed, unresolved, audits, updated = apply_review_batch(packets, ledger, batch_id, rows, [])
            self.assertEqual(1, len(confirmed))
            self.assertFalse(unresolved)
            self.assertEqual("closed", updated.status)
            self.assertEqual(1.0, updated.completion_rate)
            with self.assertRaisesRegex(ValueError, "已关闭"):
                apply_review_batch(packets, [updated], batch_id, rows, audits)

    def test_qualitative_review_batch_partial_progress_no_overwrite(self):
        with tempfile.TemporaryDirectory() as directory:
            packets, packet_path = self._planned_packets(directory)
            root = Path(directory)
            ledger_path = root / "ledger.csv"
            batch = create_review_batch(packets, root / "batch.csv", ledger_path, packet_path, priority=1)
            self.assertEqual(2, batch.group_count)
            batch_id, rows = read_batch_rows(root / "batch.csv")
            first, second = rows
            first.update({
                "action": "confirm", "selected_score": first["suggested_score"],
                "reviewer": "alice", "reviewed_at": "2026-07-31T10:00:00+08:00", "note": "核对通过",
            })
            ledger = read_batch_ledger(ledger_path)
            confirmed, unresolved, audits, updated = apply_review_batch(packets, ledger, batch_id, [first, second], [])
            self.assertEqual(1, len(confirmed))
            self.assertEqual(1, len(unresolved))
            self.assertEqual("open", updated.status)
            self.assertEqual(0.5, updated.completion_rate)
            progress = audits
            with self.assertRaisesRegex(ValueError, "禁止覆盖"):
                apply_review_batch(packets, ledger, batch_id, [first, second], progress)
            second.update({
                "action": "reject", "selected_score": "",
                "reviewer": "bob", "reviewed_at": "2026-07-31T11:00:00+08:00", "note": "证据不足",
            })
            blank_first = dict(first, action="", selected_score="", reviewer="", reviewed_at="", note="")
            confirmed, unresolved, audits2, updated = apply_review_batch(
                packets, ledger, batch_id, [blank_first, second], progress,
            )
            self.assertFalse(confirmed)
            self.assertFalse(unresolved)
            self.assertEqual("closed", updated.status)
            self.assertEqual(1.0, updated.completion_rate)

    def test_dual_review_closes_agreement_and_routes_disagreement_to_arbitration(self):
        with tempfile.TemporaryDirectory() as directory:
            packets, _ = self._planned_packets(directory)
            audits = [
                QualitativeReviewAudit(
                    "A", "甲", 2025, "X_E_ENV_SYSTEM", 80, "confirm", "80", 9,
                    "alice", "2026-07-31T10:00:00+08:00", "核对制度目标行动成效",
                ),
                QualitativeReviewAudit(
                    "A", "甲", 2025, "X_S_OCCUPATIONAL_HEALTH", 20, "confirm", "50", 10,
                    "alice", "2026-07-31T10:05:00+08:00", "补充行动证据上调",
                ),
            ]
            cases = select_dual_review_cases(packets, audits)
            self.assertEqual(2, len(cases))
            self.assertFalse(requires_dual_review(QualitativeReviewAudit(
                "A", "甲", 2025, "X_S_OCCUPATIONAL_HEALTH", 20, "confirm", "20", 10,
                "alice", "2026-07-31T10:05:00+08:00", "与建议一致",
            )))
            template = Path(directory) / "dual.csv"
            self.assertEqual(2, write_dual_review_template(template, cases))
            with template.open(encoding="utf-8-sig", newline="") as stream:
                rows = list(csv.DictReader(stream))
            for row in rows:
                self.assertFalse(row["second_action"] or row["second_reviewer"])
                if row["indicator_code"] == "X_E_ENV_SYSTEM":
                    row.update({
                        "second_action": "confirm", "second_score": "80", "second_reviewer": "bob",
                        "second_reviewed_at": "2026-07-31T12:00:00+08:00", "second_note": "复核一致",
                    })
                else:
                    row.update({
                        "second_action": "confirm", "second_score": "20", "second_reviewer": "bob",
                        "second_reviewed_at": "2026-07-31T12:05:00+08:00", "second_note": "证据仅支持20档",
                    })
            with template.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=list(rows[0].keys()), lineterminator="\n")
                writer.writeheader()
                writer.writerows(rows)
            decisions = read_dual_review_decisions(template)
            confirmed, outcomes, arbitrations, open_cases = apply_dual_review_decisions(cases, decisions)
            self.assertEqual(1, len(confirmed))
            self.assertEqual(80, confirmed[0].value)
            self.assertEqual("X_E_ENV_SYSTEM", confirmed[0].indicator_code)
            self.assertEqual({"closed_agreement", "arbitration_required"}, {item.outcome for item in outcomes})
            self.assertEqual(1, len(arbitrations))
            self.assertFalse(open_cases)

            case = arbitrations[0]
            same_side = ArbitrationDecision(
                case.company_code, case.report_year, case.indicator_code, "confirm", "20",
                "alice", "2026-07-31T13:00:00+08:00", "维持20档",
            )
            with self.assertRaisesRegex(ValueError, "区别于两名审核人"):
                apply_arbitration_decisions([case], [same_side])
            final = ArbitrationDecision(
                case.company_code, case.report_year, case.indicator_code, "confirm", "20",
                "carol", "2026-07-31T14:00:00+08:00", "仲裁采纳第二审核人20档",
            )
            resolved, unresolved, arb_audits = apply_arbitration_decisions([case], [final])
            self.assertEqual(20, resolved[0].value)
            self.assertFalse(unresolved)
            self.assertEqual("carol", arb_audits[0].arbiter)

    def test_dual_review_rejects_same_reviewer_and_unsigned_rows(self):
        with tempfile.TemporaryDirectory() as directory:
            packets, _ = self._planned_packets(directory)
            audits = [QualitativeReviewAudit(
                "A", "甲", 2025, "X_E_ENV_SYSTEM", 80, "confirm", "80", 9,
                "alice", "2026-07-31T10:00:00+08:00", "核对通过",
            )]
            cases = select_dual_review_cases(packets, audits)
            same = DualReviewDecision(
                "A", 2025, "X_E_ENV_SYSTEM", "confirm", "80",
                "alice", "2026-07-31T12:00:00+08:00", "复核一致",
            )
            with self.assertRaisesRegex(ValueError, "必须与第一审核人不同"):
                apply_dual_review_decisions(cases, [same])
            template = Path(directory) / "dual.csv"
            write_dual_review_template(template, cases)
            with template.open(encoding="utf-8-sig", newline="") as stream:
                row = next(csv.DictReader(stream))
            row.update({
                "second_action": "confirm", "second_score": "80", "second_reviewer": "bob",
                "second_reviewed_at": "2026-07-31 12:00:00", "second_note": "复核一致",
            })
            with template.open("w", encoding="utf-8-sig", newline="") as stream:
                writer = csv.DictWriter(stream, fieldnames=list(row.keys()), lineterminator="\n")
                writer.writeheader()
                writer.writerow(row)
            with self.assertRaisesRegex(ValueError, "时区"):
                read_dual_review_decisions(template)

    def test_reprioritize_evidence_gaps_orders_by_weight_and_esg_status(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            gaps = root / "gaps.csv"
            gaps.write_text(
                "company_code,company_name,report_year,indicator_code,indicator_name,indicator_weight,priority,status,next_action\n"
                "A,甲,2025,X_E_EMERGENCY,应急,1.0,2,evidence_missing,locate_additional_public_evidence\n"
                "A,甲,2025,X_E_ENV_SYSTEM,体系,4.0,1,evidence_missing,locate_additional_public_evidence\n"
                "B,乙,2025,X_E_ENV_SYSTEM,体系,4.0,1,evidence_missing,locate_additional_public_evidence\n",
                encoding="utf-8",
            )
            coverage = root / "coverage.csv"
            coverage.write_text(
                "stock_code,company_name,annual_status,esg_status\n"
                "A,甲,collected,missing\nB,乙,collected,collected\n",
                encoding="utf-8",
            )
            rows, summary = reprioritize_evidence_gaps(gaps, coverage)
            self.assertEqual(["B", "A", "A"], [row["company_code"] for row in rows])
            self.assertEqual([1, 2, 3], [row["gap_rank"] for row in rows])
            self.assertEqual("collected", rows[0]["esg_status"])
            self.assertEqual(3, summary["gap_count"])
            self.assertEqual(2, summary["high_weight_gap_count"])
            self.assertEqual(1, summary["esg_collected_gap_count"])
            self.assertFalse(summary["scoring_authorized"])
            unknown = root / "unknown.csv"
            unknown.write_text(
                "company_code,company_name,report_year,indicator_code,indicator_name,indicator_weight,priority,status,next_action\n"
                "C,丙,2025,X_E_ENV_SYSTEM,体系,4.0,1,evidence_missing,locate_additional_public_evidence\n",
                encoding="utf-8",
            )
            with self.assertRaisesRegex(ValueError, "不在覆盖审计中"):
                reprioritize_evidence_gaps(unknown, coverage)

    def test_collect_esg_qualitative_evidence_scans_esg_reports_only(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            coverage = root / "coverage.csv"
            coverage.write_text(
                "stock_code,company_name,annual_status,esg_status\n"
                "A,甲,collected,collected\nB,乙,collected,missing\nC,丙,collected,collected\n",
                encoding="utf-8",
            )
            index = root / "index.csv"
            write_document_index(index, [
                DocumentRecord(
                    "A", "甲", 2025, "annual_report", "https://official/a.pdf",
                    "https://official/a.pdf", "data/raw/A/2025/annual_report.pdf", "abc", 12,
                ),
                DocumentRecord(
                    "A", "甲", 2025, "esg_report", "https://official/a_esg.pdf",
                    "https://official/a_esg.pdf", "data/raw/A/2025/esg_report.pdf", "def", 34,
                ),
                DocumentRecord(
                    "A", "甲", 2024, "esg_report", "https://official/a_esg_2024.pdf",
                    "https://official/a_esg_2024.pdf", "data/raw/A/2024/esg_report.pdf", "ghi", 56,
                ),
            ])
            text = root / "text/A/2025/esg_report.txt"
            text.parent.mkdir(parents=True)
            text.write_text(
                "\n=== PAGE 3 ===\n公司完善生物多样性保护体系，并设定年度目标，开展专项行动并达成修复成效。\n",
                encoding="utf-8",
            )
            rows, summary = collect_esg_qualitative_evidence(
                coverage, index, root / "text", self.methodology, 2025,
            )
            self.assertEqual({"X_E_BIODIVERSITY"}, {row.indicator_code for row in rows})
            self.assertTrue(all(row.source_file.endswith("esg_report.pdf") for row in rows))
            self.assertTrue(all(row.review_status == "pending" for row in rows))
            self.assertEqual(1, summary["esg_document_count"])
            self.assertEqual(1, summary["candidate_group_count"])
            self.assertEqual(0, summary["missing_text_count"])
            self.assertFalse(summary["scoring_authorized"])
            self.assertEqual(["C"], summary["companies_without_esg_document"])

    def test_merge_qualitative_candidates_dedupes_and_rejects_year_mix(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            from aegis_esg.esg_disclosure import write_qualitative_candidates
            first = root / "first.csv"
            second = root / "second.csv"
            candidate = QualitativeEvidenceCandidate(
                "A", "甲", 2025, "X_E_ENV_SYSTEM", "环保体系", "https://a", "a.pdf", 9,
                "环境管理体系", "公司建立环境管理体系。", .75,
            )
            write_qualitative_candidates(first, [candidate])
            write_qualitative_candidates(second, [
                candidate,
                QualitativeEvidenceCandidate(
                    "A", "甲", 2025, "X_E_ENV_SYSTEM", "环保体系", "https://b", "b.pdf", 4,
                    "环境管理体系", "ESG报告披露环境管理体系。", .75,
                ),
            ])
            merged, summary = merge_qualitative_candidate_files([first, second])
            self.assertEqual(2, len(merged))
            self.assertEqual(1, summary["duplicate_candidate_count"])
            self.assertFalse(summary["scoring_authorized"])
            other_year = root / "other_year.csv"
            write_qualitative_candidates(other_year, [QualitativeEvidenceCandidate(
                "A", "甲", 2024, "X_E_ENV_SYSTEM", "环保体系", "https://c", "c.pdf", 2,
                "环境管理体系", "上年披露环境管理体系。", .75,
            )])
            with self.assertRaisesRegex(ValueError, "报告期不一致"):
                merge_qualitative_candidate_files([first, other_year])

    def test_chinese_env_table_rows_respect_year_modes_and_units(self):
        single = "指标 单位 2025 年数据\n温室气体排放强度 吨二氧化碳当量 / 万元 1.94\n"
        rows = _extract_chinese_env_table_rows(single, 2025)
        self.assertEqual([("Q_E_GHG_INTENSITY", 1940.0)], [(code, value) for code, value, _ in rows])

        current_first = "指标 单位 2025 2024\n水资源使用强度 吨/万元 0.5 0.6\n"
        rows = _extract_chinese_env_table_rows(current_first, 2025)
        self.assertEqual([("Q_E_WATER_INTENSITY", 500.0)], [(code, value) for code, value, _ in rows])

        current_last = "指标名称 单位 2023 年 2024 年 2025\n氮氧化物排放强度 千克/万元 0.1 0.2 0.3\n"
        rows = _extract_chinese_env_table_rows(current_last, 2025)
        self.assertEqual([("Q_E_NOX_INTENSITY", 300.0)], [(code, value) for code, value, _ in rows])

        rate = "指标 单位 2025 2024\n环保投入占营业收入比例（%） 1.2 1.1\n"
        rows = _extract_chinese_env_table_rows(rate, 2025)
        self.assertEqual([("Q_S_ENV_INVEST_RATE", 1.2)], [(code, value) for code, value, _ in rows])

    def test_chinese_env_table_rows_handle_real_world_variants(self):
        footnote = "指标\n单位 2024 2025\n温室气体排放强度\n注 3\n吨二氧化碳当量 / 百万元\n54.63\n60.08\n"
        rows = _extract_chinese_env_table_rows(footnote, 2025)
        self.assertEqual([("Q_E_GHG_INTENSITY", 600.8)], [(code, value) for code, value, _ in rows])

        missing_tail = "指标 单位 2025年 2024年 2023年\n综合能源消耗强度\n吨标准煤/百万元 55.46 50.21 /\n"
        rows = _extract_chinese_env_table_rows(missing_tail, 2025)
        self.assertEqual([("Q_E_ENERGY_INTENSITY", 554.6)], [(code, value) for code, value, _ in rows])

        suffix = "用水强度\n1.1991\n吨 / 百万元\n2025 年\n"
        rows = _extract_chinese_env_table_rows(suffix, 2025)
        self.assertEqual([("Q_E_WATER_INTENSITY", 11.991)], [(code, value) for code, value, _ in rows])

        kangxi = "指标 单位 2025年\n温室⽓体排放强度\n万吨⼆氧化碳当量/百万元营业收入\n0.19\n"
        rows = _extract_chinese_env_table_rows(kangxi, 2025)
        self.assertEqual([("Q_E_GHG_INTENSITY", 19000.0)], [(code, value) for code, value, _ in rows])

        wrong_unit = "能源消耗强度\n0.266\n兆瓦时 / 百万元\n2025 年\n"
        self.assertFalse(_extract_chinese_env_table_rows(wrong_unit, 2025))

        direct_revenue = (
            "指标 单位 2025 2024 2023\n"
            "单位营收温室气体排放强度（范围 1、2） tCO2e / 百万元 3.41 3.82 2.92\n"
            "单位营收综合能源消耗密度 吉焦 / 百万元 28.10 22.90 21.39\n"
            "单位营收耗水强度 吨 / 百万元 0.35 0.08 0.10\n"
        )
        rows = _extract_chinese_env_table_rows(direct_revenue, 2025)
        values = {code: value for code, value, _ in rows}
        self.assertEqual(34.1, values["Q_E_GHG_INTENSITY"])
        self.assertAlmostEqual(9.5879448, values["Q_E_ENERGY_INTENSITY"])
        self.assertEqual(3.5, values["Q_E_WATER_INTENSITY"])

        # 单位营收排放量（吨/万元）+ PDF 千分位换行（真实样例：002506.SZ）
        wrapped_unit_revenue = (
            "指标 单位 2023年 2024年 2025年\n"
            "单位营收温室气体排放量（基于位置） 吨二氧化碳当量 / 万元 0.16 0.29 0.30\n"
            "范围一及范围二温室气体排放总量（基于位置） 吨二氧化碳当量 252,302.35 477\n"
            "044.50 465,524.43\n"
        )
        rows = _extract_chinese_env_table_rows(wrapped_unit_revenue, 2025)
        self.assertEqual([("Q_E_GHG_INTENSITY", 300.0)], [(code, value) for code, value, _ in rows])

    def test_chinese_env_table_rows_reject_ambiguous_or_wrong_denominator(self):
        no_header = "温室气体排放强度 吨二氧化碳当量 / 万元 1.94\n"
        self.assertFalse(_extract_chinese_env_table_rows(no_header, 2025))
        output_value = "指标 单位 2025 2024\n能源消耗强度 吨标准煤/万元产值 0.5 0.6\n"
        self.assertFalse(_extract_chinese_env_table_rows(output_value, 2025))
        totals = "指标 单位 2025 年数据\n温室气体排放总量（范围一、范围二） 吨二氧化碳当量 30,422,875.75\n"
        self.assertFalse(_extract_chinese_env_table_rows(totals, 2025))
        scope_only = "指标 单位 2025 年数据\n范围一温室气体排放强度 吨二氧化碳当量 / 万元 1.81\n"
        self.assertFalse(_extract_chinese_env_table_rows(scope_only, 2025))
        collapsed = "指标 单位 2025 2024\n温室气体排放强度 吨二氧化碳当量 / 万元 2.613 清洁能源发电折合碳减排量 吨 7174787\n"
        self.assertFalse(_extract_chinese_env_table_rows(collapsed, 2025))

    def _annual_doc(self, revenue_text: str) -> "CompanyDocument":
        return CompanyDocument(
            "annual_report", [PageText(8, revenue_text)], "https://a", "data/raw/A/2025/annual_report.pdf",
        )

    def _esg_doc(self, text: str) -> "CompanyDocument":
        return CompanyDocument(
            "esg_report", [PageText(12, text)], "https://e", "data/raw/A/2025/esg_report.pdf",
        )

    _CN_REVENUE = (
        "主要会计数据和财务指标\n营业收入 17,050,000,000.00 16,200,000,000.00\n"
        "利润总额 2,000,000,000.00 1,800,000,000.00\n"
    )

    def test_env_intensity_derives_from_chinese_table_total_and_summary_revenue(self):
        esg = self._esg_doc(
            "指标 单位 2023 年 2024 年 2025 年\n"
            "温室气体排放总量（含范围一及范围二） 万吨二氧化碳当量 2.50 2.90 2.95\n"
            "范围一温室气体排放总量 万吨二氧化碳当量 0.06 0.21 0.15\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertEqual("Q_E_GHG_INTENSITY", items[0].indicator_code)
        self.assertAlmostEqual(2.95e7 * 1e4 / 1.705e10, items[0].value)
        self.assertEqual("pending", items[0].status.value)
        self.assertTrue(items[0].evidence_text.startswith("中文跨表派生: "))
        self.assertIn("annual_report.pdf 第8页", items[0].evidence_text)

    def test_env_intensity_sums_split_scope1_scope2_rows(self):
        esg = self._esg_doc("范围一排放 0.19 万吨\n范围二排放 5.50 万吨\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(["Q_E_GHG_INTENSITY"], [item.indicator_code for item in items])
        self.assertAlmostEqual(5.69e7 * 1e4 / 1.705e10, items[0].value)
        self.assertIn("scope1+scope2", items[0].evidence_text)

    def test_env_intensity_prefers_annotated_total_when_conflicts(self):
        esg = self._esg_doc(
            "指标 单位 2024 年 2025 年\n"
            "温室气体排放总量（范围 1 和范围 2） 吨二氧化碳当量 18520.94 22740.25\n"
            "温室气体排放总量 吨二氧化碳当量 14719.00 16000.00\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(22740.25 * 1e3 * 1e4 / 1.705e10, items[0].value)

    def test_rejects_output_value_ghg_intensity(self):
        from aegis_esg.extraction import extract_indicator_candidates, PageText
        pages = [PageText(1, "温室气体排放强度 万元产值二氧化碳排放 0.00668 吨/万元\n")]
        items = extract_indicator_candidates(pages, "A", "甲", 2025, "u", "f")
        self.assertFalse([item for item in items if item.indicator_code == "Q_E_GHG_INTENSITY"])

    def test_env_intensity_derives_from_freeform_highlight_total(self):
        esg = self._esg_doc("实际行动。\n温室气体排放总量\n2.95万吨二氧化碳当量\n员工总数\n5,987人\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(2.95e7 * 1e4 / 1.705e10, items[0].value)

    def test_env_intensity_derives_energy_and_water_totals(self):
        esg = self._esg_doc(
            "指标 单位 2025 2024\n"
            "能源消耗总量 吨标准煤 6,216.97 7,585.57\n"
            "总用水量 吨 596,001.30 580,000\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(6216.97 * 1e3 * 1e4 / 1.705e10, values["Q_E_ENERGY_INTENSITY"])
        self.assertAlmostEqual(596001.30 * 1e3 * 1e4 / 1.705e10, values["Q_E_WATER_INTENSITY"])

    def test_env_intensity_derives_clean_energy_tce_totals(self):
        esg = self._esg_doc(
            "议题 指标 单位 2024 年 2025 年\n"
            "其中：清洁能源使用量 吨标准煤 / 4,989.23\n"
            "清洁能源消耗总量 吨标煤 227.39 157.65\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        values = {
            item.indicator_code: item.value
            for item in items if item.indicator_code == "Q_E_CLEAN_ENERGY_INTENSITY"
        }
        # 两处总量不一致时放弃，不静默选值
        self.assertEqual({}, values)
        consistent = self._esg_doc(
            "关键指标 单位 2025 年\n"
            "可再生能源消耗量 吨标准煤 3,732.21\n"
        )
        items = derive_env_intensity_candidates(
            "A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), consistent],
        )
        clean = [item for item in items if item.indicator_code == "Q_E_CLEAN_ENERGY_INTENSITY"]
        self.assertEqual(1, len(clean))
        self.assertAlmostEqual(3732.21 * 1e3 * 1e4 / 1.705e10, clean[0].value)
        self.assertTrue(clean[0].evidence_text.startswith("中文跨表派生: "))
        # 发电节约/替代标煤不得冒充消耗总量
        savings = self._esg_doc("2025 年公司风电发电量可替代标准煤 12.5 万吨。\n清洁能源发电节约标准煤 8 万吨。\n")
        self.assertFalse([
            item for item in derive_env_intensity_candidates(
                "A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), savings],
            ) if item.indicator_code == "Q_E_CLEAN_ENERGY_INTENSITY"
        ])
        # 上一节“目标”标题不得跨行误杀合法总量行；清洁/可再生数值冲突则放弃
        sectioned = self._esg_doc(
            "2025年目标\n发电企业重复用水率达到 95% 以上\n"
            "清洁能源消耗量 28.42 万吨标准煤\n"
            "可再生能源消耗量 294.96 吨标准煤\n"
        )
        self.assertFalse([
            item for item in derive_env_intensity_candidates(
                "A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), sectioned],
            ) if item.indicator_code == "Q_E_CLEAN_ENERGY_INTENSITY"
        ])

    def test_env_intensity_derives_bare_pollutants_only_in_explicit_year_matrix(self):
        esg = self._esg_doc(
            "废气污染物种类 2025年度 单位\n"
            "氮氧化物 859.57 吨\n二氧化硫 353.16 吨\n烟尘 72.76 吨\n"
            "废水排放强度 2.10 吨/万元营收\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(859.57e3 * 1000 * 1e4 / 1.705e10, values["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(353.16e3 * 1000 * 1e4 / 1.705e10, values["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(72.76e3 * 1000 * 1e4 / 1.705e10, values["Q_E_PM_INTENSITY"])
        rejected = self._esg_doc("氮氧化物 859.57 吨\n二氧化硫 353.16 吨\n烟尘 72.76 吨\n")
        self.assertFalse(derive_env_intensity_candidates(
            "A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), rejected],
        ))

    def test_env_intensity_maps_zero_pollutant_from_explicit_ascending_section(self):
        esg = self._esg_doc(
            "披露指标 单位 2024 年 2025 年\n气体污染物排放\n"
            "颗粒物 千克 2 2\n氮氧化物 千克 5 0\n硫氧化物 千克 4 0\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        values = {item.indicator_code: item.value for item in items}
        self.assertEqual(0, values["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(2 * 1000 * 1e4 / 1.705e10, values["Q_E_PM_INTENSITY"])
        self.assertNotIn("Q_E_SO2_INTENSITY", values)

    def test_env_intensity_rejects_partial_scope_and_macro_narrative(self):
        esg = self._esg_doc(
            "指标 单位 2025 2024\n"
            "范围一温室气体排放总量 万吨二氧化碳当量 0.15 0.21\n"
            "直接能源消耗总量 吨标准煤 2,053,921.82 2,000,000\n"
            "循环用水总量 吨 888,000 850,000\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertFalse(items)
        macro = self._esg_doc("增加温室气体排放总量约 30 亿吨，覆盖全国碳市场。\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), macro])
        self.assertFalse(items)
        reduced = self._esg_doc("减排措施直接减少的温室气体排放总量\n吨二氧化碳当量\n15.67\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), reduced])
        self.assertFalse(items)

    def test_env_intensity_suppressed_by_disclosed_revenue_intensity(self):
        # 公司已按收入口径披露强度（即使版式未被直接规则解析）时抑制派生；
        # 同年份表头保证总量本身可抽取，断言为空即抑制生效
        esg = self._esg_doc(
            "Indicator Unit 2023 2024 2025\n"
            "Total GHG Emissions (Scope 1 + Scope 2) tCO₂e 1,618,370 1,594,055 1,498,435\n"
            "Total GHG Emission Economic Intensity (Scope 1 + Scope 2) tCO₂e /\n10,000 yuan 0.062 0.059 0.050\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertFalse(items)
        # 对照：移除收入口径强度披露后派生正常生成
        no_intensity = self._esg_doc(
            "Indicator Unit 2023 2024 2025\n"
            "Total GHG Emissions (Scope 1 + Scope 2) tCO₂e 1,618,370 1,594,055 1,498,435\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), no_intensity])
        self.assertEqual(1, len(items))
        # 产值口径披露不抑制：方法论只认营业收入分母
        output_value = self._esg_doc(
            "指标 单位 2023 年 2024 年 2025 年\n"
            "温室气体排放总量（含范围一及范围二） 万吨二氧化碳当量 2.50 2.90 2.95\n"
            "温室气体排放强度（万元产值温室气体排放强度） 吨二氧化碳当量 / 万元 0.02 0.01 0.016\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), output_value])
        self.assertEqual(1, len(items))

    def test_env_intensity_shenhua_style_wan_ton_and_operating_scope(self):
        # 神华样例：万吨单位 + 换行陈述 + 所属生产经营类企业全口径（非局部子公司）
        esg = self._esg_doc(
            "一级指标 二级指标 2023年 2024年 2025年\n"
            "二氧化硫排放总量（万吨） 2.27 2.22 1.99\n"
            "氮氧化物排放总量（万吨） 4.82 4.78 4.25\n"
            "2025年\n"
            "二氧化硫排放总量为\n"
            "1.99万吨\n"
            "2025年，公司所属生产经营类企业能源消费总量为7,282.08万吨标煤。\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        codes = {item.indicator_code for item in items}
        self.assertIn("Q_E_SO2_INTENSITY", codes)
        self.assertIn("Q_E_NOX_INTENSITY", codes)
        self.assertIn("Q_E_ENERGY_INTENSITY", codes)
        so2 = next(item for item in items if item.indicator_code == "Q_E_SO2_INTENSITY")
        self.assertAlmostEqual(1.99e7 * 1e3 * 1e4 / 1.705e10, so2.value)

    def test_summary_revenue_cnpc_style_million_yuan(self):
        from aegis_esg.extraction import _extract_summary_revenue

        text = (
            "按中国企业会计准则编制的主要财务数据\n"
            "（1）主要会计数据及财务指标\n"
            "单位：人民币百万元\n"
            "项目 2025 年 2024 年\n"
            "增减(%) 营业收入 2,864,469 2,937,981 (2.5) 3,012,812\n"
            "营业利润 234,579 255,286 (8.1) 253,522\n"
        )
        parsed = _extract_summary_revenue(text, True)
        self.assertIsNotNone(parsed)
        current, previous, _evidence = parsed
        self.assertEqual(2_864_469_000_000, current)
        self.assertEqual(2_937_981_000_000, previous)

    def test_env_intensity_cn_narrative_group_anchored_total(self):
        # 叙述式：报告主体锚定的句中总量（真实样例：600163.SH）
        esg = self._esg_doc(
            "2025年关键绩效：\n"
            "2025 年，公司下属各项目累计完成发电量28.75亿千瓦时\n"
            "2025 年，公司温室气体排放总量5532.27吨\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(5532.27 * 1e3 * 1e4 / 17.05e9, items[0].value)

    def test_env_intensity_cn_narrative_rejects_targets_and_macro(self):
        # 目标句（较基准年降低/控制在…以内）与宏观叙述不得抽取
        for text in (
            "2025 年，公司温室气体排放总量较2023年降低42%\n",
            "2025 年，公司温室气体排放总量同比减少4.35%\n",
            "到2030年，公司温室气体排放总量控制在100吨以内\n",
            "2025年扩容企业超过1300家，增加温室气体排放总量约30亿吨\n",
        ):
            esg = self._esg_doc(text)
            self.assertFalse(
                derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg]),
                text,
            )

    def test_env_intensity_million_tonnes_unit_with_scope_closure(self):
        # 百万吨二氧化碳当量单位 + 范围一/二闭环核验（真实样例：600167.SH）
        esg = self._esg_doc(
            "直接温室气体排放（范围1） 百万吨二氧化碳当量\n0.0264\n"
            "间接温室气体排放（范围2） 百万吨二氧化碳当量\n0.0004\n"
            "温室气体排放总量（范围1+范围2） 百万吨二氧化碳当量\n0.0268\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(0.0268 * 1e9 * 1e4 / 17.05e9, items[0].value)

    def test_env_intensity_spaced_scope_annotation(self):
        # 带空格的范围标注（真实样例：300880.SZ “（范围 1+ 范围 2）”）
        esg = self._esg_doc("温室气体排放总量\n（范围 1+ 范围 2）\n2.95万吨二氧化碳当量\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))

    def test_chinese_env_table_disclosure_item_header_and_wanyuan_revenue_unit(self):
        # “披露项 单位”表头 + 万元营收单位（真实样例：300207.SZ 欣旺达）
        from aegis_esg.extraction import _extract_chinese_env_table_rows
        rows = _extract_chinese_env_table_rows(
            "披露项 单位 2023 2024 2025\n"
            "温室气体排放强度 吨二氧化碳当量 / 万元营收 0.21 0.18 0.23\n",
            2025,
        )
        self.assertEqual(1, len(rows))
        code, value, _ = rows[0]
        self.assertEqual("Q_E_GHG_INTENSITY", code)
        self.assertAlmostEqual(230.0, value)

    def test_chinese_hazardous_waste_density_postfix_billion_revenue_unit(self):
        from aegis_esg.extraction import _extract_chinese_env_table_rows
        rows = _extract_chinese_env_table_rows(
            "指标 2023 年 2024 年 2025 年 单位\n"
            "危险废弃物密度 2.9 3.2 3.62 吨 / 亿元人民币营业收入\n",
            2025,
        )
        self.assertEqual(1, len(rows))
        self.assertEqual("Q_E_HAZ_WASTE_INTENSITY", rows[0][0])
        self.assertAlmostEqual(.362, rows[0][1])

    def test_chinese_hazardous_waste_revenue_unit_variants(self):
        from aegis_esg.extraction import _extract_chinese_env_table_rows
        cases = (
            ("指标 单位 2025 年 2024 年\n危险废弃物产生强度 吨 / 万元营收 0.0006 0.0007\n", .6),
            ("类别 2025 年 2024 年 单位\n危险废弃物产生强度 0.0002 0.0001 吨 / 万元人民币营业收入\n", .2),
            ("指标 单位 2025 年 2024 年\n单位营收危险废物密度 吨 / 百万元 0.06 0.05\n", .6),
        )
        for text, expected in cases:
            rows = _extract_chinese_env_table_rows(text, 2025)
            self.assertEqual(1, len(rows), text)
            self.assertAlmostEqual(expected, rows[0][1])

    def test_chinese_hazardous_waste_named_header_and_single_revenue_value(self):
        from aegis_esg.extraction import _extract_chinese_env_table_rows
        named = _extract_chinese_env_table_rows(
            "指标名称 指标单位 2024 年数值 2025 年数值\n"
            "危险废物产生强度 吨 / 万元 0.00034 0.00235\n", 2025,
        )
        single = _extract_chinese_env_table_rows(
            "危险废弃物产生强度 0.01 吨 / 百万元营业收入\n", 2025,
        )
        self.assertAlmostEqual(2.35, named[0][1])
        self.assertAlmostEqual(.1, single[0][1])

    def test_env_intensity_pollutant_statement_rows_gram_canonical(self):
        # 法定披露分号终止总量句；SO2/NOx/PM方法论口径为克/万元（真实样例：000600.SZ）
        esg = self._esg_doc(
            "优于国家超低排放标准。氮氧化物排放总量 4,697.8 吨；"
            "二氧化硫排放总量 2,724.11 吨；烟尘排放总量 391.06 吨。\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(4697.8 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(2724.11 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(391.06 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_PM_INTENSITY"])

    def test_env_intensity_partial_scope_sentence_rejected(self):
        # 句首部分口径锚点覆盖全句（真实样例：000883.SZ“4 家火电企业二氧化硫…氮氧化物…烟尘…”）
        esg = self._esg_doc(
            "脱硫和除尘设施投运率 100%。4 家火电企业二氧化硫排放量为 1,074.22 吨，"
            "氮氧化物排放量为 2,632.63 吨，烟尘排放量为 149.26 吨。\n"
        )
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg]))

    def test_env_intensity_scope3_prefixed_statement_rejected(self):
        # 范围三前缀总量不得当作范围一+二口径（真实样例：002506.SZ）
        esg = self._esg_doc("报告期内，公司范围三温室气体排放总量为 26,410,296.89 吨二氧化碳当量。\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg]))

    def test_env_intensity_scope_rows_arabic_numeral_not_total(self):
        # 阿拉伯数字范围标签的子项不得当作集团总量；拼版残块不制造假冲突（真实样例：603396.SH）
        esg = self._esg_doc(
            "公司总部与金辰自动化温室气体排放数据\n"
            "温室气体排放总量\n范围1温室气体排放总量\n3,582.42\n300.60\n吨二氧化碳当量\n"
            "范围2温室气体排放总量\n3,281.82\n吨二氧化碳当量\n"
            "指标 单位 2024年 2025年\n"
            "温室气体排放总量 吨二氧化碳当量 6,181.36 3,582.42\n"
            "范围1温室气体排放总量 吨二氧化碳当量 307.28 300.60\n"
            "范围2温室气体排放总量 吨二氧化碳当量 5,874.08 3,281.82\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(3582.42 * 1e3 * 1e4 / 17.05e9, items[0].value)

    def test_env_intensity_revenue_prefixed_disclosure_suppresses(self):
        # “单位营收/每百万营收”前缀即收入口径披露，抑制派生（真实样例：002506.SZ、000600.SZ）
        for text in (
            "单位营收温室气体排放量（基于位置） 吨二氧化碳当量 / 万元 0.16 0.29 0.30\n温室气体排放总量（含范围一及范围二） 万吨二氧化碳当量 2.95\n",
            "每百万营收能源消耗总量 万吨标准煤 0.0775 0.0792 0.0789\n能源消耗总量 万吨标准煤 1,797.17\n",
        ):
            esg = self._esg_doc(text)
            self.assertFalse(
                derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg]),
                text[:30],
            )

    def test_env_intensity_wrapped_thousand_separator_year_table(self):
        # PDF 把千分位数字拆行时仍取2025列总量（真实样例：002506.SZ 477,044.50）
        esg = self._esg_doc(
            "指标 单位 2023年 2024年 2025年\n"
            "范围一及范围二温室气体排放总量（基于位置） 吨二氧化碳当量 252,302.35 477\n"
            "044.50 465,524.43\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(465524.43 * 1e3 * 1e4 / 17.05e9, items[0].value)

    def test_env_intensity_so2_header_unit_and_chemical_formula_tables(self):
        # 表头单位行（300932.SZ）与废气化学式年表（600968.SH）
        header = self._esg_doc(
            "▎废气排放数据\n污染物种类 2025 年排放量（吨）\n"
            "废气污染物排放总量 1.0341\n烟尘 / 颗粒物排放总量 0.4470\n"
            "二氧化硫排放总量 0.0212\n氮氧化物排放总量 0.0212\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), header])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(0.0212e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])
        formula = self._esg_doc(
            "废气污染物排放 单位 2023 2024 2025\n"
            "SO2 吨 1.27 0.81 1.27\n"
            "NOX 吨 21.29 21.29 19.41\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), formula])
        so2 = [item for item in items if item.indicator_code == "Q_E_SO2_INTENSITY"]
        self.assertEqual(1, len(so2))
        self.assertAlmostEqual(1.27e6 * 1e4 / 17.05e9, so2[0].value)

    def test_env_intensity_short_label_so2_year_and_total_columns(self):
        # 短标签两列表（000791.SZ）与分单位+2025总计列（605011.SH）
        two_year = self._esg_doc(
            "废气污染物种类 单位 2024 年数值 2025 年数值\n"
            "氮氧化物 吨 2,645.91 2,708.82\n"
            "二氧化硫 吨 1,647.42 1,525.13\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), two_year])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1525.13e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(2708.82e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])
        multi = self._esg_doc(
            "指标 单位 上海金联 临江环保 丽水杭丽 安吉天子湖\n2025总计\n"
            "二氧化硫\n吨\n1.61\n16.95\n16.16\n15.01\n49.73\n"
            "氮氧化物\n吨\n37.08\n39.85\n76.62\n38.30\n191.85\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), multi])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(49.73e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(191.85e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])

    def test_env_intensity_production_value_disclosure_not_suppressing(self):
        # 产值口径强度披露不抑制派生（真实样例：000600.SZ 千克/万元产值）
        esg = self._esg_doc(
            "氮氧化物（NOx）排放强度 千克/万元产值 2.0144 2.1299 2.2803\n"
            "氮氧化物排放总量 4,697.8 吨。\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))

    def test_env_intensity_year_column_precedence_over_stray_single_value(self):
        # 年列表列定位值优先；同页单值行捕获首年列的拼版残片不制造假冲突（真实样例：000027.SZ）
        esg = self._esg_doc(
            "指标 单位 2023 年 2024 年 2025 年\n"
            "废水排放量 万立方米 36.28 50.93 58.24\n"
            "废水排放量 万立方米 36.28\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(58.24 * 1e7 * 1e4 / 17.05e9, items[0].value)

    def test_env_intensity_emission_reduction_compound_prefix_allowed(self):
        # “减排标准/减排任务”等复合名词不是目标措辞（真实样例：000600.SZ“深度减排标准。氮氧化物…”）
        esg = self._esg_doc("优于国家超低排放标准和河北省深度减排标准。氮氧化物排放总量 4,697.8 吨；\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        self.assertEqual(1, len(items))

    def test_env_intensity_external_flue_gas_actual_rows_not_permit_limits(self):
        # “外排废气中…量”是实际排放；紧邻的核定年度排放量是许可上限，不得混入（真实样例：600028.SH）
        esg = self._esg_doc(
            "指标 单位 2023 2024 2025\n"
            "外排废气中二氧化硫量 吨 4,661 4,652 4,481\n"
            "外排废气中氮氧化物量 吨 19,984 18,482 18,484\n"
            "核定的年度二氧化硫排放量 吨 9,000 9,000 9,000\n"
            "核定的年度氮氧化物排放量 吨 30,000 30,000 30,000\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertEqual({"Q_E_NOX_INTENSITY", "Q_E_SO2_INTENSITY"}, set(by_code))
        self.assertAlmostEqual(18484 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(4481 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])

    def test_env_intensity_external_flue_gas_split_pdf_table(self):
        # PDF抽取将每个单元格拆行时，仍按显式年份列取本期值（真实样例：600028.SH）
        esg = self._esg_doc(
            "指标\n单位\n2024 2025\n"
            "外排废气中二氧化硫量\n吨\n4,652\n4,481\n"
            "外排废气中氮氧化物量\n吨\n18,482\n18,484\n"
            "核定的年度二氧化硫排放总量\n吨\n46,258\n37,208\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertEqual({"Q_E_NOX_INTENSITY", "Q_E_SO2_INTENSITY"}, set(by_code))
        self.assertAlmostEqual(18484 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(4481 * 1e6 * 1e4 / 17.05e9, by_code["Q_E_SO2_INTENSITY"])

    def test_env_intensity_single_year_split_nox_pm_not_sulfur_oxides(self):
        # 单年拆行KPI只映射同名NOx和明确PM，不把硫氧化物等同SO2（真实样例：600023.SH）
        esg = self._esg_doc(
            "指标\n单位\n2025年\n氮氧化物排放量\n吨\n21242\n"
            "硫氧化物排放量\n吨\n10865\n"
            "悬浮粒子与颗粒物（PM）排放量\n吨\n1038\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertEqual({"Q_E_NOX_INTENSITY", "Q_E_PM_INTENSITY"}, set(by_code))
        self.assertAlmostEqual(21242e6 * 1e4 / 17.05e9, by_code["Q_E_NOX_INTENSITY"])
        self.assertAlmostEqual(1038e6 * 1e4 / 17.05e9, by_code["Q_E_PM_INTENSITY"])

    def test_env_intensity_pm_parenthetical_and_year_infix_labels(self):
        for label in ("颗粒物（PM）排放量", "颗粒物(PM)年排放量"):
            esg = self._esg_doc(f"指标 单位 2025 年\n{label} 吨 7.55\n")
            items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
            self.assertEqual(["Q_E_PM_INTENSITY"], [item.indicator_code for item in items], label)

    def test_env_intensity_chinese_postfix_unit_year_table(self):
        # 年份在前、单位在末的集团附录表按2025列取值（真实样例：601727.SH）
        esg = self._esg_doc(
            "指标 2023 年 2024 年 2025 年 单位\n"
            "氮氧化物排放量 50.61 51.06 41.42 吨\n"
            "硫氧化物排放量 0.71 0.92 0.54 吨\n"
            "颗粒物排放量 26.36 39.31 48.54 吨\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertEqual({"Q_E_NOX_INTENSITY", "Q_E_PM_INTENSITY"}, set(by_code))
        self.assertAlmostEqual(48.54e6 * 1e4 / 17.05e9, by_code["Q_E_PM_INTENSITY"])

    def test_env_intensity_pollutant_table_order_proved_by_per_value_years(self):
        # 区间标题本身不决定列序；首行逐值年份才锚定后续三值行（真实样例：600956.SH）
        esg = self._esg_doc(
            "2023-2025 年新天废气排放量\n"
            "指标 硫氧化物排放量 单位 千克 2025 年 28.63 2024 年 37.23 2023 年 10.44\n"
            "氮氧化物排放量 千克 11,564.93 14,191.81 5,314.12\n"
            "颗粒物排放量 千克 1,046.07 1,360.38 509.19\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1046.07e3 * 1e4 / 17.05e9, by_code["Q_E_PM_INTENSITY"])
        self.assertNotIn("Q_E_SO2_INTENSITY", by_code)
        no_anchor = self._esg_doc("2023-2025 年新天废气排放量\n颗粒物排放量 千克 1,046.07 1,360.38 509.19\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), no_anchor]))

    def test_env_intensity_summary_revenue_rmb_million_scale(self):
        # 财务摘要常按列重复“人民币百万元”，分母必须还原为人民币元（真实样例：600028.SH）
        annual = self._annual_doc(
            "主要财务数据及指标\n1 按中国企业会计准则编制的主要会计数据和财务指标\n"
            "项目 截至 12 月 31 日止年度\n2025 年 2024 年\n人民币\n百万元\n人民币\n百万元\n"
            "营业收入 2,783,583 3,074,562\n利润总额 43,184 70,513\n"
        )
        esg = self._esg_doc(
            "指标\n单位\n2024 2025\n外排废气中氮氧化物量\n吨\n18,482\n18,484\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(18484 * 1e6 * 1e4 / (2783583 * 1e6), items[0].value)

    def test_env_intensity_summary_revenue_explicit_thousand_currency_scale(self):
        annual = self._annual_doc(
            "主要会计数据和财务指标\n单位：千元 币种：人民币\n"
            "营业收入 125,958,695 115,456,181\n利润总额 8,000,000 7,000,000\n"
        )
        esg = self._esg_doc("指标 单位 2025 年\n颗粒物（PM）排放量 吨 48.54\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(48.54e6 * 1e4 / 125958695e3, items[0].value)

    def test_env_intensity_chinese_consolidated_income_statement_revenue_fallback(self):
        annual = CompanyDocument("annual_report", [PageText(101, (
            "合并利润表\n2025 年 1—12 月\n单位：元 币种：人民币\n"
            "项目 附注 2025 年度 2024 年度\n一、营业总收入 七、61 35,425,723,476.94 35,423,751,313.61\n"
            "其中：营业收入 七、61 35,425,723,476.94 35,423,751,313.61\n"
        ))], "url", "annual.pdf")
        esg = self._esg_doc("指标 单位 2025 年\n颗粒物（PM）排放量 吨 814.30\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(814.3e6 * 1e4 / 35425723476.94, items[0].value)

    def test_env_intensity_solid_and_haz_waste_current_first_table(self):
        # 一般固废/危废产生量年列表派生（真实样例：000922.SZ）
        esg = self._esg_doc(
            "指标名称 单位 2025年 2024年\n"
            "一般工业固体废物产生量 吨 10860 11367.87\n"
            "危险废物产生量 吨 492.1 456.45\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(10860 * 1e3 * 1e4 / 17.05e9, by_code["Q_E_SOLID_WASTE_INTENSITY"])
        self.assertAlmostEqual(492.1 * 1e3 * 1e4 / 17.05e9, by_code["Q_E_HAZ_WASTE_INTENSITY"])

    def test_ghg_reduction_current_first_two_period_table(self):
        # 当前年优先两期总量行派生减排率，两期范围闭环均通过（真实样例：603396.SH版式）
        esg = self._esg_doc(
            "指标 单位 2025 2024\n"
            "温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
            "范围1温室气体排放总量 吨二氧化碳当量 300.60 307.28\n"
            "范围2温室气体排放总量 吨二氧化碳当量 3,281.82 5,874.08\n"
        )
        items = derive_ghg_reduction_candidates("A", "甲", 2025, [esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual((6181.36 - 3582.42) / 6181.36 * 100, items[0].value)

    def test_ghg_reduction_current_last_three_year_table(self):
        # 当前年靠后三年表取末两列；排放量上升时减排率为负（真实样例：000400.SZ版式）
        esg = self._esg_doc(
            "指标 单位 2023 年 2024 年 2025 年\n"
            "温室气体排放总量 吨二氧化碳当量 2.50 2.90 2.95\n"
        )
        items = derive_ghg_reduction_candidates("A", "甲", 2025, [esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual((2.90 - 2.95) / 2.90 * 100, items[0].value)

    def test_ghg_reduction_scope12_location_based_total_with_market_row(self):
        esg = self._esg_doc(
            "披露项 单位 2023 年 2024 年 2025 年\n"
            "范围一温室气体排放量 吨二氧化碳当量 1,309.95 1,295.60 5,205.78\n"
            "范围二温室气体排放量（基于位置） 吨二氧化碳当量 250,992.40 475,748.90 460,318.65\n"
            "范围一及范围二温室气体排放总量（基于位置） 吨二氧化碳当量 252,302.35 477,044.50 465,524.43\n"
            "范围一及范围二温室气体排放总量（基于市场） 吨二氧化碳当量 250,177.45 324,256.22 168,299.90\n"
        )
        items = derive_ghg_reduction_candidates("A", "甲", 2025, [esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual((477044.50 - 465524.43) / 477044.50 * 100, items[0].value)

    def test_ghg_reduction_scope_closure_failure_rejected(self):
        # 本期或上期任一时期不满足范围一+二闭环即拒绝
        for scope2_row in (
            "范围2温室气体排放总量 吨二氧化碳当量 3,281.82 5,874.08\n",
            "范围2温室气体排放总量 吨二氧化碳当量 3,000.00 5,874.08\n",
        ):
            esg = self._esg_doc(
                "指标 单位 2025 2024\n"
                "温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
                "范围1温室气体排放总量 吨二氧化碳当量 300.60 307.28\n"
                + scope2_row
            )
            items = derive_ghg_reduction_candidates("A", "甲", 2025, [esg])
            if scope2_row.startswith("范围2温室气体排放总量 吨二氧化碳当量 3,281.82"):
                self.assertEqual(1, len(items))
            else:
                self.assertFalse(items)

    def test_ghg_reduction_scope3_contamination_rejected(self):
        # 同页范围三正值且总量未标注范围一+二时拒绝（真实样例：600803.SH口径）
        esg = self._esg_doc(
            "指标 单位 2025 2024\n"
            "温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
            "范围三温室气体排放量 吨二氧化碳当量 415.80 390.20\n"
        )
        self.assertFalse(derive_ghg_reduction_candidates("A", "甲", 2025, [esg]))

    def test_ghg_reduction_direct_disclosure_suppresses(self):
        # 公司已披露同口径减排率/同比下降时抑制派生；目标措辞不构成披露
        disclosed = self._esg_doc(
            "指标 单位 2025 2024\n"
            "温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
            "2025年，公司温室气体排放总量同比下降 42.05%。\n"
        )
        self.assertFalse(derive_ghg_reduction_candidates("A", "甲", 2025, [disclosed]))
        target = self._esg_doc(
            "指标 单位 2025 2024\n"
            "温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
            "公司设定温室气体减排率目标 5%。\n"
        )
        self.assertEqual(1, len(derive_ghg_reduction_candidates("A", "甲", 2025, [target])))

    def test_ghg_reduction_conflicting_tables_skipped(self):
        # 两表减排率不一致时放弃，不静默选值
        annual = self._annual_doc(
            "指标 单位 2025 2024\n温室气体排放总量 吨二氧化碳当量 3,582.42 6,181.36\n"
        )
        esg = self._esg_doc(
            "指标 单位 2025 2024\n温室气体排放总量 吨二氧化碳当量 3,500.00 6,181.36\n"
        )
        self.assertFalse(derive_ghg_reduction_candidates("A", "甲", 2025, [annual, esg]))

    def test_ghg_reduction_single_year_mode_no_derivation(self):
        # 单年表头只有本期值，不能派生两期减排率
        esg = self._esg_doc(
            "指标 单位 2025 年数据\n温室气体排放总量 吨二氧化碳当量 3,582.42\n"
        )
        self.assertFalse(derive_ghg_reduction_candidates("A", "甲", 2025, [esg]))

    def test_ghg_reduction_period_boundary_break_rejected(self):
        # 附注明示上年仅统计集团本部/不含子公司时两期不可比（真实样例：600903.SH注2）
        esg = self._esg_doc(
            "指标 单位 2023 2024 2025\n"
            "温室气体排放总量（范围一+范围二） 吨二氧化碳当量 1,997.56 2,010.71 17,421.38\n"
            "注2：2023年-2024年温室气体排放量仅统计了集团本部数据，不含子公司。\n"
        )
        self.assertFalse(derive_ghg_reduction_candidates("A", "甲", 2025, [esg]))

    def test_ghg_reduction_direct_single_year_table_row(self):
        # 单年表头“同比下降 %”行直接披露（真实样例：601991.SH），强度行不匹配
        from aegis_esg.extraction import _extract_chinese_ghg_reduction_direct
        text = (
            "指标 单位 2025 年\n"
            "温室气体排放总量 万吨 19694.26\n"
            "温室气体排放同比下降 % 6.63\n"
            "综合碳排放强度同比下降 % 5.14\n"
        )
        self.assertEqual((6.63, "中文减排率直接披露: 温室气体排放同比下降 % 6.63"),
                         _extract_chinese_ghg_reduction_direct(text, 2025))

    def test_ghg_reduction_direct_narrative_yoy(self):
        # 公司锚定同比减少叙述（真实样例：300776.SZ）
        from aegis_esg.extraction import _extract_chinese_ghg_reduction_direct
        text = "公司2025年温室气体排放总量同比减少4.35%。公司通过优化生产交付规划，科学调整出货时间。\n"
        value, evidence = _extract_chinese_ghg_reduction_direct(text, 2025)
        self.assertEqual(4.35, value)
        self.assertTrue(evidence.startswith("中文减排率直接披露: "))

    def test_ghg_reduction_direct_rejects_non_yoy_and_target(self):
        # 目标/峰值/累计/较基准年/强度口径与净排放一律拒绝，非报告期锚定拒绝
        from aegis_esg.extraction import _extract_chinese_ghg_reduction_direct
        for text in (
            "公司计划2026年温室气体排放总量同比减少5%。\n",
            "温室气体净排放量比峰值下降 7%-10%。\n",
            "温室气体盘查与产品碳足迹认证，累计下降 25.41%。\n",
            "公司温室气体排放量较 2019 年下降 48%。\n",
            "公司2025年温室气体排放强度同比下降 3.0%。\n",
            "公司2030年温室气体排放总量同比减少 6%。\n",
            "温室气体排放强度，2030 年减少 6%，直至到 2050 年减少 80%。\n",
        ):
            self.assertIsNone(_extract_chinese_ghg_reduction_direct(text, 2025), text)

    def test_ghg_reduction_direct_auto_confirm_closed_loop(self):
        # 直接披露候选经严格前缀自动确认闭环
        candidate = Observation(
            "A", "甲", 2025, "Q_E_GHG_REDUCTION_RATE", 6.63, ValueStatus.PENDING,
            source_file="esg_report.pdf", confidence=.94,
            evidence_text="中文减排率直接披露: 温室气体排放同比下降 % 6.63",
        )
        confirmed, unresolved, decisions = resolve_pending_candidates([candidate])
        self.assertEqual([6.63], [item.value for item in confirmed])
        self.assertFalse(unresolved)
        self.assertEqual("strict_extraction_evidence_consistent", decisions[0].reason)

    def test_social_invest_note_table_donation(self):
        # 营业外支出附注：本期发生额首列即当期对外捐赠（真实样例：000400.SZ/000531.SZ）
        from aegis_esg.social_invest import derive_social_invest_candidates
        annual = self._annual_doc(
            self._CN_REVENUE +
            "\n56、营业外支出\n单位：元\n项目 本期发生额 上期发生额 计入当期非经常性损益的金额\n"
            "对外捐赠 224,016.92 97,606.72 224,016.92\n非流动资产毁损报废损失 4,377,443.43 3,438,525.72 4,377,443.43\n"
        )
        items = derive_social_invest_candidates("A", "甲", 2025, [annual])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(224016.92 * 100 / 17.05e9, items[0].value)

    def test_social_invest_note_glued_requires_page_unit(self):
        # 附注行首值=第三值为结构锚点，但单位须本页“单位：元”明示（000407跨页场景不猜单位）
        from aegis_esg.social_invest import derive_social_invest_candidates
        annual = self._annual_doc(
            self._CN_REVENUE +
            "\n罚款支出 69,393.67 60,000.00 69,393.67\n对外捐赠 169,500.00 69,585.00 169,500.00\n其他 1,535,543.01 4,688,451.41 1,535,543.01\n"
        )
        self.assertFalse(derive_social_invest_candidates("A", "甲", 2025, [annual]))

    def test_social_invest_value_first_and_narrative(self):
        # ESG亮点块值先式与公司锚定叙述式（真实样例：000531.SZ、000027.SZ）
        from aegis_esg.social_invest import derive_social_invest_candidates
        esg = self._esg_doc(
            "环保投入 4,153.91万元 环境领域违法违规事件 0件\n"
            "2025年，公司安全生产投入为 2,636.20 万元。\n"
        )
        items = derive_social_invest_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(4153.91e4 * 100 / 17.05e9, by_code["Q_S_ENV_INVEST_RATE"])
        self.assertAlmostEqual(2636.2e4 * 100 / 17.05e9, by_code["Q_S_SAFETY_INVEST_RATE"])

    def test_social_invest_suppressed_by_disclosed_rate(self):
        # 公司已披露占营收比例时抑制派生（真实样例：000543.SZ、000400.SZ）
        from aegis_esg.social_invest import derive_social_invest_candidates
        esg = self._esg_doc("环保总投入 3.67 亿元 环保总投入占营业收入比例 1.35%\n")
        self.assertFalse(derive_social_invest_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg]))

    def test_social_invest_conflicting_totals_skipped(self):
        # 年报营业外支出与ESG披露不一致时放弃（真实样例：000400.SZ 22.4万 vs 10万）
        from aegis_esg.social_invest import derive_social_invest_candidates
        annual = self._annual_doc(
            self._CN_REVENUE +
            "\n营业外支出\n单位：元\n项目 本期发生额 上期发生额 计入当期非经常性损益的金额\n"
            "对外捐赠 224,016.92 97,606.72 224,016.92\n"
        )
        esg = self._esg_doc("公司对外捐赠 10 万元。\n")
        self.assertFalse(derive_social_invest_candidates("A", "甲", 2025, [annual, esg]))

    def test_social_invest_rejects_cumulative_plan_and_non_invest_labels(self):
        # 累计口径/预算计划/费用计提/接受捐赠/税会差异一律拒绝
        from aegis_esg.social_invest import derive_social_invest_candidates
        annual = self._annual_doc(self._CN_REVENUE)
        for text in (
            "累计环保投入 5,000 万元。\n",
            "公司计划环保投入 5,000 万元。\n",
            "足额计提安全生产费用共计 1,105 万元。\n",
            "接受捐赠 500 万元。\n",
            "公益捐赠支出超出标准部分 699,613.13 174,903.28\n",
            "下一年度环保投入预算 5,000 万元。\n",
        ):
            self.assertFalse(
                derive_social_invest_candidates("A", "甲", 2025, [annual, self._esg_doc(text)]),
                text,
            )

    def test_env_intensity_skips_existing_candidates_and_conflicts(self):
        esg = self._esg_doc("温室气体排放总量\n2.95万吨二氧化碳当量\n")
        documents = [self._annual_doc(self._CN_REVENUE), esg]
        skipped = derive_env_intensity_candidates(
            "A", "甲", 2025, documents, frozenset({"Q_E_GHG_INTENSITY"}),
        )
        self.assertFalse(skipped)
        conflict = [
            self._annual_doc(self._CN_REVENUE),
            self._esg_doc("温室气体排放总量\n2.95万吨二氧化碳当量\n"),
            CompanyDocument("esg_report", [PageText(20, "温室气体排放总量\n3.10万吨二氧化碳当量\n")], "https://e2", "data/raw/A/2025/esg_report2.pdf"),
        ]
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, conflict))

    def test_env_intensity_requires_revenue_and_plausible_value(self):
        esg = self._esg_doc("温室气体排放总量\n2.95万吨二氧化碳当量\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [esg]))
        # 总量相对营收超出量级防护区间时拒绝（拦截单位错配）
        huge = self._esg_doc("温室气体排放总量\n50,000万吨二氧化碳当量\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), huge]))

    def test_env_intensity_derives_from_english_group_total_and_rmb_statement(self):
        annual = CompanyDocument("annual_report", [PageText(99, (
            "Consolidated Statement of Comprehensive Income\n"
            "For the year ended 31 December 2025\n"
            "2025 2024\nNote RMB'million RMB'million\n"
            "Revenue 收入 5 5,000 4,800\n"
            "Profit for the year 620 590\n"
        ))], "https://a", "data/raw/A/2025/annual_report.pdf")
        esg = self._esg_doc(
            "During the Reporting Period, the Group's total GHG emissions were "
            "154,166.64 tonnes of carbon dioxide equivalent.\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(154166.64 * 1e3 * 1e4 / 5e9, items[0].value)
        self.assertTrue(items[0].evidence_text.startswith("English cross-document derived: "))

    def test_env_intensity_rejects_equity_basis_and_foreign_currency(self):
        annual_hkd = CompanyDocument("annual_report", [PageText(99, (
            "Consolidated Statement of Profit or Loss\n"
            "2025 2024\nHK$'000 HK$'000\n"
            "Revenue 5,000,000 4,800,000\n"
        ))], "https://a", "data/raw/A/2025/annual_report.pdf")
        esg = self._esg_doc("The Group's total GHG emissions were 154,166.64 tonnes of carbon dioxide equivalent.\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [annual_hkd, esg]))
        equity = self._esg_doc(
            "The Group's total GHG emissions were 45,783 kilotonnes of carbon dioxide "
            "equivalent (CO2e) on an equity basis.\n"
        )
        annual_rmb = CompanyDocument("annual_report", [PageText(99, (
            "Consolidated Statement of Comprehensive Income\n"
            "2025 2024\nRMB'million RMB'million\n"
            "Revenue 5,000 4,800\n"
        ))], "https://a", "data/raw/A/2025/annual_report.pdf")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [annual_rmb, equity]))

    def test_env_intensity_english_year_header_table(self):
        annual = CompanyDocument("annual_report", [PageText(99, (
            "Consolidated Statement of Comprehensive Income\n"
            "2025 2024\nRMB'million RMB'million\n"
            "Revenue 5,000 4,800\n"
        ))], "https://a", "data/raw/A/2025/annual_report.pdf")
        esg = self._esg_doc(
            "Indicator Unit 2023 2024 2025\n"
            "Total GHG Emissions (Scope 1 + Scope 2) tCO₂e 1,618,370 1,594,055 1,498,435\n"
            "Total water consumption tonnes 800,000 790,000 750,000\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        values = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(1498435 * 1e3 * 1e4 / 5e9, values["Q_E_GHG_INTENSITY"])
        self.assertAlmostEqual(750000 * 1e3 * 1e4 / 5e9, values["Q_E_WATER_INTENSITY"])
        # 无年份表头时拒绝猜列
        no_header = self._esg_doc("Total GHG Emissions (Scope 1 + Scope 2) tCO₂e 1,618,370 1,594,055 1,498,435\n")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [annual, no_header]))

    def test_env_intensity_english_four_year_postfix_unit_and_rmb_unit_statement(self):
        # 四年升序表、行尾单位及“Expressed in RMB”原币单位（真实样例：01713.HK）
        annual = CompanyDocument("annual_report", [PageText(83, (
            "Consolidated Statement of Profit or Loss\n2025\n(Expressed in RMB)\nNotes 2025 2024\n"
            "I. Operating revenue V.33 4,968,130,028.88 4,775,571,527.88\n"
        ))], "url", "annual.pdf")
        esg = self._esg_doc(
            "Emissions Year 2025 Year 2024 Year 2023 Year 2022 Unit\nAir Pollutant\n"
            "Nitrogen oxides 2,384.50 2,210.72 1,486.56 2,223.87 kg\n"
            "Sulphur oxides 8.00 8.78 5.12 4.74 kg\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(2384.5 * 1000 * 1e4 / 4968130028.88, by_code["Q_E_NOX_INTENSITY"])
        # sulphur oxides不自动等同二氧化硫
        self.assertNotIn("Q_E_SO2_INTENSITY", by_code)

    def test_env_intensity_english_group_boiler_actual_so2_nox_statement(self):
        # 报告年、集团主体、化学式和实际质量同时闭环（真实样例：01277.HK）
        annual = CompanyDocument("annual_report", [PageText(117, (
            "Consolidated Statement of Profit or Loss and Other Comprehensive Income\nRMB’000\n"
            "Revenue 5 5,293,266 5,655,829\n"
        ))], "url", "annual.pdf")
        esg = self._esg_doc(
            "In 2025, exhaust gas emission from boilers by the Group was 81.267 million standard cubic meters, "
            "of which, sulfur dioxide (SO2) was 12.59 tonnes, and nitrogen oxide (NOx) was 13.01 tonnes. "
            "The target for exhaust gas emissions in 2026 is to maintain the same level.\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, esg])
        by_code = {item.indicator_code: item.value for item in items}
        self.assertAlmostEqual(12.59e6 * 1e4 / 5.293266e9, by_code["Q_E_SO2_INTENSITY"])
        self.assertAlmostEqual(13.01e6 * 1e4 / 5.293266e9, by_code["Q_E_NOX_INTENSITY"])

    def test_env_intensity_english_boiler_target_or_product_sulfur_rejected(self):
        annual = CompanyDocument("annual_report", [PageText(117, (
            "Consolidated Statement of Profit or Loss\nRMB’000\nRevenue 5 5,293,266 5,655,829\n"
        ))], "url", "annual.pdf")
        for text in (
            "The Group targets sulfur dioxide (SO2) emissions of 12.59 tonnes in 2026.",
            "The Group's clean coal products have sulfur content below 0.6% to reduce sulfur dioxide emissions.",
        ):
            self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [annual, self._esg_doc(text)]))

    def test_env_intensity_ghg_total_scope_closure(self):
        # 范围一+范围二闭环时接受总量
        closing = self._esg_doc(
            "指标 单位 2023 年 2024 年 2025 年\n"
            "温室气体排放总量 吨二氧化碳当量 227,656 221,495 157,996\n"
            "直接温室气体排放量（范围一） 吨二氧化碳当量 3,725.44 3,142.65 3,081.89\n"
            "间接温室气体排放量（范围二） 吨二氧化碳当量 223,930.56 218,353 154,914.11\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), closing])
        self.assertEqual(1, len(items))
        self.assertAlmostEqual(157996 * 1e3 * 1e4 / 1.705e10, items[0].value)
        # 总量含范围三（范围一+范围二不闭环）时拒绝
        scope3 = self._esg_doc(
            "Indicator Unit 2023 2024 2025\n"
            "Scope 1 Greenhouse Gas Emissions tCO2e 58,333.53 109,895.55 137,645.38\n"
            "Scope 2 Greenhouse Gas Emissions tCO2e 367,858.05 488,540.24 511,814.87\n"
            "Scope 3 Greenhouse Gas Emissions tCO2e - 3,614,426.24 4,110,318.06\n"
            "Total Greenhouse Gas Emissions tCO2e 426,191.57 4,212,862.03 4,759,778.31\n"
        )
        annual = CompanyDocument("annual_report", [PageText(99, (
            "Consolidated Statement of Comprehensive Income\n"
            "2025 2024\nRMB'million RMB'million\n"
            "Revenue 5,000 4,800\n"
        ))], "https://a", "data/raw/A/2025/annual_report.pdf")
        self.assertFalse(derive_env_intensity_candidates("A", "甲", 2025, [annual, scope3]))
        # 显式标注范围一+范围二的总量不受同页范围三披露影响
        annotated = self._esg_doc(
            "Indicator Unit 2023 2024 2025\n"
            "Total GHG Emissions (Scope 1 + Scope 2) tCO2e 1,618,370 1,594,055 1,498,435\n"
            "Scope 3 GHG emissions tCO2e - - 12,345\n"
        )
        items = derive_env_intensity_candidates("A", "甲", 2025, [annual, annotated])
        self.assertEqual(1, len(items))

    def test_env_intensity_derived_candidates_auto_confirm(self):
        from aegis_esg.resolution import resolve_pending_candidates
        esg = self._esg_doc("温室气体排放总量\n2.95万吨二氧化碳当量\n")
        items = derive_env_intensity_candidates("A", "甲", 2025, [self._annual_doc(self._CN_REVENUE), esg])
        confirmed, unresolved, _ = resolve_pending_candidates(items)
        self.assertEqual(1, len(confirmed))
        self.assertFalse(unresolved)
        self.assertEqual("confirmed", confirmed[0].status.value)

    def test_merge_confirmed_observations_dedupes_and_rejects_conflicts(self):
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            first = root / "first.csv"
            second = root / "second.csv"
            write_observations(first, [Observation(
                "600900.SH", "长江电力", 2025, "Q_E_GHG_INTENSITY", 1.25,
                ValueStatus.CONFIRMED, source_url="https://a", confidence=1.0,
            )])
            write_observations(second, [
                Observation(
                    "600900.SH", "长江电力", 2025, "Q_E_GHG_INTENSITY", 1.25,
                    ValueStatus.CONFIRMED, source_url="https://a", confidence=1.0,
                ),
                Observation(
                    "600900.SH", "长江电力", 2025, "X_E_ENV_SYSTEM", 80,
                    ValueStatus.CONFIRMED, source_url="https://b", confidence=1.0,
                ),
            ])
            merged, summary = merge_confirmed_observations([first, second], self.methodology)
            self.assertEqual(2, len(merged))
            self.assertEqual(2, summary["merged_observation_count"])
            self.assertFalse(summary["publishable"])
            conflict = root / "conflict.csv"
            write_observations(conflict, [Observation(
                "600900.SH", "长江电力", 2025, "Q_E_GHG_INTENSITY", 2.5,
                ValueStatus.CONFIRMED, source_url="https://c", confidence=1.0,
            )])
            with self.assertRaisesRegex(ValueError, "冲突"):
                merge_confirmed_observations([first, conflict], self.methodology)
            pending = root / "pending.csv"
            write_observations(pending, [Observation(
                "600900.SH", "长江电力", 2025, "Q_E_GHG_INTENSITY", None, ValueStatus.PENDING,
            )])
            with self.assertRaisesRegex(ValueError, "非confirmed"):
                merge_confirmed_observations([pending], self.methodology)


if __name__ == "__main__":
    unittest.main()
